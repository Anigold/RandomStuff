from openpyxl import load_workbook
import xlrd
from bs4 import BeautifulSoup
import pprint
from math import ceil
import shutil

output_file_path 	= './Output Files/'
new_files_path 	 	= './Put New Files in Here/'
template_files_path = './Templates/'

bagel_sort_order = [
	'Honey Wheat',
	'Cinnamon Raisin',
	'Granola',
	'Rosemary Salt',
	'Salt',
	'Garlic',
	'Italian',
	'Zaatar',
	'Plain',
	'Sesame',
	'Poppy',
	'Onion',
	'Long Island Wheat'	
]

bagel_mapping_strings = {
 
	'Bagel Honey Whole Wheat': "HWW",
	'Bagel Cinnamon Raisin': "CR",
	'Bagel Granola': "GRAN",
	'Bagel Rosemary Salt': "ROSE",
	'Bagel Salt': "SALT",
	'Bagel Garlic': "GARLIC",
	'Bagel Spicy Italian': "ITALIAN",
	'Bagel Za`atar': "ZA'ATAR",
	'Bagel Plain': "PLAIN",
	'Bagel Sesame': "SESAME",
	'Bagel Poppy': "POPPY",
	'Bagel Onion': "ONION",
	'Bagel Long Island Wheat': "LIW",
	'Bagel Long Island': "LI",

}

bagel_mapping = {
 
	'070108': "HWW",
	'070104': "CR",
	'070116': "GRAN",
	'070139': "ROSE",
	'070107': "SALT",
	'070106': "GARLIC",
	'070130': "ITALIAN",
	'070143': "ZA'ATAR",
	'070101': "PLAIN",
	'070102': "SESAME",
	'070103': "POPPY",
	'070105': "ONION",
	'070138': "LIW",
	'070110': "LI",

}

stores_mapping = {
	'01CTB': 7,
	'02DTB': 9,
	'03TRIP': 8,
	'04IB': 6,
	'05FRSH': 5,
	'06EHP': 10
}

# Read data from OVEN PRODUCTION
soup = BeautifulSoup(open(new_files_path + 'ZZ Bagel Production - Oven.xls'), 'html.parser')
tables = soup.find('table')
rows = tables.find_all('tr')
data = []
section = False
current_section = ''
for row in rows:
	cols = row.find_all('td')
	cols = [ele.text.strip() for ele in cols]
	
	if cols:
		data.append([ele for ele in cols if ele])

bagel_oven_data = []
for i in data:
	if i[0].split(" ")[0] != 'Category:' and i[0] != 'Totals':
		name = bagel_mapping[i[0]] if i[0] in bagel_mapping else 'Unnamed Bagel'

		default_value = i[3] + i[4]
		default_rounded = 0.5 * ceil(2.0 * float(default_value))

		retail_value = i[5]

		bagel_oven_data.append({
			'Name': name,
			'Default': default_rounded,
			'Retail': retail_value
			})

	elif i[0] == 'Totals':
		break

# Access template
bagel_early_output_workbook = load_workbook(template_files_path + 'Bagel Distribution EARLY.xlsx')
bagel_early_output_sheet = bagel_early_output_workbook.active
for i in range(8, 22):
	bagel_name = bagel_early_output_sheet[f'B{i}'].value
	for j in bagel_oven_data:
		
		if j['Name'] == bagel_name:
			bagel_early_output_sheet.cell(row=i, column=7).value = j['Default']
			bagel_early_output_sheet.cell(row=i, column=8).value = j['Retail']

bagel_early_output_workbook.save(output_file_path + 'Bagel Distribution EARLY.xlsx')

# Read data from STORES PRODUCTION
soup = BeautifulSoup(open(new_files_path + 'ZZ Our Stores Bagel Production.xls'), 'html.parser')
tables = soup.find('table')
rows = tables.find_all('tr')
data = []

# Get headers
headers = [row.find_all('th') for row in rows]
stores = []
i = 0
for header in headers[0]:

	if header.contents[0] != 'Item' and header.contents[0] != 'Totals':
		stores.append({
			'Name': header.contents[0],
			'Bagels': [],
			'Order': i
			})

	++i
for row in rows:
	cols = row.find_all('td')
	cols = [ele.text.strip() for ele in cols]
	
	if cols:
		data.append([ele for ele in cols if ele])
		
for i in data:

	if i[0].split(" ")[0] != 'Route:' and i[0] != 'Totals:':
		name = bagel_mapping_strings[i[0]]
		for pos, val in enumerate(i[2:]):
			stores[pos]['Bagels'].append({
				"Name": name,
				"Quantity": val
				})



# Access template
bagel_early_output_workbook = load_workbook(template_files_path + 'Bagel Distribution RTL.xlsx')
bagel_early_output_sheet = bagel_early_output_workbook.active
for i in range(8, 22):
	bagel_name = bagel_early_output_sheet[f'B{i}'].value
	for store in stores:

		for bagel in store['Bagels']:
			if bagel['Name'] == bagel_name:
				
				bagel_early_output_sheet.cell(row=i, column=stores_mapping[store['Name']]).value = float(bagel['Quantity'])

bagel_early_output_workbook.save(output_file_path + 'Bagel Distribution RTL.xlsx')

with open(output_file_path + 'Hello, Bakers!.txt', 'w+') as hello_bakers_file:
	message = '''
	If there are any problems or miscalculations with the output files, please leave me a note so that I may correct them.
	- Andrew
	'''
	hello_bakers_file.write(message)
	hello_bakers_file.close()

if __name__ == '__main__':
	pass
