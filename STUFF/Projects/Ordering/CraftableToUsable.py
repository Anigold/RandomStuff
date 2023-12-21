import pprint
from PyPDF2 import PdfReader
from os import listdir, remove
from os.path import isfile, join
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from io import StringIO
from pdfminer.high_level import extract_text_to_fp
from pdfminer.layout import LAParams

base_path = 'C:/Users/Will/Desktop/Andrew/Projects/Ordering/'

def extract_table_column_data(html: str, column_data: dict) -> dict:

	output_data = {}
	soup 		= BeautifulSoup(html, 'html.parser')

	for key in column_data:

		column_info = [div.extract() for div in soup.select(f'[style*="{column_data[key]["styles"]}"]')]

		if key not in output_data: output_data[key] = column_info

	return output_data

def convert_to_html(path: str, string_buffer: StringIO) -> None:
	with open(path, 'rb') as pdf_file:
		extract_text_to_fp(pdf_file, string_buffer, laparams=LAParams(), output_type='html', codec=None)
	return

def retrieve_item_ordering_information(column_info: dict) -> dict:

	skus 	   = []
	quantities = []
	for sku in column_info['item_skus']:
		
		sku_spans = sku.find_all('span')
		sku 	  = ''
		for j in sku_spans:
			if 'SKU: ' in j.text: sku = j.text.replace('SKU: ', '').strip().replace('\n', '')

		if sku: skus.append(sku)

	for quantity in column_info['item_quantities']:
		if quantity.text.strip().isnumeric(): quantities.append(quantity.text.strip())

	return {skus[i]: quantities[i] for i in range(len(skus))}


	
def get_files(path: str) -> list:
	return [file for file in listdir(path) if isfile(join(path, file))]

if __name__ == '__main__':

	order_sheets_path = join(base_path, 'OrderSheets')
	order_sheets 	  = get_files(order_sheets_path)
	
	for order_sheet in order_sheets:

		output = StringIO()
		
		# Extract PDF text to string builder
		convert_to_html(join(order_sheets_path, order_sheet), output)

		# Convert PDF text to HTML page
		with open(f'temp{order_sheet.split(".")[0]}.html', 'w') as html_file:
			html_file.write(output.getvalue())

		# Parse HTML page
		with open(f'temp{order_sheet.split(".")[0]}.html') as fp:
			column_info = extract_table_column_data(fp, {'item_skus': {'styles': 'left:52px'}, 'item_quantities': {'styles': 'left:352px'}})
			
		items = retrieve_item_ordering_information(column_info)
		pprint.pprint(items)
		workbook = Workbook()
		sheet = workbook.active

		for pos, sku in enumerate(items):
			quantity = items[sku]

			sheet.cell(row=pos+1, column=1).value = sku
			sheet.cell(row=pos+1, column=2).value = quantity

		workbook.save(filename=f'{order_sheet.split(".")[0]}_order_list.xlsx')

		remove(join(base_path, f'temp{order_sheet.split(".")[0]}.html'))
