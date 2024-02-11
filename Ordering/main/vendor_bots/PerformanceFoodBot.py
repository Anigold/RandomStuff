from .VendorBot import VendorBot
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium import webdriver
from openpyxl import Workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class PerformanceFoodBot(VendorBot):

    def __init__(self, driver: webdriver, username, password) -> None:
        super().__init__()
        self.name = "Performance Food"
        self.driver = driver
        self.username = username
        self.password = password

        self.store_ids = {}

    def login(self) -> None:
        pass

    def logout(self) -> None:
        pass
    
    def switch_store(self, store_id: str) -> None:
        pass

    def format_for_file_upload(self, item_data: dict, path_to_save: str) -> None:
        
        workbook = Workbook()
        sheet = workbook.active

        for pos, sku in enumerate(item_data):
            quantity = item_data[sku]

            sheet.cell(row=pos+1, column=1).value = int(sku)
            sheet.cell(row=pos+1, column=2).value = int(quantity)
            sheet.cell(row=pos+1, column=3).value = 'CS'
        
        workbook.save(filename=f'{path_to_save}.xlsx')

        return
    
    def retrieve_pricing_sheet(self, guide_name: str) -> None:
       pass

    def end_session(self) -> None:
        self.driver.close()
        return