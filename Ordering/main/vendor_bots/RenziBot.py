from .VendorBot import VendorBot
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium import webdriver
from openpyxl import Workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class RenziBot(VendorBot):

    def __init__(self, driver: webdriver, username, password) -> None:
        super().__init__()
        self.name = "Renzi"
        self.driver = driver
        self.username = username
        self.password = password

        self.store_ids = {
            'Bakery': "11104",
            'Collegetown': "11106",
            'Triphammer': "11105",
            'Easthill': "11108",
            'Downtown': "11107"
        }

    def login(self) -> None:
	
        self.driver.get('https://connect.renzifoodservice.com/pnet/eOrder')
        
        username_input = self.driver.find_element(By.NAME, 'UserName')
        password_input = self.driver.find_element(By.NAME, 'Password')

        username_input.send_keys(self.username)
        password_input.send_keys(self.password)

        submit_button = self.driver.find_element(By.NAME, 'SubmitBtn')
        submit_button.click()

        try:
            WebDriverWait(self.driver, 10).until(EC.alert_is_present())
            self.driver.switch_to.alert.accept()
        except:
            pass

        time.sleep(5)

        return

    def logout(self) -> None:
        logout_button = self.driver.find_element(By.XPATH, '//span[@title="Sign Off"]')
        logout_button.click()
        time.sleep(2)

        try:
            WebDriverWait(self.driver, 10).until(EC.alert_is_present())
            self.driver.switch_to.alert.accept()
        except:
            pass

        try:
            WebDriverWait(self.driver, 10).until(EC.alert_is_present())
            self.driver.switch_to.alert.accept()
        except:
            pass
        
        return
    
    def switch_store(self, store_id: str) -> None:

        store_dropdown = self.driver.find_element(By.NAME, 'selectedCustomer')
        store_dropdown.click()

        time.sleep(3)

        store_id = f'  1,  1,  1,{store_id}'
        store_option = Select(store_dropdown.find_element(By.XPATH, f'.//option[value="{store_id}"]'))
        store_option.select_by_value(store_id)

        time.sleep(3)

        return

    def format_for_file_upload(self, item_data: dict, path_to_save: str) -> None:
        # CSV-style Excel file with "Item Code, Quantity, and Broken Case"
        workbook = Workbook()
        sheet = workbook.active

        for pos, sku in enumerate(item_data):
            quantity = item_data[sku]

            sheet.cell(row=pos+1, column=1).value = int(sku)
            sheet.cell(row=pos+1, column=2).value = int(quantity)
            sheet.cell(row=pos+1, column=3).value = int(0)
        
        workbook.save(filename=f'{path_to_save}.xlsx')

        return
    
    def retrieve_pricing_sheet(self, guide_name: str) -> None:
        self.login()

        create_order_button = self.driver.find_element(By.ID, 'mainmenu-order')
        create_order_button.click()
        time.sleep(10)

        WebDriverWait(self.driver, 10).until(EC.frame_to_be_available_and_switch_to_it(self.driver.find_element(By.XPATH, '//iframe[@class="bigrounded"]')))
        WebDriverWait(self.driver, 10).until(EC.frame_to_be_available_and_switch_to_it(self.driver.find_element(By.XPATH, '//iframe[@id="Custom-iFrame0"]')))
        custom_guide_select = Select(self.driver.find_element(By.NAME, 'previousOrders'))
        custom_guide_select.select_by_visible_text(guide_name)
        time.sleep(5)

        submit_button = self.driver.find_element(By.ID, 'startOrderButton')
        submit_button.click()
        self.driver.switch_to.default_content()
        
        
        WebDriverWait(self.driver, 10).until(EC.frame_to_be_available_and_switch_to_it(self.driver.find_element(By.XPATH, '//iframe[@id="ContentFrame"]')))
        time.sleep(10)
        advanced_dropdown = self.driver.find_element(By.ID, 'pagemenuli-adv')
        advanced_dropdown.click()
        time.sleep(5)

        export_dropdown = self.driver.find_element(By.ID, 'pagemenuli-exp')
        export_dropdown.click()
        time.sleep(5)

        standard_dropdown = self.driver.find_element(By.ID, 'pagemenuli-standExp')
        standard_dropdown.click()
        time.sleep(5)

        excel_button = self.driver.find_element(By.ID, 'pagemenu-exp2')
        excel_button.click()
        time.sleep(5)

        self.logout()

        return

    def end_session(self) -> None:
        self.driver.close()
        return