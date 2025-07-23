from backend.models.order import Order
from WorkBot.refactor.backend.models.order_item import OrderItem
from openpyxl import load_workbook, Workbook
from pathlib import Path

class OrderParser:

    def parse_excel(self, file_path: Path) -> Order:
        stem = file_path.stem
        vendor, store, date = stem.split("_")

        order = Order(store=store, vendor=vendor, date=date)

        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            sku, name, quantity, cost_per, total_cost = row
            item = OrderItem(sku, name, quantity, cost_per, total_cost)
            order.items.append(item)
        
        wb.close()
        return order

    def to_excel_workbook(self, order: Order) -> 'Workbook':
        wb = Workbook()
        ws = wb.active
        ws.title = "Order"

        headers = ["SKU", "Name", "Quantity", "Cost Per", "Total Cost"]
        ws.append(headers)

        for item in order.items:
            ws.append([item.sku, item.name, item.quantity, item.cost_per, item.total_cost])
        
        return wb