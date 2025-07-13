from models.Order import Order, OrderItem
from openpyxl import load_workbook
from pathlib import Path

class OrderParser:
    def parse_excel(self, file_path: Path) -> Order:
        stem = file_path.stem
        vendor, store, date = stem.split("_")
        order = Order(store=store, vendor=vendor, date=date)

        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            sku, name, quantity, cost_per, total_cost = row
            item = OrderItem(sku, name, quantity, cost_per, total_cost)
            order.items.append(item)
        
        return order
