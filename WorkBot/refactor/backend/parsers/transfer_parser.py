from backend.models.transfer import Transfer
from backend.models.transfer_item import TransferItem
from openpyxl import load_workbook, Workbook
from pathlib import Path

class TransferParser:

    def parse_excel(self, file_path: Path) -> Transfer:
        """
        Parse a Transfer Excel file into a Transfer object.

        Filename format: <store_from>_<store_to>_<date>.xlsx
        """
        stem = file_path.stem
        store_from, store_to, date = stem.split("_")

        transfer = Transfer(store_from=store_from, store_to=store_to, date=date)

        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, quantity = row
            item = TransferItem(name=name, quantity=quantity)
            transfer.transfer_items.append(item)

        wb.close()
        return transfer

    def to_excel_workbook(self, transfer: Transfer) -> Workbook:
        """
        Serialize a Transfer object to an Excel Workbook.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Transfer"

        headers = ["Name", "Quantity"]
        ws.append(headers)

        for item in transfer.transfer_items:
            ws.append([item.name, item.quantity])

        return wb
