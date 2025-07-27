# import backend.config as config

from backend.utils.logger import Logger

''' MANAGERS '''
from backend.bots.craftable_bot.craftable_bot import CraftableBot
from backend.coordinators.vendor_coordinator import VendorCoordinator
from backend.coordinators.store_coordinator import StoreCoordinator
from backend.coordinators.order_coordinator import OrderCoordinator
from backend.coordinators.transfer_coordinator import TransferCoordinator

''' OBJECTS '''
from backend.models.order import Order
from backend.models.transfer import Transfer
from backend.models.transfer_item import TransferItem

from backend.utils.emailer.emailer import Emailer, Email
from backend.utils.emailer.services.gmail_service import GmailService
from backend.utils.emailer.services.outlook_service import OutlookService

# from backend.transferring.Transfer import Transfer, TransferItem
# from backend.emailer.Emailer import Emailer, Email
# from backend.emailer.Services.Gmail import GmailService
# from backend.emailer.Services.Outlook import OutlookService

'''  HELPERS '''
from backend.bots.craftable_bot.helpers import get_craftable_username_password
from backend.utils.helpers import string_to_datetime

from backend.bots.bot_mixins import SeleniumBotMixin

''' STANDARD LIBRARY '''
from datetime import datetime
import socket
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

@Logger.attach_logger
class WorkBot:
    
    def __init__(self):
        self.logger.info('Initializing WorkBot...')

        self.vendor_coordinator   = VendorCoordinator()
        self.order_coordinator    = OrderCoordinator()
        self.store_coordinator    = StoreCoordinator()
        self.transfer_coordinator = TransferCoordinator()

        username, password = get_craftable_username_password()

        self.craft_bot = CraftableBot(username, password,
                                    order_coordinator=self.order_coordinator,
                                    transfer_coordinator=self.transfer_coordinator)

        self.logger.info('WorkBot initialized successfully.')
               # Establish host service
        
        # Establish emailer service based on which computer we're one.
        # We will move this to config later.
        email_host = socket.gethostname()
        if email_host == 'Purchasing':
            self.emailer = Emailer(OutlookService())
        else:
            self.emailer = None  

        self.logger.info('WorkBot initialized successfully.')

    def download_orders(self, stores, vendors=[], download_pdf=True, update=True):
        self.craft_bot.download_orders(stores, vendors, download_pdf=download_pdf, update=update)
    
    def sort_orders(self):
        self.order_coordinator.sort_orders()

    def delete_orders(self, stores, vendors=[]):
        self.craft_bot.delete_orders(stores, vendors)

    def input_transfers(self):
        transfer_file_dir = self.transfer_coordinator.get_transfer_files_directory()
         
        transfers = [TransferCoordinator.create_transfer_from_excel(transfer_file) for transfer_file in transfer_file_dir.iterdir() if transfer_file.is_file()]

        self.craft_bot.input_transfers(transfers)

    def convert_order_to_transfer(self, order_store, vendor, store_to):
        '''This is so niche that we only need to use it once a week, but it removes tedium from my life so it stays.'''
        # print(order.items, flush=True)

        order = self.get_orders(order_store, vendor)[0]
        if order.vendor == 'Ithaca Bakery': 
            store_from = 'Bakery'
        else:
            store_from = order.vendor

        transfer_items = []
        for item in order.items:
            transfer_item = TransferItem(name=item.name, quantity=item.quantity)
            transfer_items.append(transfer_item)

        order_datetime = string_to_datetime(order.date)

        transfer = Transfer(
            store_from=store_from,
            store_to=order.store,
            date=order_datetime,
            items=transfer_items
            )
        
        return self.transfer_coordinator.save_transfer(transfer=transfer)
        
    def generate_vendor_upload_files(
        self,
        stores: list[str],
        vendors: list[str],
        start_date: str = None,
        end_date: str = None
    ) -> list[Path]:
        """
        Entrypoint for generating vendor upload files for the given filters.
        Delegates to the OrderCoordinator.
        """
        order_file_paths = self.order_coordinator.get_orders_from_file(
            stores=stores,
            vendors=vendors
        )

       

        context_map = {}
        for file_path in order_file_paths:
            order = self.order_coordinator.read_order_file(file_path)
            vendor_info = self.vendor_coordinator.get_vendor_information(order.vendor)

            context_map[str(file_path)] = {
                "store": order.store,
                "vendor_info": vendor_info,
                "date_str": order.date,
            }

        return self.order_coordinator.generate_vendor_upload_files(
            stores=stores,
            vendors=vendors,
            start_date=start_date,
            end_date=end_date,
            context_map=context_map
        )

    def close_craftable_session(self):
        self.craft_bot.close_session()

    def welcome_to_work(self) -> None:
        
        today = self._get_today_date_and_day()

        return f'''\n{today[1]}, {today[0]}'''

    def get_orders(self, stores: list, vendors: list = []) -> list:
        return self.order_coordinator.get_orders_from_file(stores=stores, vendors=vendors)

    def get_vendor_information(self, vendor_name: str) -> dict:
        return self.vendor_coordinator.get_vendor_information(vendor_name)

    def get_store_information(self, store_name: str) -> dict:
        return self.store_coordinator.find_store_by_name(store_name=store_name)
    
    def combine_orders(self, vendors: list) -> None:
        self.order_coordinator.combine_orders(vendors)
  

    def generate_vendor_order_emails(self, vendors: list[str], stores: list[str] = []) -> list[Email]:
        orders = self.get_orders(stores=stores, vendors=vendors)
        if not orders:
            return []

        grouped_orders = self._group_orders_by_vendor(orders)
        emails = []

        for vendor_name, vendor_orders in grouped_orders.items():
            to_emails   = self._get_vendor_recipients(vendor_name)
            subject     = self._build_email_subject(vendor_name)
            body        = self._build_vendor_email_body(vendor_name, vendor_orders)
            attachments = self._get_email_attachments(vendor_orders)

            email = Email(
                to=tuple(to_emails),
                subject=subject,
                body=body,
                attachments=tuple(attachments) if attachments else None
            )

            self.emailer.create_email(email)
            self.emailer.display_email(email)
            emails.append(email)

        return emails

    def _group_orders_by_vendor(self, orders: list) -> dict:
        grouped = defaultdict(list)
        for order in orders:
            order_obj = self.order_coordinator.read_order_file(order)
            grouped[order_obj.vendor].append(order)
        return grouped

    def _build_email_subject(self, vendor_name: str) -> str:
        date_str = datetime.now().strftime("%B %d, %Y")
        return f"Orders for {vendor_name} ({date_str})"

    def _get_vendor_recipients(self, vendor_name: str) -> list[str]:
        try:
            vendor_info = self.vendor_coordinator.get_vendor_info(vendor_name)
            if vendor_info.ordering.email:
                return [vendor_info.ordering.email]
        except Exception:
            pass
        return ["default@vendor.com"]

    def _build_vendor_email_body(self, vendor_name: str, orders: list) -> str:
        date_str = datetime.now().strftime("%B %d, %Y")
        lines = [
            f"Hello {vendor_name},",
            "",
            f"Please find below our orders for {date_str}:"
        ]

        for order in orders:
            order_obj = self.order_coordinator.read_order_file(order)
            lines.append(f"\nStore: {order_obj.store}")
            for item in order_obj.items:
                lines.append(f"  - {item.quantity} x {item.name}")

        lines += [
            "", "Let us know if there are any issues.",
            "", "Thank you!", 
            '---',
            "Andrew Goldsmith", 
            'Purchasing Manager', 
            'Ithaca Bakery', 
            '(607) 273-7110'
        ]
        return "\n".join(lines)

    def _get_email_attachments(self, orders: list) -> list[str]:
        attachments = []
        for order in orders:
            order_obj = self.order_coordinator.read_order_file(order)
            pdf_file_path = self.order_coordinator.get_order_file_path(order_obj, format='pdf')

            if pdf_file_path.exists():
                attachments.append(str(pdf_file_path))

        return attachments


    def generate_store_order_emails(self, stores: list[str]):
        orders = self.get_orders(stores=stores)
        if not orders: return None

        emails = []
        store_orders_table = {}
        for order in orders:
            if order.store not in store_orders_table:
                store_orders_table[order.store] = [order]
            else:
                store_orders_table[order.store].append(order)
        

        for store in store_orders_table:
            to_emails = []
            store_contacts = self.get_store_information(store).contacts
            
            for contact in store_contacts:
                if contact['title'] == 'Inventory Clerk': to_emails.append(contact['email'])

            subject = f'Orders for the Week: {store}'
            attachments = []

            for store_order in store_orders_table[store]:
                pdf_dir      = self.order_coordinator.get_vendor_orders_directory(store_order.vendor)
                pdf_filename = f"{self.order_coordinator.generate_filename(store_order, file_extension='.pdf')}"
                pdf_path     = pdf_dir / pdf_filename

                if pdf_path.exists(): attachments.append(str(pdf_path))

            body = 'Hello, here are the orders you have placed for the week.'
            
            # Need to abstract out the CC of operations director
            cc = ('jennithacabakery@gmail.com')
            
            email = Email(
                to=tuple(to_emails),
                subject=subject,
                body=body,
                cc=cc,
                attachments=tuple(attachments) if attachments else None
            )       
           
            self.emailer.create_email(email)
            emails.append(email) 
          
        for email in emails: self.emailer.display_email(email)

    def shutdown(self) -> None:
        """Exits the CLI loop."""
        self.close_craftable_session()
        print("Exiting WorkBot CLI.")
        # exit()

    def _get_today_date_and_day(self):
        today = datetime.today()
        day_of_week = today.strftime("%A")
        day = today.day
        suffix = "th" if 11 <= day <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
        long_date = today.strftime(f"%B {day}{suffix}, %Y")
        return long_date, day_of_week
    
    def archive_all_current_orders(self, stores: list[str] = None, vendors: list[str] = None) -> None:
        """
        Archives all current (non-completed) orders.
        If no vendors are supplied, all vendor directories are searched.
        """
        vendors = vendors or []  # Treat None as empty list (i.e., all vendors)

        # Get current (non-archived) order files
        order_files = self.order_coordinator.get_orders_from_file(
            stores=stores or [],
            vendors=vendors
        )

        for file_path in order_files:
            try:
                order = self.order_coordinator.read_order_file(file_path)
                self.order_coordinator(order)
            except Exception as e:
                self.logger.warning(f"[Archive] Skipped {file_path.name}: {e}")



    def split_natalies(self) -> None:
        natalies_excel_path = Path('C:/Users/Will/Desktop/Natalies.xlsx')
        natalie_flavors = []

        workbook = load_workbook(natalies_excel_path)
        sheet = workbook.active
        for row in sheet.iter_rows():
            if row[0] not in ["", None, ' '] and row[0] is not None:
                natalie_flavors.append(row[0].value)
        flavors = [i for i in natalie_flavors if i is not None]

        perf_orders = self.order_coordinator.get_orders_from_file(vendors=['Performance Food'])
        flf_orders = self.order_coordinator.get_orders_from_file(vendors=['FingerLakes Farms'])

        orders = perf_orders + flf_orders
        
        store_indices = {
            'Collegetown': 2,
            'Downtown': 5,
            'Easthill': 8,
            'Triphammer': 11,
            'Bakery': 14
        }

        for order_path in orders:
            order = self.order_coordinator.parse_order_file(order_path)
            print(order)
            for item in order.items:
                if 'Natalie' in item.name and item.name.split(' - ')[1] in flavors:
                    flavor = item.name.split(' - ')[1]
                    sheet.cell(row=flavors.index(flavor)+4, column=store_indices[order.store]).value = item.quantity

        workbook.save('C:/Users/Will/Desktop/Natalies.xlsx')