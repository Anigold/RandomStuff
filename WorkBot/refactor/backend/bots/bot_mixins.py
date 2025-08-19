from __future__ import annotations
from abc import ABC, abstractmethod
from openpyxl import load_workbook, Workbook
import re
import time
from contextlib import contextmanager

class SeleniumBotMixin(ABC):

    def __init__(self, downloads_path: str = None, username: str = None, password: str = None) -> None:
        self._driver             = None
        self._driver_initialized = False
        self.username            = username
        self.password            = password
        self.downloads_path      = downloads_path
        self.is_logged_in        = False

        # Track nested usage so inner calls don’t kill the browser
        self._session_depth = 0

    def start_session(self) -> None:
        """Create a WebDriver if one isn't already running."""
        if not self._driver_initialized or self._driver is None:
            from backend.bots.helpers import create_driver, create_options
            options            = create_options(self.downloads_path)
            self._driver       = create_driver(options)
            self._driver_initialized = True

    def end_session(self) -> None:
        """
        Hard close of the browser process when a session truly ends.
        Safe to call multiple times.
        """
        try:
            if self._driver_initialized and self._driver:
                self._driver.quit()
        finally:
            # Reset everything so a fresh session can be started later
            self._driver = None
            self._driver_initialized = False
            self.is_logged_in = False

    @contextmanager
    def session(self, login: bool = False):
        """
        Nesting-safe context:
          - Starts the browser on first entry
          - (Optionally) logs in once
          - Tears down only when the outermost context exits
        """
        created_here = False
        try:
            if self._session_depth == 0:
                # spin up a new browser
                self.start_session()
                created_here = True
                if login and not self.is_logged_in:
                    self.login()
            self._session_depth += 1
            yield self
        finally:
            # unwind depth and close only when outermost finishes
            self._session_depth -= 1
            if created_here and self._session_depth == 0:
                self.end_session()
                # slight pause can prevent race conditions on some OS/driver combos
                time.sleep(1.5)

    @staticmethod
    def with_session(login: bool = False):
        """
        Decorator to auto-wrap a method inside `self.session(login=...)`.
        Usage:
            @SeleniumBotMixin.with_session(login=True)
            def do_stuff(self): ...
        """
        def deco(func):
            def wrapper(self, *args, **kwargs):
                with self.session(login=login):
                    return func(self, *args, **kwargs)
            return wrapper
        return deco

    def __enter__(self):
        # no auto-login here; pick per-call with session(login=True) when needed
        return self.session().__enter__()

    def __exit__(self, exc_type, exc, tb):
        return self.session().__exit__(exc_type, exc, tb)

    @property
    def driver(self):
        """
        Lazy accessor. If you prefer to *force* session usage, you could
        assert self._session_depth > 0 here. For backward-compatibility,
        we keep lazy startup.
        """
        if not self._driver_initialized or self._driver is None:
            # Start a session if someone touches driver directly.
            # Strongly prefer using `with self.session():` or the decorator.
            self.start_session()
        return self._driver
    
    def login(self) -> None: 
        ...

    def logout(self) -> None:
        ...

    # def end_session(self) -> None:
    #     if self._driver_initialized and self._driver:
    #         self._driver.quit()
    #         self._driver_initialized = False
    #     return self.driver.close()
    

class PricingBotMixin(ABC):

    UNIT_NORMALIZATION = {
        '#': 'LB',
        'LB': 'LB',
        'LBS': 'LB',
        'OZ': 'OZ',
        'GAL': 'GAL',
        'GALLON': 'GAL',
        'EA': 'EA',
        'EACH': 'EA',
        'CT': 'EA',
        'CNT': 'EA',
        'DOZ': 'DZ',
        'PT': 'PT',
        'QT': 'QT',
        'HD': 'EA',
        'HEAD': 'EA',
        'CS': 'CS',
        'CASE': 'CS',
    }

    PACK_SIZE_PATTERN = re.compile(
        r'''
        ^(?:                                    # Start match
            (?P<packs>\d+)[/x]\s*               # Pack multiplier (e.g. '4/', '2x') (optional)
        )?
        (?P<low>(?:\d+)?(?:\.\d+)?|\d+)         # Main quantity (to the tenth of a unit, '.1))
        (?:-(?P<high>(?:\d+)?(?:\.\d+)?|\d+))?  # Optional range (to the tenth of a unit, '.1')
        \s*
        (?P<unit>[a-zA-Z#]+)                    # Unit of measure (e.g. 'LB', 'EA', 'QT')
        ''',
        re.VERBOSE,
    )

    def __init__(self) -> None:
        self.special_cases = {}
    
    def get_pricing_info_from_sheet(self, path_to_pricing_sheet: str) -> dict:
        pass
    
    def get_skus_from_file(self, path_to_skus: str) -> dict:
        pass

    def normalize_units(unit: str) -> str:

        if not unit: return unit

        unit = unit.upper()
        units = {
			'#': 'LB',
			'LB': 'LB',
			'CNT': 'EA',
			'EA': 'EA',
			'CT': 'EA',
			'DOZ': 'DZ',
			'PINT': 'PT',
            'HD': 'EA',
            'HEAD': 'EA'
		}

        return units[unit] if unit in units else unit
    
    def helper_iter_rows(sheet):
        for row in sheet.iter_rows():
            yield [cell.value for cell in row]

    def get_skus_from_file(self, path_to_skus) -> dict:
        workbook = load_workbook(path_to_skus)
        sheet = workbook.active

        skus = {}
        for pos, row in enumerate(sheet.iter_rows()):
            if pos < 2: continue

            sku = str(row[self.skus_col].value)

            if not sku: continue

            if sku not in skus:
                skus[sku] = pos+1
            
        return skus
    
    def format_vendor_pricing_sheet(self, pricing_sheet_path: str, save_path: str) -> None:
        price_info = self.get_pricing_info_from_sheet(pricing_sheet_path)

        # print('creating workbook',  flush=True)
        workbook = Workbook()
        sheet = workbook.active

        # print('importing pricing data', flush=True)
        for row_pos, item_name in enumerate(price_info):

          
            item_info     = price_info[item_name]
            # print(item_name)
            # pprint(item_info)
            sku           = item_info['SKU']
            cost          = item_info['cost']
            case_size     = item_info['case_size']
            quantity 	  = float(item_info['quantity'])
            unit 	 	  = item_info['units']
            cost_per_unit = f'{round(cost/quantity, 2)} per {unit}'
			
            row = [item_name, sku, cost_per_unit, cost, case_size]
            for col_position, col_value in enumerate(row):
                sheet.cell(row=row_pos+1, column=col_position+1).value = col_value

        # print('Saving workbook', flush=True)
        # print(save_path, flush=True)
        return workbook.save(filename=f'{save_path}')
    
    @classmethod
    def helper_format_size_units(cls, pack_size_str):


        # Expected pack strings:

        #   pack/size unit (12/6 oz)
        #   low-high unit (5-6 EA)
        #   size unit (5 LB)
        #

        if not pack_size_str:
            return None, None
        
        # Ensure space exists between number and unit (e.g., '10Gal' → '10 Gal')
        pack_size_str = re.sub(r'(?<=\d)(?=[A-Za-z#])', ' ', pack_size_str)

        match = cls.PACK_SIZE_PATTERN.search(pack_size_str)
        if not match:
            return None, None  # Return None if no valid match is found

        packs = int(match.group('packs')) if match.group('packs') else 1
        low   = float(match.group('low')) if match.group('low') else 1
        high  = float(match.group('high')) if match.group('high') else low
        unit  = match.group('unit').upper()

        # Normalize the unit
        unit = cls.UNIT_NORMALIZATION.get(unit, unit)
        
        # Compute final quantity (average if range is given)
        avg_size = (low + high) / 2
        total_quantity = avg_size * packs

        # print(f'{pack_size_str} -> ({total_quantity}, {unit})')
        return total_quantity, unit

    # def helper_format_size_units(pack: str, size: str) -> list:
	# 	# Check that pack size is float compatible
		
	# 	# if not isinstance(pack, float):
	# 	# 	try:
	# 	# 		pack = float(pack)
	# 	# 	except:
	# 	# 		raise Exception('Pack size is unable to be coerced to float.')

	# 	# Split value and units from size
    #     unit_string = ''
    #     size_string = ''
		
    #     on_number = False
    #     numeric_non_numerics = ['-', '.']
    #     for pos, char in enumerate(size):
			
    #         if char == ' ':
    #             continue

    #         if not on_number and char.isnumeric():
    #             on_number = True
    #             size_string += char
    #             continue

    #         if not on_number and not char.isnumeric():
    #             if size_string == '': size_string = '1'
    #             unit_string += char
    #             continue

    #         if on_number and char.isnumeric():
    #             size_string += char
    #             continue

    #         if on_number and not char.isnumeric():
                
				
    #             if char in numeric_non_numerics:
    #                 size_string += char
    #                 continue

    #             if char == '#':
    #                 unit_string = char
    #                 break

    #             on_number = False
    #             unit_string += char
    #             continue

    #     if '-' in size_string:
    #         lower, upper = size_string.split('-')
    #         size_string  = (float(lower) + float(upper))/2
    #     else:
    #         size_string = float(size_string)

    #     return [size_string * float(pack), unit_string]

    @abstractmethod
    def retrieve_pricing_sheet(self) -> None:
        pass

    def _special_case_info(self, unit: str, pack: float) -> dict:
        return {'unit': unit, 'pack': pack}
    
class OTPMixin(ABC):
    def __init__(self, retrieve_otp_callback):
        self.retrieve_otp_callback = retrieve_otp_callback