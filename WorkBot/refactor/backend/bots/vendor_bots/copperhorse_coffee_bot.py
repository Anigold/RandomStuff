from .vendor_bot import VendorBot
import os
from pprint import pprint
from openpyxl import Workbook, load_workbook

class CopperHorseBot(VendorBot):

    def __init__(self) -> None:
        super().__init__()
        
        self.name                 = 'Copper Horse Coffee'
        self.minimum_order_amount = 500_00 # $500 in cents


    def format_for_file_upload(self, item_data: dict, path_to_save: str) -> None:
        # CSV-style Excel file with "Item Code, Quantity"
        workbook = Workbook()
        sheet = workbook.active
         
        for pos, name in enumerate(item_data):
            quantity = item_data[name]['quantity']
            sheet.cell(row=pos+1, column=1).value = True
            sheet.cell(row=pos+1, column=2).value = name
            sheet.cell(row=pos+1, column=3).value = int(quantity)
        
        workbook.save(filename=f'{path_to_save}.xlsx')

    '''
    Takes the pre-formatted orders and merges them to one sheet.

    Args:
        path_to_orders: string representation of the path to the folder containing all the orders
    '''
    def combine_orders(self, path_to_orders: list, path_to_save: str) -> None:
        
        combined_orders = {}

        for order_file in path_to_orders:
            order_book = load_workbook(order_file)
            order_sheet = order_book.active

            for row in order_sheet.iter_rows(min_row=2):
                name = row[1].value.split('\n')[0]
                quantity = int(row[2].value)

                if name not in combined_orders:
                    combined_orders[name] = quantity
                else:
                    combined_orders[name] += quantity

        combined_orders_book  = Workbook()
        combined_orders_sheet = combined_orders_book.active

        for pos, item in enumerate(combined_orders):
            combined_orders_sheet.cell(row=pos+1, column=1).value = item
            combined_orders_sheet.cell(row=pos+1, column=2).value = combined_orders[item]

        combined_orders_book.save(f'{path_to_save}\\combined_order.xlsx')

        return