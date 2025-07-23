from .vendor_bot import VendorBot, SeleniumBotMixin, PricingBotMixin, OTPMixin
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium import webdriver
from openpyxl import Workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from csv import writer, reader
from pprint import pprint
from backend.emailer.Emailer import Emailer
from backend.emailer.Services.Outlook import OutlookService
import re

class USFoodsOTPProvider:

    def __init__(self, email_service):
        self.email_service = email_service or OutlookService()

    def __call__(self):
        print('Expecting OTP Email...refreshing inbox.', flush=True)
        self.email_service.refresh_inbox()
        time.sleep(20)
        # message = self.email_service.get_recent_messages(
        #     subject_filter='Your US Foods one time passcode',
        #     max_age_minutes=10
        # )
        inbox = self.email_service.get_inbox()
        messages = inbox.Items
        messages.Sort("[ReceivedTime]", True)
        for pos, i in enumerate(messages):
            if pos > 0: break
            # print(i.Subject)
            # print(i.Body)
            return self.extract_code(i.body)

        # print('Heres the message', flush=True)
        # print(message, flush=True)
        # print(next(message), flush=True)
        # return self.extract_code(message)
    
    def extract_code(self, message_body):
        print('Extracting!', flush=True)
        print(message_body, flush=True)
        match = re.search(r'\b\d{6}\b', message_body)
        return match.group() if match else None

    
class USFoodsBot(VendorBot, SeleniumBotMixin, PricingBotMixin):

    def __init__(self, driver: webdriver = None, username = None, password = None, otp_provider = None) -> None:
        VendorBot.__init__(self)
        SeleniumBotMixin.__init__(self, driver, username, password)
        self.name                 = "US Foods"
        self.minimum_order_amount = 500_00
        self.otp_provider         = otp_provider(None)

        self.store_ids = {
            'BAKERY':      '91602987',
            'COLLEGETOWN': '11602976',
            'TRIPHAMMER':  '1602994',
            'EASTHILL':    '31602998',
            'DOWNTOWN':    '21602990',
            'Bakery':      '91602987',
            'Collegetown': '11602976',
            'Triphammer':  '1602994',
            'Easthill':    '31602998',
            'Downtown':    '21602990',
        }

        self.special_cases = {
		    '88076': {'unit': 'EA', 'pack': 1},
		    '88055': {'unit': 'EA', 'pack': 36},
	    }

        self.site_map = {
            'home_page': 'https://order.usfoods.com/desktop/search/browse'
        }

    def login(self) -> None:
        
        if self.is_logged_in:
            return
        

        self.driver.get(self.site_map['home_page'])
        # self.driver.get('https://connect.renzifoodservice.com/pnet/eOrder')

        # try:
        #     WebDriverWait(self.driver, 10).until(EC.alert_is_present())
        #     self.driver.switch_to.alert.accept()
        # except:
        #     pass
        time.sleep(15)
        login_button_redirect = self.driver.find_element(By.XPATH, '//ion-button[@data-cy="guest-login-header-button"]')
        login_button_redirect.click()

        try:
            WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.ID, 'signInName-facade')))
        except:
            pass

        username_input = self.driver.find_element(By.ID, 'signInName-facade')
        # password_input = self.driver.find_element(By.NAME, 'Password')

        username_input.send_keys(self.username)
        # password_input.send_keys(self.password)

        submit_button = self.driver.find_element(By.ID, 'next')
        submit_button.click()

        try:
            WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, 'Email')))
            send_otp_email_button = self.driver.find_element(By.ID, 'Email')
            send_otp_email_button.click()
            time.sleep(10)
            otp_code = self.otp_provider()
            time.sleep(5)
            print(otp_code, flush=True)
            
            
        except Exception as e:
            print(f'Something fucked up: {e}')

        # opt_code = USFoodsOTPProvider()
        # outlook_service = OutlookService()
        # outlook_service.refresh_inbox()
        # time.sleep(7)
        # for message in outlook_service.get_recent_messages(subject_filter="Your US Foods one time password"):
        #     print(message.Body)
        #     break

        time.sleep(5)

        self.is_logged_in = True
        return


    def logout(self) -> None:
        logout_button = self.driver.find_element(By.XPATH, '//span[@title="Sign Off"]')
        logout_button.click()
        time.sleep(2)

        try:
            WebDriverWait(self.driver, 10).until(EC.alert_is_present())
            self.driver.switch_to.alert.accept()
        except:
            pass

        try:
            WebDriverWait(self.driver, 10).until(EC.alert_is_present())
            self.driver.switch_to.alert.accept()
        except:
            pass
        
        self.is_logged_in = False
        return
    
    def switch_store(self, store_id: str) -> None:

        store_dropdown = self.driver.find_element(By.NAME, 'selectedCustomer')
        store_dropdown.click()

        time.sleep(3)

        store_id = f'  1,  1,  1,{store_id}'
        store_option = Select(store_dropdown.find_element(By.XPATH, f'.//option[value="{store_id}"]'))
        store_option.select_by_value(store_id)

        time.sleep(3)

        return

    def format_for_file_upload(self, item_data: dict, path_to_save: str, store: str) -> None:
        # CSV-style Excel file with "Item Code, Quantity, and Broken Case"
        # workbook = Workbook()
        # sheet = workbook.active

        headers = ['CUSTOMER NUMBER', 'DISTRIBUTOR', 'DEPARTMENT', 'DATE', 'PO NUMBER', 
                   'PRODUCT NUMBER', 'CUST PROD #', 'DESCRIPTION', 'BRAND', 'PACK SIZE',
                   'CS PRICE', 'EA PRICE', 'CS', 'EA', 'EXTENDED PRICE', 'ORDER #',
                   'STOCK STATUS', 'EXCEPTIONS / AUTO-SUB', 'SHORTED',
                   ]
        
        with open(f'{path_to_save}.csv', 'w', newline='') as csv_file:

            csv_writer = writer(csv_file, delimiter=',')

            csv_writer.writerow(headers)
            for item in item_data:
                
                quantity = item['quantity']
                sku = item['sku']

                csv_writer.writerow([self.store_ids[store], '2195', '0', '2/23/2025', store, sku,
                                     '', '', '', '', '', '',
                                    quantity, '0'
                                    '', '', '', '',])
                
        # for pos, header in enumerate(headers):
        #     sheet.cell(row=1, column=pos+1).value = header
            
        # for pos, sku in enumerate(item_data):
        #     quantity = item_data[sku]['quantity']
        #     #sku = item_data[name]['sku']

        #     sheet.cell(row=pos+2, column=2).value = '2195' # DISTRIBUTOR NUMBER
        #     sheet.cell(row=pos+2, column=3).value = '0' # DEPARTMENT NUMBER
        #     sheet.cell(row=pos+2, column=4).value = '2/23/2025' # Need to figure out how to implement dynamic date
        #     sheet.cell(row=pos+2, column=1).value = self.store_ids[store] # Store ID
        #     sheet.cell(row=pos+2, column=5).value = store # PO Number is the store name
        #     sheet.cell(row=pos+2, column=6).value = sku

        #     sheet.cell(row=pos+2, column=7).value  = ''
        #     sheet.cell(row=pos+2, column=9).value  = ''
        #     sheet.cell(row=pos+2, column=8).value  = ''
        #     sheet.cell(row=pos+2, column=10).value = ''
        #     sheet.cell(row=pos+2, column=11).value = ''
        #     sheet.cell(row=pos+2, column=12).value = ''

        #     sheet.cell(row=pos+2, column=13).value = quantity
        #     sheet.cell(row=pos+2, column=14).value = '0'

        #     sheet.cell(row=pos+2, column=15).value = ''
        #     sheet.cell(row=pos+2, column=16).value = ''
        #     sheet.cell(row=pos+2, column=17).value = ''
        #     sheet.cell(row=pos+2, column=18).value = ''

        
        # workbook.save(filename=f'{path_to_save}.csv')

        return
    
    def retrieve_pricing_sheet(self, guide_name: str) -> str:

        self.login()

        create_order_button = self.driver.find_element(By.ID, 'mainmenu-order')
        create_order_button.click()
        time.sleep(10)

        WebDriverWait(self.driver, 10).until(EC.frame_to_be_available_and_switch_to_it(self.driver.find_element(By.XPATH, '//iframe[@class="bigrounded"]')))
        WebDriverWait(self.driver, 10).until(EC.frame_to_be_available_and_switch_to_it(self.driver.find_element(By.XPATH, '//iframe[@id="Custom-iFrame0"]')))
        custom_guide_select = Select(self.driver.find_element(By.NAME, 'previousOrders'))
        custom_guide_select.select_by_visible_text(guide_name)
        time.sleep(5)

        submit_button = self.driver.find_element(By.ID, 'startOrderButton')
        submit_button.click()
        self.driver.switch_to.default_content()
        
        
        WebDriverWait(self.driver, 10).until(EC.frame_to_be_available_and_switch_to_it(self.driver.find_element(By.XPATH, '//iframe[@id="ContentFrame"]')))
        time.sleep(10)
        advanced_dropdown = self.driver.find_element(By.ID, 'pagemenuli-adv')
        advanced_dropdown.click()
        time.sleep(5)

        export_dropdown = self.driver.find_element(By.ID, 'pagemenuli-exp')
        export_dropdown.click()
        time.sleep(5)

        standard_dropdown = self.driver.find_element(By.ID, 'pagemenuli-standExp')
        standard_dropdown.click()
        time.sleep(5)

        excel_button = self.driver.find_element(By.ID, 'pagemenu-exp2')
        excel_button.click()
        time.sleep(10)

        # We go back to the home page to avoid force logouts
        # home_button = self.driver.find_element(By.CLASS_NAME, 'sprite-Home')
        # home_button.click()
        # time.sleep(10)

        return 'export.csv'

    def get_pricing_info_from_sheet(self, path_to_pricing_sheet: str) -> dict:
        item_info = {}
        with open(path_to_pricing_sheet) as us_food_info:
            csv_reader = reader(us_food_info, delimiter=',')
            for pos, row in enumerate(csv_reader):

                if pos == 0: continue
                if row[7] == 0 or row[7] == '0': continue
                if row[5] == '' or row[5] is None: continue
                
                item_sku 		= row[2]
                item_name 		= row[3]
                # print(item_name, flush=True)

                package_info = row[5]

                quantity, units = PricingBotMixin.helper_format_size_units(package_info)

                # if '/' in package_info:
                #     size = package_info.split('/')[0]
                #     pack = package_info.split('/')[1]
                #     quantity, units = PricingBotMixin.helper_format_size_units(pack, size)
                # else:    
                #     quantity, units = PricingBotMixin.helper_format_size_units(1, row[5])

                cost 			= float(row[7].replace('$', ''))
                # print(f'{item_name}: {quantity} {units}')
                if item_sku in self.special_cases:
                    quantity, units = PricingBotMixin.helper_format_size_units(self.special_cases[item_sku]['pack'], self.special_cases[item_sku]['unit'])

                if item_name not in item_info:
                    item_info[item_name] = {
                        'SKU': item_sku,
                        'quantity': quantity,
                        'units': PricingBotMixin.normalize_units(units),
                        'cost': cost,
                        'case_size': f'{quantity} {units}'
                    }
                print(item_name)
                pprint(item_info[item_name])
        return item_info

    # def get_pricing_info_from_sheet(self, path_to_pricing_sheet: str) -> dict:
    #     item_info = {}
    #     with open(path_to_pricing_sheet) as file:
    #         for pos, row in enumerate(file):

    #             if pos == 0: continue

    #             row_info = row.split('\t')
    #             print(row_info, flush=True)
    #             item_sku  		= row_info[0].split('"')[1]
    #             item_name 		= row_info[7]
    #             cost   	  		= float(row_info[8])
                
    #             if item_sku in self.special_cases:
    #                 quantity, units = PricingBotMixin.helper_format_size_units(self.special_cases[item_sku]['pack'], self.special_cases[item_sku]['unit'])
    #             else:
    #                 quantity, units = PricingBotMixin.helper_format_size_units(row_info[4], row_info[5])

            
    #             if item_name not in item_info:
    #                 item_info[item_name] = {
    #                     'SKU': item_sku,
    #                     'quantity': quantity,
    #                     'units': PricingBotMixin.normalize_units(units),
    #                     'cost': cost,
    #                     'case_size': f'{row_info[4]} / {row_info[5]}'
    #                 }

    #     return item_info