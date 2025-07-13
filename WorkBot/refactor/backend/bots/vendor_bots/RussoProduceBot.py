from .VendorBot import VendorBot, SeleniumBotMixin, PricingBotMixin
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium import webdriver
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from csv import writer
from datetime import date
from pprint import pprint

class RussoProduceBot(VendorBot, PricingBotMixin):

    def __init__(self, **kwargs) -> None:
        VendorBot.__init__(self)
        PricingBotMixin.__init__(self)
        self.name = 'Russo Produce'
        self.minimum_order_case = 5

        self.store_ids = {}

        self.special_cases = {
            # 'Alfalfa Sprouts': {'unit': 'LB', 'pack': 2.25},
            # 'Apple Empire': self._special_case_info('LB', 36),
            # 'Arugula Baby': self._special_case_info('LB', 4),
            # 'Bananas': {'unit': 'LB', 'pack': 40},
            # 'Basil (Cleaned)': {'unit': 'LB', 'pack': .15625},
            # 'Broccoli Crowns': self._special_case_info('LB', 20),
            # 'Brussel Sprouts': {'unit': 'LB', 'pack': 25},
            # 'Cabbage NAPA': self._special_case_info('LB', 30),
            # 'Cabbage Shredded': {'unit': 'LB', 'pack': 25},
            # 'Carrots (25lb)': self._special_case_info('LB', 25),
            # 'Carrots (50lb)': self._special_case_info('LB', 50),
            # 'Cauliflower': {'unit': 'EA', 'pack': 12},
            # 'Celery': self._special_case_info('LB', 20),
            # 'Cilantro': {'unit': 'EA', 'pack': 1},
            # 'Collard Greens': {'unit': 'EA', 'pack': 9},
            # 'Cucumbers English': {'unit': 'EA', 'pack': 12},
            # 'Garlic Peeled': {'unit': 'LB', 'pack': 5},
            # 'Ginger Root': {'unit': 'LB', 'pack': 1},
            # 'Grapes Seedless': {'unit': 'LB', 'pack': 18},
            # 'Green Beans': self._special_case_info('LB', 20),
            # 'Honeydew Melon': {'unit': 'EA', 'pack': 5.5},
            # 'Kale': {'unit': 'EA', 'pack': 24},
            # 'Kiwi': {'unit': 'EA', 'pack': 1},
            # 'Lemon': {'unit': 'EA', 'pack': 140},
            # 'Lettuce Green Leaf Fillet': self._special_case_info('LB', 10),
            # 'Limes': {'unit': 'LB', 'pack': 10},
            # 'Mushrooms Portobello Caps': self._special_case_info('LB', 5),
            # 'Mushrooms Thick Sliced': {'unit': 'LB', 'pack': 10},
            # 'Mushrooms Whole Button': {'unit': 'LB', 'pack': 10},
            # 'Onion Red': {'unit': 'LB', 'pack': 25},
            # 'Onion Spanish': {'unit': 'LB', 'pack': 50},
            # 'Onions Green  (Scallions)': self._special_case_info('EA', 48),
            # 'Oregano': self._special_case_info('EA', 1),
            # 'Parsley': self._special_case_info('EA', 1),
            # 'Pepper Green Bell (Large) (RENZI = 20lb)': {'unit': 'LB', 'pack': 25},
            # 'Pepper Green Bell Xtra Large': {'unit': 'LB', 'pack': 25},
            # 'Pepper Red Bell (11 lb)': {'unit': 'LB', 'pack': 11},
            # 'Peppers Jalapeno': {'unit': 'LB', 'pack': 1},
            # 'Pineapple': self._special_case_info('EA', 7.5),
            # 'Pomegranate Whole': self._special_case_info('EA', 30),
            # 'Potato Chef': self._special_case_info('LB', 50),
            # 'Potato Red "A"': self._special_case_info('LB', 50),
            # 'Potato Russet': self._special_case_info('LB', 50),
            # 'Romaine Hearts (Washed & Trimmed)': self._special_case_info('LB', 12),
            # 'Salad Blend (Heritage/Arcadia)': self._special_case_info('LB', 3),
            # 'Radish': self._special_case_info('LB', 14),
            # 'Spinach Baby': self._special_case_info('LB', 10),
            # 'Strawberries Fresh': self._special_case_info('LB', 8),
            # 'Snow Peas': self._special_case_info('LB', 10),
            # 'Sugar Snap Peas (24 lb bags Sysco)': self._special_case_info('LB', 10),
            # 'Tomato Grape': self._special_case_info('LB', 20),
            # 'Tomatos 5x6': self._special_case_info('LB', 25),
        }

    def format_for_file_upload(self, item_data: dict, path_to_save: str):
        pass
    
    def get_pricing_info_from_sheet(self, path_to_pricing_sheet: str) -> dict:
        workbook = load_workbook(path_to_pricing_sheet)
        sheet    = workbook.active

        row_info  = list(PricingBotMixin.helper_iter_rows(sheet))[1:]
        item_info = {}

        numeric_non_numerics = ['.', '-', '/']
        for row in row_info:
            item_sku  = str(row[1])
            item_name = str(row[0])

            if row[3] == None or row[2] == None or row[1] == None: continue

            # print(row, flush=True)
            pack_size = ''
            unit = ''
            if item_name not in self.special_cases:
                # continue
                # print(item_name)
                pack_size_info = row[2]
                
                # Extract units
                print(item_name)
                for char in str(pack_size_info):
                    if char.isnumeric() or char in numeric_non_numerics:
                        pack_size += char

                    else: unit += char

                # Average if necessary
                if '-' in pack_size:
                    lower, upper = pack_size.split('-')
                    pack_size = ( int(lower) + int(upper) ) / 2

    
                # Combine if necessary
                elif '/' in pack_size:
                    packs = pack_size.split('/')[0] or 1
                    size = pack_size.split('/')[1] or 1

                    pack_size = ( float(packs) * float(size) )

            else:
                unit = self.special_cases[item_name]['unit']
                pack_size = self.special_cases[item_name]['pack']

            cost = float(row[3])

            if pack_size == None or pack_size == '': continue
            if item_name not in item_info:
                item_info[item_name] = {
                    'SKU': item_sku,
                    'quantity': pack_size,
                    'units': PricingBotMixin.normalize_units(unit),
                    'cost': cost,
                    'case_size': row[2]
                }
            

        return item_info
    
    def retrieve_pricing_sheet(self, pricing_guide) -> None:
        return 'Russo Produce_IBProduce.xlsx'