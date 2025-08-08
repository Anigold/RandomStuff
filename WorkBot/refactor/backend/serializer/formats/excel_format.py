from openpyxl import Workbook, load_workbook
from pathlib import Path
from typing import Any
from .base_format import BaseFormat

class ExcelFormat(BaseFormat):
    
    default_suffix = ".xlsx"

    def write(self, headers: list[str], rows: list[list[Any]]) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        return wb

    def read(self, file_path: Path) -> list[list[Any]]:
        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active
        data = [[cell for cell in row] for row in sheet.iter_rows(min_row=2, values_only=True)]
        wb.close()
        return data