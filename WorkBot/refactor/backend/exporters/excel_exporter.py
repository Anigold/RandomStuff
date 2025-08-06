from backend.exporters.exporter import Exporter
from backend.models.order import Order
from backend.models.item import Item
from backend.models.transfer import Transfer
from openpyxl import Workbook
from .adapters import ExportAdapter

@Exporter.register_exporter(Order, "excel")
class ExcelOrderExporter(Exporter):
    def export(self, order: Order, adapter: ExportAdapter = None, context: dict = None) -> Workbook:
        wb = Workbook()
        ws = wb.active

        # Default headers and row logic
        default_headers = ["SKU", "Name", "Quantity", "Cost Per", "Total Cost"]
        default_row_fn = lambda item: [item.sku, item.name, item.quantity, item.cost_per, item.total_cost]

        # Use adapter to modify headers (with context if needed)
        headers = adapter.modify_headers(default_headers, context=context) if adapter else default_headers


        ws.append(headers)

        for item in order.items:
            row = default_row_fn(item)
            if adapter:
                row = adapter.modify_row(row, item=item, context=context)
            ws.append(row)

        return wb


@Exporter.register_exporter(Item, "excel")
class ExcelItemExporter(Exporter):
    def export(self, item: Item) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Quantity"])
        ws.append([item.name, item.quantity])
        return wb


@Exporter.register_exporter(Transfer, 'excel')
class ExcelTransferExporter(Exporter):
    def export(self, transfer: Transfer, adapter: ExportAdapter = None, context: dict = None) -> Workbook:
        wb = Workbook()
        ws = wb.active

        default_headers = ['Item Name', 'Quantity']
        default_row_fn = lambda transfer_item: [transfer_item.name, transfer_item.quantity]

        headers = adapter.modify_headers(default_headers, context=context) if adapter else default_headers

        ws.append(headers)

        for item in transfer.transfer_items:
            row = default_row_fn(item)
            if adapter:
                row = adapter.modify_row(row, item=item, context=context)
            ws.append(row)

        return wb