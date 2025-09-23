# backend/app/formatters/excel_formatter.py

from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List
from openpyxl import Workbook, load_workbook
from .base_format import BaseFormatter


class ExcelFormatter(BaseFormatter):
    """Domain-agnostic Excel formatter: expects a dict with 'headers' and 'rows'."""

    def format_name(self) -> str:
        return "xlsx"

    def dumps(self, obj: Dict[str, Any], context: dict | None = None) -> bytes:
        """
        obj should be of the form:
        {
            "headers": ["col1", "col2", ...],
            "rows": [
                ["a", "b", ...],
                ["c", "d", ...],
            ]
        }
        """
        wb = Workbook()
        ws = wb.active
        ws.append(obj.get("headers", []))
        for row in obj.get("rows", []):
            ws.append(row)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def loads(self, data: bytes) -> Dict[str, Any]:

        buf = BytesIO(data)
        wb = load_workbook(buf)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"headers": [], "rows": []}
        headers, *body = rows
        return {"headers": list(headers), "rows": [list(r) for r in body]}

    def load_path(self, path: Path) -> Dict[str, Any]:
        wb = load_workbook(path)
        ws = wb.active
        rows = [[cell for cell in row] for row in ws.iter_rows(min_row=2, values_only=True)]
        
        if not rows:
            return {"headers": [], "rows": []}
        headers, *body = rows
        return {"headers": list(headers), "rows": [list(r) for r in body]}
