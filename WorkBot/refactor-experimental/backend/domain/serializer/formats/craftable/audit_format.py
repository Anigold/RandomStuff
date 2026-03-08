from pathlib import Path
from typing import Any, Dict, Optional
from openpyxl import load_workbook

from backend.core.interfaces.formatter import BaseFormatter


class CraftableAuditFormatter(BaseFormatter[Dict[str, Any]]):

    def format_name(self) -> str:
        return "audit"

    def dumps(self, obj: Dict[str, Any], format: Optional[str] = None, context: dict | None = None) -> bytes:
        raise NotImplementedError("CraftableAuditFormatter is import-only.")

    def loads(self, data: bytes, format: Optional[str] = None) -> Dict[str, Any]:
        raise NotImplementedError(
            "CraftableAuditFormatter.loads(bytes) is not supported. Use load_path()."
        )

    def load_path(self, path: Path, context: dict | None = None) -> Dict[str, Any]:
        context = context or {}

        wb = load_workbook(path, data_only=True)

        sheet_name = "4. Summary By Item"
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f'Expected sheet "{sheet_name}" not found in {path.name}. Found sheets: {wb.sheetnames}'
            )

        ws = wb[sheet_name]
        header_row_idx = 7

        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in ws[header_row_idx]
        ]

        rows = [
            list(row)
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True)
            if any(cell is not None and str(cell).strip() != "" for cell in row)
        ]

        return {
            "headers": headers,
            "rows": rows,
            "metadata": {
                "source": "craftable",
                **context,
            },
        }