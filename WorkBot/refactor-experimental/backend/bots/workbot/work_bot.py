# ==========================================================
#                         WORKBOT
# ==========================================================
"""
Coordinates all system domains (Orders, Vendors, Transfers, Emails)
and automates Craftable operations.
"""

# Standard Library
from datetime import datetime
from pathlib import Path
from typing import List

# Third-Party Libraries
from openpyxl import load_workbook

# Internal
from backend.infra.logger import Logger
from backend.bots.craftable_bot.craftable_bot import CraftableBot
from backend.app.services import (
    OrderServices,
    VendorServices,
    StoreServices,
    TransferServices,
    EmailServices,
)
from backend.domain.models import (
    Order,
    Transfer, TransferItem,
    Vendor,
    Store,
    Item
)
from backend.adapters.emailer.emailer import Email


@Logger.attach_logger
class WorkBot:
 
# ------------------------------------------------------
# INIT / SETUP
# ------------------------------------------------------
    def __init__(self,
                orders_service:    OrderServices,
                transfers_service: TransferServices,
                vendors_service:   VendorServices,
                stores_service:    StoreServices,
                emails_service:    EmailServices,
                craftable_bot:     CraftableBot,
                 ):
        self.logger.info('Initializing WorkBot...')

        self.orders =    orders_service
        self.transfers = transfers_service
        self.vendors =   vendors_service
        self.stores =    stores_service

        self.emails = emails_service

        self.craft_bot = craftable_bot

        self.logger.info('WorkBot initialized successfully.')

# ------------------------------------------------------
# CRAFTABLE BOT ACTIONS
# ------------------------------------------------------
    def download_craftable_orders(self, stores, vendors=[], download_pdf=True, update=True):
        return self.craft_bot.download_orders(stores, vendors, download_pdf=download_pdf, update=update)

    def delete_craftable_orders(self, stores, vendors=[]):
        return self.craft_bot.delete_orders(stores, vendors)

    def input_craftable_transfers(self):
        transfers = self.transfers.list_transfers()
        self.logger.info(transfers)
        return self.craft_bot.input_transfers(transfers)

    def download_audits(self, stores: list[str], start_date: str, end_date: str) -> None:
        self.craft_bot.download_audits(stores, start_date, end_date)

    def add_items_to_craftable_order(self, store: str, vendor: str, items: List[Item]) -> None:
        ...
    
    def remove_items_from_craftable_order(self, store: str, vendor: str, items: List[Item]) -> None:
        ...

# ------------------------------------------------------
# ORDER MANAGEMENT
# ------------------------------------------------------
    def combine_orders(self, vendor: str) -> None:
        return self.orders.combine_orders(vendor)
    
    def get_orders(self, stores: list[str], vendors: list[str]) -> list[Order]:
        self.logger.info(f'Retrieving orders for: stores={stores}, vendors={vendors}')
        return self.orders.get_orders(stores=stores, vendors=vendors)

    def archive_all_current_orders(self, stores: list[str] = None, vendors: list[str] = None) -> None:
        vendors = vendors or []

        orders = self.get_orders(stores, vendors)

        for order in orders:
            try:
                self.orders.archive_order(order)
            except Exception as e:
                self.logger.warning(f'[Archive] Skipped {order}: {e}')

# ------------------------------------------------------
# TRANSFER MANAGEMENT
# ------------------------------------------------------
    def get_transfers(self) -> list[Transfer]:
        '''
        Retrieves saved transfer objects from file based on optional filters.

        Args:
            stores (list[str], optional): List of store names to filter by.
            start_date (str, optional): Start date filter in YYYY-MM-DD format.
            end_date (str, optional): End date filter in YYYY-MM-DD format.

        Returns:
            list[Transfer]: List of parsed transfer domain objects.
        '''
        return self.transfers.list_transfers()

    def convert_order_to_transfer(self, destination, vendor, origin):
        self.logger.info(f'Beginning order-transfer conversion: {destination}-{vendor} -> {origin}')
        order = self.get_orders([destination], [vendor])[0]
        origin = 'Bakery' if order.vendor == 'Ithaca Bakery' else order.vendor # YOU NEED TO FIX THIS FOR WHEN YOU HAVE TO DO IT FOR A DIFFERENT STORE

        transfer_items = [
            TransferItem(name=item.name, quantity=item.quantity)
            for item in order.items
        ]

        transfer = Transfer(
            transfer_items=transfer_items,
            origin=origin,
            destination=order.store,
            transfer_date=order.date
        )

        return self.transfers.save_transfer(transfer=transfer)

# ------------------------------------------------------
# EMAIL OPERATIONS
# ------------------------------------------------------
    def generate_vendor_order_emails(self, vendors: list[str], stores: list[str] = []) -> list[Email]:
        return self.emails.send_vendor_order_emails(stores=stores, vendors=vendors)

    def generate_store_order_emails(self, stores: list[str]):
        return self.emails.send_store_order_emails(stores=stores)

# ------------------------------------------------------
# SUPPORT / LOOKUP
# ------------------------------------------------------ 
    def list_all_vendors(self) -> List[Vendor]:
        return self.vendors.list_vendors()
    
    def list_all_stores(self) -> List[Store]:
        return self.stores.list_stores()
    
    def get_vendor_information(self, vendor_name: str) -> Vendor:
        return self.vendors.get_vendor(vendor_name)

    def get_store_information(self, store_name: str) -> dict:
        return self.stores.get_store(store_name)

# ------------------------------------------------------
# FILE GENERATION
# ------------------------------------------------------ 
    def generate_vendor_upload_files(
                self,
                stores: list[str],
                vendors: list[str],
                start_date: str = None,
                end_date: str = None
            ) -> list[Path]:

                self.logger.info(f'Generating vendor upload files for stores={stores}, vendors={vendors}, '
                                f'start_date={start_date}, end_date={end_date}')

                orders = self.get_orders(stores, vendors)
                self.logger.info(f'Found {len(orders)} orders to process.')

                context_map = {}
                for order in orders:
            
                    vendor_info = self.vendors.get_vendor(order.vendor)

                    context_map[f'{order.store}|{order.vendor}'] = {
                        'store':       order.store,
                        'vendor_info': vendor_info,
                        'date_str':    order.date,
                    }
                    
                self.logger.info(f'Context map built with {len(context_map)} entries. Delegating to OrderServices.')

                result_paths = self.orders.generate_vendor_uploads(
                    vendors=vendors,
                    stores=stores,
                    start_date=start_date,
                    end_date=end_date,
                    context_map=context_map
                )

                self.logger.info(f'Vendor upload file generation complete. {len(result_paths)} files created.')

                return result_paths

    def split_natalies(self) -> None:
        natalies_excel_path = Path('C:/Users/Will/Desktop/Natalies.xlsx')
        workbook = load_workbook(natalies_excel_path)
        sheet = workbook.active

        flavors = [
            row[0].value for row in sheet.iter_rows()
            if row[0].value and row[0].value.strip()
        ]

        orders = self.get_orders(
            stores=['Bakery', 'Easthill', 'Collegetown', 'Triphammer', 'Downtown'],
            vendors=['Performance Food', 'FingerLakes Farms']
        )

        store_indices = {
            'Collegetown': 2,
            'Downtown': 5,
            'Easthill': 8,
            'Triphammer': 11,
            'Bakery': 14
        }

        for order in orders:
            for item in order.items:
                if 'Natalie' in item.name:
                    flavor = item.name.split(' - ')[1]
                    if flavor in flavors:
                        row = flavors.index(flavor) + 4
                        col = store_indices.get(order.store, None)
                        if col:
                            sheet.cell(row=row, column=col).value = item.quantity

        workbook.save('C:/Users/Will/Desktop/Natalies.xlsx')
   
# ------------------------------------------------------
# MAINTENANCE / UTILITIES
# ------------------------------------------------------  
    def close_craftable_session(self):
        self.craft_bot.close_session()

    def welcome_to_work(self) -> str:
        today = self._get_today_date_and_day()
        return f'''\n{today[1]}, {today[0]}'''

    def shutdown(self) -> None:
        self.close_craftable_session()
        print('Exiting WorkBot CLI.')

    def _get_today_date_and_day(self):
        today = datetime.today()
        day_of_week = today.strftime('%A')
        day = today.day
        suffix = 'th' if 11 <= day <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')
        long_date = today.strftime(f'%B {day}{suffix}, %Y')
        return long_date, day_of_week

    def testing_function(self, stores: str, vendors: str) -> None:
        return self.craft_bot.new_download_orders(stores, vendors)
    