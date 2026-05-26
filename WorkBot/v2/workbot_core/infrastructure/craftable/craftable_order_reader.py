from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from workbot_core.application.dto.order_import_row import (
    OrderImportRow,
    OrderLineImportRow,
)


@dataclass(frozen=True, slots=True)
class CraftableOrderReaderConfig:
    store_id: str
    vendor_id: str
    order_date: date
    delivery_date: date | None = None

    source: str = "craftable"
    source_reference: str | None = None

    sheet_name: str | None = None

    sku_header: str = "SKU"
    item_name_header: str = "Name"
    quantity_header: str = "Quantity"
    unit_price_header: str = "Cost per"
    total_cost_header: str = "Total Cost"


class CraftableOrderReader:
    """Reads a Craftable order spreadsheet and converts it to OrderImportRow.

    This reader does not save anything and does not know about repositories.
    """

    def read_file(
        self,
        path: Path,
        *,
        config: CraftableOrderReaderConfig,
    ) -> OrderImportRow:
        if not path.exists():
            raise FileNotFoundError(f"Craftable order file does not exist: {path}")

        workbook = load_workbook(path, data_only=True)

        if config.sheet_name is None:
            sheet = workbook.active
        else:
            sheet = workbook[config.sheet_name]

        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise ValueError(f"Craftable order file is empty: {path}")

        header_index = self._find_header_row(
            rows,
            required_headers=(
                config.item_name_header,
                config.quantity_header,
            ),
        )

        header_index = self._find_header_row(
            rows,
            required_headers=(
                config.sku_header,
                config.item_name_header,
                config.quantity_header,
                config.unit_price_header,
                config.total_cost_header,
        ),
)

        headers = self._normalize_headers(rows[header_index])
        line_rows = rows[header_index + 1 :]

        sku_index = self._header_index(headers, config.sku_header)
        item_name_index = self._header_index(headers, config.item_name_header)
        quantity_index = self._header_index(headers, config.quantity_header)
        unit_price_index = self._header_index(headers, config.unit_price_header)
        total_cost_index = self._header_index(headers, config.total_cost_header)

        lines: list[OrderLineImportRow] = []

        for raw_row in line_rows:
            source_item_name = self._cell_str_at(raw_row, item_name_index)

            if source_item_name is None:
                continue

            quantity = self._decimal_at(raw_row, quantity_index)

            if quantity is None:
                continue

            source_vendor_sku = self._cell_str_at(raw_row, sku_index)
            unit_price = self._decimal_at(raw_row, unit_price_index)
            total_cost = self._decimal_at(raw_row, total_cost_index)

            lines.append(
                OrderLineImportRow(
                    source_item_name=source_item_name,
                    source_vendor_sku=source_vendor_sku,
                    quantity=quantity,
                    unit_price=unit_price,
                    total_cost=total_cost,
                )
            )

        source_reference = config.source_reference or str(path)

        return OrderImportRow(
            store_id=config.store_id,
            vendor_id=config.vendor_id,
            order_date=config.order_date,
            delivery_date=config.delivery_date,
            source=config.source,
            source_reference=source_reference,
            lines=tuple(lines),
        )

    def _find_header_row(
        self,
        rows: list[tuple[Any, ...]],
        *,
        required_headers: tuple[str, ...],
    ) -> int:
        normalized_required = {
            self._normalize_header(header)
            for header in required_headers
        }

        for index, row in enumerate(rows):
            normalized_row = {
                self._normalize_header(value)
                for value in row
                if value is not None
            }

            if normalized_required.issubset(normalized_row):
                return index

        raise ValueError(
            "Could not find header row containing: "
            + ", ".join(required_headers)
        )

    def _normalize_headers(self, row: tuple[Any, ...]) -> tuple[str, ...]:
        return tuple(self._normalize_header(value) for value in row)

    def _header_index(self, headers: tuple[str, ...], header: str) -> int:
        normalized = self._normalize_header(header)

        try:
            return headers.index(normalized)
        except ValueError as exc:
            raise ValueError(f"Missing required header: {header}") from exc

    def _optional_header_index(
        self,
        headers: tuple[str, ...],
        header: str,
    ) -> int | None:
        normalized = self._normalize_header(header)

        try:
            return headers.index(normalized)
        except ValueError:
            return None

    @staticmethod
    def _normalize_header(value: object) -> str:
        return " ".join(str(value).strip().casefold().split())

    @staticmethod
    def _cell_str_at(row: tuple[Any, ...], index: int | None) -> str | None:
        if index is None or index >= len(row):
            return None

        value = row[index]

        if value is None:
            return None

        cleaned = str(value).strip()

        return cleaned or None

    @classmethod
    def _decimal_at(cls, row: tuple[Any, ...], index: int) -> Decimal | None:
        value = cls._cell_str_at(row, index)

        if value is None:
            return None

        try:
            return Decimal(value)
        except InvalidOperation as exc:
            raise ValueError(f"Invalid quantity value: {value!r}") from exc