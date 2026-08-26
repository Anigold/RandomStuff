from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.domain.models import Item, VendorItemInfo


DATE_FORMAT = "%m/%d/%Y"


@dataclass(slots=True)
class PurchaseVendorInfoRow:
    item_name: str
    vendor_name: str
    vendor_id: str
    sku: str
    order_unit: str
    last_price: float | None
    last_ordered: str | None


def update_items_from_purchase_log(
    *,
    workbot,
    purchase_log_path: Path,
) -> None:
    latest_rows = _collect_latest_vendor_rows_by_item_name(
        workbot=workbot,
        purchase_log_path=purchase_log_path,
    )
    workbot.logger.info(f'Number of rows: {len(latest_rows)}')
    updated_count = 0
    added_count = 0
    changed_price_count = 0
    skipped_missing_items = 0

    count = 0
    for row in latest_rows:
        workbot.logger.info(f'{count + 1}')
        try:
            item = workbot.items.get_item_by_name(row.item_name)
        except FileNotFoundError:
            skipped_missing_items += 1
            continue

        changed, added, price_changed = _apply_vendor_info(item, row)

        if changed:
            workbot.items.save(item)
            updated_count += 1
            added_count += int(added)
            changed_price_count += int(price_changed)

        count = count + 1
        
    print("Vendor item info update complete.")
    print(f"Items updated: {updated_count}")
    print(f"Vendor entries added: {added_count}")
    print(f"Prices changed: {changed_price_count}")
    print(f"Skipped missing items: {skipped_missing_items}")


def _collect_latest_vendor_rows_by_item_name(
    *,
    workbot,
    purchase_log_path: Path,
) -> list[PurchaseVendorInfoRow]:
    wb = load_workbook(purchase_log_path, data_only=True)
    ws = wb.active

    headers = _read_headers(ws)

    latest: dict[tuple[str, str, str], tuple[datetime, PurchaseVendorInfoRow]] = {}

    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        row = _row_to_dict(headers, excel_row)

        item_name = _clean_str(row.get("ITEM"))
        vendor_name = _clean_str(row.get("VENDOR"))
        sku = _clean_str(row.get("SKU"))

        if not item_name or not vendor_name or not sku:
            continue

        received_date = _parse_date(row.get("RECEIVED DATE"))
        invoice_date = _parse_date(row.get("INVOICE DATE"))
        effective_date = received_date or invoice_date

        if effective_date is None:
            continue
        
        vendor = workbot.get_vendor_information(vendor_name)
        if not vendor: continue
        vendor_id = vendor.id

        purchase_row = PurchaseVendorInfoRow(
            item_name=item_name,
            vendor_name=vendor_name,
            vendor_id=vendor_id,
            sku=sku,
            order_unit=_clean_str(row.get("PACK SIZE")),
            last_price=_to_float(row.get("CU PRICE")),
            last_ordered=effective_date.strftime("%Y-%m-%d"),
        )

        key = (
            item_name.casefold(),
            vendor_id,
            sku.casefold(),
        )

        current = latest.get(key)
        if current is None or effective_date > current[0]:
            latest[key] = (effective_date, purchase_row)

    return [entry[1] for entry in latest.values()]

def _apply_vendor_info(
    item: Item,
    row: PurchaseVendorInfoRow,
) -> tuple[bool, bool, bool]:
    item.vendor_info = item.vendor_info or {}

    key = _vendor_info_key(row.vendor_id, row.sku)
    existing = item.vendor_info.get(key)

    if existing is None:
        item.vendor_info[key] = VendorItemInfo(
            vendor_id=row.vendor_id,
            vendor_name=row.vendor_name,
            sku=row.sku,
            order_unit=row.order_unit,
            last_price=row.last_price,
            last_ordered=row.last_ordered,
            is_primary=False,
            is_orderable=True,
            is_active=True,
        )
        return True, True, False

    changed = False
    price_changed = False

    if row.last_price is not None and existing.last_price != row.last_price:
        existing.last_price = row.last_price
        changed = True
        price_changed = True

    if row.last_ordered and existing.last_ordered != row.last_ordered:
        existing.last_ordered = row.last_ordered
        changed = True

    if row.order_unit and existing.order_unit != row.order_unit:
        existing.order_unit = row.order_unit
        changed = True

    if not existing.vendor_name and row.vendor_name:
        existing.vendor_name = row.vendor_name
        changed = True

    return changed, False, price_changed


def _vendor_info_key(vendor_id: str, sku: str) -> str:
    return f"{vendor_id}|{sku}"


def _read_headers(ws) -> list[str]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return [_clean_str(value) for value in header_row]


def _row_to_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    return dict(zip(headers, row))


def _clean_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_date(value: Any) -> datetime | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    text = _clean_str(value)
    if not text:
        return None

    try:
        return datetime.strptime(text, DATE_FORMAT)
    except ValueError:
        return None


def _to_float(value: Any) -> float | None:
    if value is None:
        return None

    text = str(value).replace("$", "").replace(",", "").strip()
    if not text:
        return None

    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        return None