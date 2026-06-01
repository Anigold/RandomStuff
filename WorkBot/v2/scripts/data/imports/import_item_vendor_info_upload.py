from __future__ import annotations

import argparse
from dataclasses import replace
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from workbot_core.domain.models.item_vendor_info import ItemVendorInfo
from workbot_core.infrastructure.database.repositories.item_repository import (
    SqlItemRepository,
)
from workbot_core.infrastructure.database.repositories.item_vendor_info_repository import (
    SqlItemVendorInfoRepository,
)
from workbot_core.infrastructure.database.repositories.vendor_repository import (
    SqlVendorRepository,
)
from workbot_core.infrastructure.database.session import create_session
from workbot_core.utils.ids import IdGenerator


MAX_VENDOR_COLUMNS = 10


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import item and item vendor information from an upload workbook."
    )

    parser.add_argument(
        "path",
        type=Path,
        help="Path to the item/vendor info upload workbook.",
    )

    parser.add_argument(
        "--sheet-name",
        help="Optional sheet name. Defaults to the active sheet.",
    )

    parser.add_argument(
        "--output",
        type=Path,
        default=Path("item_vendor_info_upload_report.xlsx"),
        help="Output report workbook path.",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Read and validate the file, but roll back database changes.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()

    rows = read_upload_rows(args.path, sheet_name=args.sheet_name)
    report = UploadReport()

    with create_session() as session:
        items = SqlItemRepository(session)
        vendors = SqlVendorRepository(session)
        item_vendor_infos = SqlItemVendorInfoRepository(session)

        for row_number, row in rows:
            item_name = clean_text(row.get("ITEM NAME"))

            if not item_name:
                report.skipped_rows.append(
                    {
                        "row": row_number,
                        "reason": "Missing ITEM NAME",
                    }
                )
                continue

            item = items.get_by_name(item_name)

            if item is None:
                report.missing_items[item_name] = {
                    "row": row_number,
                    "item_name": item_name,
                }
                continue

            updated_item = update_item_from_upload_row(item, row)
            items.save(updated_item)

            report.updated_items.append(
                {
                    "row": row_number,
                    "item_name": updated_item.name,
                    "item_id": updated_item.id,
                    "category": updated_item.category,
                    "subcategory": updated_item.subcategory,
                    "count_unit_quantity": updated_item.count_unit_quantity,
                    "count_unit_measure": updated_item.count_unit_measure,
                    "custom_each_name": updated_item.custom_each_name,
                    "each_quantity": updated_item.each_quantity,
                    "each_measure": updated_item.each_measure,
                    "weight_quantity": updated_item.weight_quantity,
                    "weight_measure": updated_item.weight_measure,
                    "volume_quantity": updated_item.volume_quantity,
                    "volume_measure": updated_item.volume_measure,
                }
            )

            item = updated_item

            for index in range(1, MAX_VENDOR_COLUMNS + 1):
                vendor_name = clean_text(row.get(f"VENDOR NAME {index}"))

                if not vendor_name:
                    continue

                vendor = vendors.get_by_name(vendor_name)

                sku = clean_text(row.get(f"SKU {index}"))
                purchase_unit = clean_text(row.get(f"PU {index}"))
                pack_size = optional_decimal(row.get(f"QTY {index}"))
                price = optional_decimal(row.get(f"COT (PU) {index}"))

                if vendor is None:
                    report.missing_vendors[vendor_name] = {
                        "row": row_number,
                        "vendor_name": vendor_name,
                        "item_name": item_name,
                        "sku": sku,
                    }
                    continue

                if not sku:
                    report.skipped_vendor_infos.append(
                        {
                            "row": row_number,
                            "item_name": item_name,
                            "vendor_name": vendor_name,
                            "reason": f"Missing SKU {index}",
                        }
                    )
                    continue

                existing = item_vendor_infos.get_by_item_vendor_sku(
                    item_id=item.id,
                    vendor_id=vendor.id,
                    vendor_sku=sku,
                )

                if existing is None:
                    info = ItemVendorInfo(
                        id=IdGenerator.item_vendor_info_id(
                            exists=item_vendor_infos.exists
                        ),
                        item_id=item.id,
                        vendor_id=vendor.id,
                        vendor_sku=sku,
                        purchase_unit=purchase_unit,
                        pack_size=pack_size,
                        price=price,
                        is_active=True,
                    )

                    item_vendor_infos.save(info)

                    report.created.append(
                        {
                            "row": row_number,
                            "item_name": item.name,
                            "item_id": item.id,
                            "vendor_name": vendor.name,
                            "vendor_id": vendor.id,
                            "vendor_sku": sku,
                            "purchase_unit": purchase_unit,
                            "pack_size": pack_size,
                            "price": price,
                        }
                    )

                else:
                    updated = replace(
                        existing,
                        purchase_unit=purchase_unit,
                        pack_size=pack_size,
                        price=price,
                        is_active=True,
                    )

                    item_vendor_infos.save(updated)

                    report.updated.append(
                        {
                            "row": row_number,
                            "item_name": item.name,
                            "item_id": item.id,
                            "vendor_name": vendor.name,
                            "vendor_id": vendor.id,
                            "vendor_sku": sku,
                            "purchase_unit": purchase_unit,
                            "pack_size": pack_size,
                            "price": price,
                        }
                    )

        if args.dry_run:
            session.rollback()
            action = "rolled back dry run"
        else:
            session.commit()
            action = "committed"    

    write_report(args.output, report, action=action, dry_run=args.dry_run)

    print("Item vendor info upload complete.")
    print(f"  Action:                 {action}")
    print(f"  Updated item metadata:  {len(report.updated_items)}")
    print(f"  Created vendor info:    {len(report.created)}")
    print(f"  Updated vendor info:    {len(report.updated)}")
    print(f"  Missing items:          {len(report.missing_items)}")
    print(f"  Missing vendors:        {len(report.missing_vendors)}")
    print(f"  Skipped rows:           {len(report.skipped_rows)}")
    print(f"  Skipped vendor infos:   {len(report.skipped_vendor_infos)}")
    print(f"  Report:                 {args.output}")


class UploadReport:
    def __init__(self) -> None:
        self.updated_items: list[dict[str, Any]] = []

        self.created: list[dict[str, Any]] = []
        self.updated: list[dict[str, Any]] = []
        self.missing_items: dict[str, dict[str, Any]] = {}
        self.missing_vendors: dict[str, dict[str, Any]] = {}
        self.skipped_rows: list[dict[str, Any]] = []
        self.skipped_vendor_infos: list[dict[str, Any]] = []

    @property
    def has_blocking_errors(self) -> bool:
        return False


def read_upload_rows(
    path: Path,
    *,
    sheet_name: str | None,
) -> list[tuple[int, dict[str, Any]]]:
    if not path.exists():
        raise FileNotFoundError(f"Upload file does not exist: {path}")

    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise ValueError(f"Upload file is empty: {path}")

    headers = [normalize_header(value) for value in rows[0]]

    if "ITEM NAME" not in headers:
        raise ValueError("Missing required header: ITEM NAME")

    output: list[tuple[int, dict[str, Any]]] = []

    for excel_row_number, values in enumerate(rows[1:], start=2):
        row = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
            if header
        }

        if is_blank_row(row):
            continue

        output.append((excel_row_number, row))

    return output


def write_report(path: Path, report: UploadReport, *, action: str, dry_run: bool) -> None:
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"

    summary.append(["Action", action])
    summary.append(["Dry run", dry_run])
    summary.append(["Updated item metadata", len(report.updated_items)])
    summary.append(["Created vendor info", len(report.created)])
    summary.append(["Updated vendor info", len(report.updated)])
    summary.append(["Missing items", len(report.missing_items)])
    summary.append(["Missing vendors", len(report.missing_vendors)])
    summary.append(["Skipped rows", len(report.skipped_rows)])
    summary.append(["Skipped vendor info rows", len(report.skipped_vendor_infos)])

    add_sheet(
        workbook,
        "Updated Items",
        report.updated_items,
        [
            "row",
            "item_name",
            "item_id",
            "category",
            "subcategory",
            "count_unit_quantity",
            "count_unit_measure",
            "custom_each_name",
            "each_quantity",
            "each_measure",
            "weight_quantity",
            "weight_measure",
            "volume_quantity",
            "volume_measure",
        ],
    )

    add_sheet(
        workbook,
        "Created Vendor Info",
        report.created,
        [
            "row",
            "item_name",
            "item_id",
            "vendor_name",
            "vendor_id",
            "vendor_sku",
            "purchase_unit",
            "pack_size",
            "price",
        ],
    )

    add_sheet(
        workbook,
        "Updated Vendor Info",
        report.updated,
        [
            "row",
            "item_name",
            "item_id",
            "vendor_name",
            "vendor_id",
            "vendor_sku",
            "purchase_unit",
            "pack_size",
            "price",
        ],
    )

    add_sheet(
        workbook,
        "Missing Items",
        list(report.missing_items.values()),
        [
            "row",
            "item_name",
        ],
    )

    add_sheet(
        workbook,
        "Missing Vendors",
        list(report.missing_vendors.values()),
        [
            "row",
            "vendor_name",
            "item_name",
            "sku",
        ],
    )

    add_sheet(
        workbook,
        "Skipped Rows",
        report.skipped_rows,
        [
            "row",
            "reason",
        ],
    )

    add_sheet(
        workbook,
        "Skipped Vendor Infos",
        report.skipped_vendor_infos,
        [
            "row",
            "item_name",
            "vendor_name",
            "reason",
        ],
    )

    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            sheet.column_dimensions[column_letter].width = min(max_length + 2, 48)

    workbook.save(path)


def add_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
    headers: list[str],
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)

    for row in rows:
        sheet.append([format_cell_value(row.get(header)) for header in headers])


def update_item_from_upload_row(item, row: dict[str, Any]):
    return replace(
        item,
        category=clean_text(row.get("CATEGORY")),
        subcategory=clean_text(row.get("SUBCATEGORY")),
        count_unit_quantity=optional_decimal(row.get("CU QTY")),
        count_unit_measure=clean_text(row.get("CU MEASURE")),
        custom_each_name=clean_text(row.get("CUSTOME EACH")),
        each_quantity=optional_decimal(row.get("EACH QTY")),
        each_measure=clean_text(row.get("EACH MEASURE")),
        weight_quantity=optional_decimal(row.get("WEIGHT QTY")),
        weight_measure=clean_text(row.get("WEIGHT MEASURE")),
        volume_quantity=optional_decimal(row.get("VOLUME QTY")),
        volume_measure=clean_text(row.get("VOLUME MEASURE")),
    )


def clean_text(value: object) -> str | None:
    if value is None:
        return None

    cleaned = str(value).strip()

    return cleaned or None


def normalize_header(value: object) -> str:
    if value is None:
        return ""

    return " ".join(str(value).strip().upper().split())


def optional_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, int | float):
        return Decimal(str(value))

    cleaned = str(value).strip()

    if not cleaned:
        return None

    cleaned = (
        cleaned
        .replace(",", "")
        .replace("$", "")
        .replace("€", "")
        .replace("£", "")
    )

    # Handle accounting-style negative values, e.g. "(12.50)"
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = f"-{cleaned[1:-1]}"

    try:
        return Decimal(cleaned)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid decimal value: {value!r}") from exc


def format_cell_value(value: object) -> object:
    if isinstance(value, Decimal):
        return str(value)

    return value


def is_blank_row(row: dict[str, Any]) -> bool:
    return all(value is None or str(value).strip() == "" for value in row.values())


if __name__ == "__main__":
    main()