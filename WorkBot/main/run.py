from backend.craftable_bot.CraftableBot import CraftableBot
import undetected_chromedriver as uc

from dotenv import load_dotenv
from os import getenv
from os import listdir, remove as os_remove, rename
from os.path import isfile, join

from backend.helpers.selenium_helpers import create_driver, create_options

from backend.helpers import  FormatItemData

from openpyxl import load_workbook, Workbook

# from backend.printing.Printer import Printer
from backend.emailer.Emailer import Emailer
from backend.emailer.Services.Service import Email
from backend.emailer.Services.Outlook import OutlookService
from backend.emailer.Services.Gmail import GmailService

# from backend.printing.Printer import Printer
from backend.transferring.TransferManager import TransferManager
from backend.transferring.Transfer import Transfer, TransferItem
from backend.pricing.PriceComparator import PriceComparator
from backend.ordering.OrderManager import OrderManager
from backend.ordering.Order import Order
from backend.vendors.VendorManager import VendorManager

# from backend.vendors.vendor_bots.USFoodsBot import USFoodsBot
from datetime import date
import json
from config.paths import VENDORS_DIR, ITEMS_DIR
from pathlib import Path
from pprint import pprint
import time
import re
import uuid 

from backend.workbot.WorkBot import WorkBot

dotenv = load_dotenv()

CRAFTABLE_USERNAME = getenv('CRAFTABLE_USERNAME')
CRAFTABLE_PASSWORD = getenv('CRAFTABLE_PASSWORD')

SOURCE_PATH = Path(__file__).parent / 'backend'

ORDER_FILES_PATH   = SOURCE_PATH / 'ordering' / 'OrderFiles'
PRICING_FILES_PATH = SOURCE_PATH / 'pricing'
DOWNLOAD_PATH      = SOURCE_PATH / 'downloads'
TRANSFER_PATH      = SOURCE_PATH / 'transferring'
DATABASE_PATH      = SOURCE_PATH / 'database'

def get_files(path: str) -> list:
	return [file for file in listdir(path) if isfile(join(path, file))]

def get_excel_files(path: Path) -> list[Path]:
	return [path / file for file in listdir(path) if isfile(join(path, file)) and file.endswith('.xlsx') and '~' not in file]



def produce_pricing_and_email(driver) -> None:

    vendors = [
        'Renzi',
        'Sysco',
        'Performance Food'
    ]

    for vendor in vendors:
        
        credentials = get_credentials(vendor)
        bot         = get_bot(vendor)(driver, credentials['username'], credentials['password'])

        for pricing_guide in 'IBProduce':
            file_name     = bot.retrieve_pricing_sheet(pricing_guide)
            new_file_name = f'{PRICING_FILES_PATH}{bot.name} {pricing_guide} {date.today()}.{file_name.split(".")[-1]}'

            rename(f'{DOWNLOAD_PATH}{file_name}', new_file_name)

    emailer             = Emailer(service=Outlook)
    recipients          = ('kitchen.ibctb@gmail.com', 'milesbrous@gmail.com', 'mimimehaffey@gmail.com')
    produce_sheet_path  = tuple([join(f'{PRICING_FILES_PATH}', 'ProducePricing.xlsx')])
    email               = Email(recipients, 'Produce Pricing', '', cc=None, attachments=produce_sheet_path)

    emailer.create_email(email)
    emailer.display_email(email)
    return

def download_pricing_sheets(driver, vendors=['Performance Food',], guides=['IBProduce']) -> None:

    for vendor in vendors:

        vendor_manager = VendorManager()

        bot = vendor_manager.initialize_vendor(vendor, driver)
        # print(bot, flush=True)
        for pricing_guide in guides:

            if vendor not in [
                 'US Foods', 
                 'Sysco', 
                 'Performance Food'
                ]:
                file_name = bot.retrieve_pricing_sheet(pricing_guide)
                # file_name = None
                new_file_name = PRICING_FILES_PATH / 'VendorSheets' / f'{bot.name}_{pricing_guide}_{date.today()}.{file_name.split(".")[1]}'

            elif vendor == 'US Foods':
                file_name = 'US Foods_IBProduce.csv'
                new_file_name = PRICING_FILES_PATH / 'VendorSheets' / f'US Foods_IBProduce_{date.today()}.csv'
            
            elif vendor == 'Sysco':
                file_name = 'Sysco_IBProduce.csv'
                new_file_name = PRICING_FILES_PATH / 'VendorSheets' / f'Sysco_IBProduce_{date.today()}.csv'
            
            elif vendor == 'Performance Food':
                file_name = 'Performance Food_IBProduce.xlsx'
                new_file_name = PRICING_FILES_PATH / 'VendorSheets' / f'Performance Food_IBProduce_{date.today()}.xlsx'

            new_file_name = PRICING_FILES_PATH / 'VendorSheets' / f'{bot.name}_{pricing_guide}_{date.today()}.{file_name.split(".")[1]}'

            
            # new_file_name = PRICING_FILES_PATH / 'VendorSheets' / 'US Foods_IBProduce_2025-03-10.csv'
            rename( DOWNLOAD_PATH / file_name, new_file_name)
            bot.format_vendor_pricing_sheet(new_file_name, new_file_name.with_suffix('.xlsx'))
        
    return

def generate_pricing_sheets(vendors=['Sysco', 'Performance Food', 'US Foods', 'Russo Produce',], guides=['IBProduce']):
        pricer = PriceComparator()
        # pricer.item_skus_file_path = f'{PRICING_FILES_PATH}\\ItemSkus.xlsx'
        for guide in guides:
            pricer.generate_pricing_sheet(f'{PRICING_FILES_PATH}\\Templates\\{guide}.xlsx', 
                                          [f'{PRICING_FILES_PATH}\\VendorSheets\\{vendor}_{guide}_{date.today()}.xlsx' for vendor in vendors], 
                                          f'{PRICING_FILES_PATH}\\Pricing Guides\\{guide}\\{guide} {date.today()}.xlsx')
        return

def delete_all_files_without_extension(directory: str, extension: str) -> None:
    for file in listdir(directory):
        if isfile(join(directory, file)) and not file.endswith(extension):
            os_remove(f'{directory}\\{file}')
    return




if __name__ == '__main__':

    vendors = [ 
        # 'Sysco', 
        # 'Performance Food',
        # 'US Foods',
        # 'Renzi',
        'UNFI',
        # 'Hill & Markes',
        # 'Johnston Paper',
        # 'Regional Distributors, Inc.',
        # 'Peters Supply',
        # 'SANICO',
        # 'Copper Horse Coffee',
        # 'Equal Exchange',
        # 'Eurocafe Imports',
        # 'Macro Mamas',
        # 'Coca-Cola',
        # 'FingerLakes Farms',
        # 'Ithaca Bakery',
        # 'Webstaurant',
        # 'Hillcrest Dairy',
        # 'Hillcrest Foods',
        # 'BakeMark',
        # 'Lentz',
        # 'Keck\'s Food Service',
        # 'Dawn',
        # 'Cortland Produce Inc.',
        # 'A.L. George',
        # 'BALKAN BEVERAGE LLC',
        # 'Casa',
        # 'Palmer',
        # 'ACE ENDICO',
        # 'Russo Produce',
        # 'BEHLOG & SON, INC.',
    ]

    stores = [
        #  'BAKERY',
         'TRIPHAMMER',
         'COLLEGETOWN',
         'EASTHILL',
         'DOWNTOWN'
    ]
    
    
    # gmail_service = GmailService()

    # emailer = Emailer(gmail_service)

    # test_email = Email(
    #     to=('andrew.ctb.ithaca@gmail.com',),
    #     subject='Test Email',
    #     body='This is a test email for the automated service.'
    # )

    # test_email_data = gmail_service.create_email(test_email)
    # gmail_service.display_email(test_email_data)

    # input('Press Enter to send email')
    # gmail_service.send_email(test_email_data)


    from backend.helpers.XLSXToJSON import xlsx_to_json_flat

    # xlsx_to_json_flat(DOWNLOAD_PATH / 'IthacaBakeryItems.xlsx', DOWNLOAD_PATH / 'IthacaBakeryItems.json')



    def overwrite_all_ids(file_path: str) -> None:
        """
        Overwrites the 'ID' field in every row of a JSON array with a new UUID.

        Args:
            file_path (str): Path to the JSON file. The file will be overwritten.
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        for item in data:
            item["ID"] = str(uuid.uuid4())

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

        print(f"✅ Overwrote IDs for {len(data)} items in: {file_path}")
    

    # overwrite_all_ids(ITEMS_DIR / 'IthacaBakeryItems.json')
    
    import json
    from pathlib import Path
    from uuid import uuid4

    def convert_to_item_first(input_path: str, output_path: str = "items_structured.json") -> None:
        """
        Converts a flat item/vendor JSON file into a structured item-first format
        with vendor information grouped under each item.

        Args:
            input_path (str): Path to the flat JSON file exported from Excel.
            output_path (str): Path to save the structured item-first JSON.
        """
        input_path = Path(input_path)
        output_path = Path(output_path)

        with open(input_path, "r", encoding="utf-8") as f:
            raw_items = json.load(f)

        structured_items = []

        for item in raw_items:
            item_id = item.get("ID") or str(uuid4())
            item_name = item.get("ITEM NAME", "").strip()

            vendors = []
            for i in range(1, 11):
                vendor = item.get(f"VENDOR NAME {i}")
                if not vendor:
                    continue

                vendor_info = {
                    "vendor": vendor,
                    "sku": item.get(f"SKU {i}"),
                    "unit": item.get(f"EDI ORDERING UOM {i}"),
                    "quantity": item.get(f"QTY {i}"),
                    "cost": item.get(f"COST (PU) {i}"),
                    "case_size": item.get(f"PU {i}")
                }
                vendors.append(vendor_info)

            structured_items.append({
                "id": item_id,
                "name": item_name,
                "vendors": vendors
            })

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(structured_items, f, indent=4, ensure_ascii=False)

        print(f"✅ Converted {len(structured_items)} items")
        print(f"📁 Saved to: {output_path}")


    # convert_to_item_first(ITEMS_DIR / 'IthacaBakeryItems.json', ITEMS_DIR / 'IthacaBakeryItemsITEMFIRST.json')
    
    '''Pricing Sheet Protocol'''
    options = create_options(DOWNLOAD_PATH)
    driver  = uc.Chrome(options=options, use_subprocess=True)
    download_pricing_sheets(driver)
    delete_all_files_without_extension(PRICING_FILES_PATH / 'VendorSheets', '.xlsx')
    input('Press ENTER to stop waiting.')
    generate_pricing_sheets()
    



    # NEED A WHITELIST OF ITEMS OTHERWISE ALL ~2500 ITEMS WILL BE LISTED
    # ANSWER: Use the par list and count only those items with a par above 0.
    # Use the OrderForm Excel file, it has the pars and ordered amount baked in.



    
    #{
    #   item_name: {store_name: {onhand: qty, par: par, ordered: 0}, store_name: quantity}
    #}

    def load_item_counts(files: list) -> dict:

        item_counts = {}
        filename_store_map = {
            'East Hill Plaza': 'Easthill',
            'Syracuse':        'Syracuse',
            'College Ave':     'Collegetown',
            'Triphammer Rd':   'Triphammer',
            'State St':        'Downtown'
        }

        for filename in filename_store_map:
            
            orderform_path = [i for i in files if filename in i.name][0]
        
            orderform_workbook = load_workbook(orderform_path)
            orderform_sheet = orderform_workbook.active

            store_name = filename_store_map[filename]

            for row in orderform_sheet.iter_rows(min_row=6):
                item_par = row[2].value
                if item_par <= 0: continue
                 
                item_name  = row[1].value
                order_qty  = row[5].value
                onhand_qty = row[3].value
                
                if item_name not in item_counts:
                    item_counts[item_name] = {store_name: {'onhand': onhand_qty, 'par': item_par, 'ordered': order_qty}}
                else:
                    if store_name not in item_counts[item_name]:
                        item_counts[item_name][store_name] = {'onhand': onhand_qty, 'par': item_par, 'ordered': order_qty}
        
        return item_counts
    # pprint(item_counts)

    def item_counts_to_excel_workbook(item_counts) -> Workbook:

        comparison_workbook = Workbook()
        sheet = comparison_workbook.active
        stores = []
        store_col_offset = 3
        for pos, item in enumerate(item_counts):
            item_name = item.split(' ')
            item_unit = item_name[-1]
            item_name = ' '.join(item_name[:-1])

            item_info = item_counts[item]
            sheet.cell(row=pos+2, column=1).value = item_name
            sheet.cell(row=pos+2, column=2).value = item_unit
            for store in item_info:
                onhand  = item_info[store]['onhand']
                ordered = item_info[store]['ordered']
                par     = item_info[store]['par']

                if store not in stores: stores.append(store)
                store_index = stores.index(store)

                store_col_pos = (store_index+1) * store_col_offset
                sheet.cell(row=1, column=store_col_pos).value = store
                sheet.cell(row=pos+2, column=store_col_pos).value = onhand
                sheet.cell(row=pos+2, column=store_col_pos+1).value = par
                sheet.cell(row=pos+2, column=store_col_pos+2).value = ordered

        return comparison_workbook


    # dir_path = DOWNLOAD_PATH / 'testing'
    # orderform_file_paths = [i for i in dir_path.iterdir() if i.is_file() and 'Order Form' in i.name]

    # item_counts = load_item_counts(orderform_file_paths)
    # item_count_workbook = item_counts_to_excel_workbook(item_counts)
    # item_count_workbook.save('testing.xlsx')


            

    # outlook_service = OutlookService()
    # outlook_service.refresh_inbox()
    # time.sleep(5)
    # # # print(outlook_service.get_recent_messages(subject_filter="Your US Foods one time password", max_age_minutes=60))
    # for message in outlook_service.get_recent_messages(
    #                                 subject_filter='Your US Foods one time passcode', 
    #                                 max_age_minutes=60):
    #     print(message)
    #     break
    
    # def extract_otp_from_body(body: str) -> str | None:
    #     match = re.search(r'\b\d{6}\b', body)
    #     if match:
    #         return match.group()
    #     return None

    # inbox = outlook_service.get_inbox()
    # messages = inbox.Items
    # messages.Sort("[ReceivedTime]", True)
    # for pos, i in enumerate(messages):
    #     if pos > 0: break
    #     print(i.Subject)
    #     # print(i.Body)
    #     otp = extract_otp_from_body(i.Body)
    #     print(otp)

    # vendor_manager = VendorManager()
    # us_foods_bot = vendor_manager.initialize_vendor('US Foods')
    # us_foods_bot.login()


        # 1. Open audit
        # 2. Extract necessary item counts
            # a. Name: Col2, Unit Count: Col4, Total Count: Col-3
        # 3. Update combine sheet
            # a. Use order combine function?
        # 4. Close audit



    # import sqlite3

    # def init_db(db_path=f'{DATABASE_PATH}/orders.db'):

    #     conn = sqlite3.connect(db_path)
    #     cursor = conn.cursor()

    #     cursor.execute('''
    #     CREATE TABLE IF NOT EXISTS stores (
    #                    id INTEGER PRIMARY KEY,
    #                    store_id INTEGER,
    #                    store_name TEXT UNIQUE,
    #                    )
                       
    #                    ''')