import json
from pathlib import Path
from openpyxl import load_workbook

def xlsx_to_json_flat(xlsx_path: str, output_path: str = None) -> None:
    """
    Convert an XLSX file to a JSON file using the first row as headers.

    Args:
        xlsx_path (str): Path to the Excel file.
        output_path (str, optional): Path to save the JSON output. Defaults to <input>.json.
    """
    xlsx_path = Path(xlsx_path)
    if output_path is None:
        output_path = xlsx_path.with_suffix('.json')
    else:
        output_path = Path(output_path)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"File not found: {xlsx_path}")

    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The Excel file is empty.")

    headers = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[0])]
    json_data = []

    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue  # skip empty rows
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        json_data.append(row_dict)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, indent=4, ensure_ascii=False)

    print(f"Saved JSON to: {output_path}")

