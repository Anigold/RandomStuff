from openpyxl import load_workbook
from os import listdir, remove
from os.path import isfile, join



def extract_item_data_from_excel_file(path: str) -> dict:

    item_data = {}

    workbook = load_workbook(path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        sku, name, quantity, cost_per, total_cost = row
        if sku.value not in item_data:
            item_data[sku.value] = {
                'name': name.value,
                'quantity': quantity.value,
                'cost_per': cost_per.value,
                'total_cost': total_cost.value,
            }
    
    return item_data

