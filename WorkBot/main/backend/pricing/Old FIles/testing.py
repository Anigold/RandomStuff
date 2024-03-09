from openpyxl import load_workbook, Workbook
from datetime import datetime
from pprint import pprint

def hoffmire_pricing(path_to_excel_file):
	
	workbook = load_workbook(path_to_excel_file)
	sheet = workbook.active

	items = {}
	for row in sheet.iter_rows(min_row=7):

		name 		  = row[3].value
		unit 		  = row[4].value


		purchase_date = row[9].value
		quantity 	  = row[11].value
		total_price   = row[12].value

		if not name or name == '':
			continue
			
		if name not in items:
			items[name] = {
				'SKU': name,
				'units': unit,
				'purchase_date': purchase_date,
				'quantity': quantity,
				'cost': total_price
			}
		else:
			new_month, new_day, new_year 			 = purchase_date.split('/') 
			current_month, current_day, current_year = items[name]['purchase_date'].split('/')
			
			if datetime(int(new_year), int(new_month), int(new_day)) > datetime(int(current_year), int(current_month), int(current_day)):
				items[name] = {
					'SKU': name,
					'units': unit,
					'purchase_date': purchase_date,
					'quantity': quantity,
					'cost': total_price
				}

	return items

if __name__ == '__main__':
	items = hoffmire_pricing('C:\\Users\\Will\\Desktop\\Andrew\\Projects\\Pricing\\VendorSheets\\Hoffmire Produce.xlsx')
	workbook = Workbook()

	sheet = workbook.active
	for pos, item_name in enumerate(items):

		name = item_name
		sku = items[item_name]['SKU']
		quantity = items[item_name]['quantity']
		units = items[item_name]['units']
		cost = items[item_name]['cost']

		sheet.cell(row=pos+1, column=1).value = item_name
		sheet.cell(row=pos+1, column=2).value = sku
		sheet.cell(row=pos+1, column=3).value = int(quantity)
		sheet.cell(row=pos+1, column=4).value = units

	workbook.save(filename='C:\\Users\\Will\\Desktop\\Andrew\\Projects\\Pricing\\VendorSheets\\Hoffmiresss.xlsx')
	pprint(items)