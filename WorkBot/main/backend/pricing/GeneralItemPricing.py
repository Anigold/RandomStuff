from PyPDF2 import PdfReader
import re
import pprint
from openpyxl import load_workbook
import xlrd
from os import listdir
from os.path import isfile, join
import shutil
from pathlib import Path
import csv
from datetime import datetime

class Vendor:

	def __init__(self, path_to_pricing_sheet: str, path_to_skus: str) -> None:
		self.path_to_pricing_sheet = path_to_pricing_sheet
		self.path_to_skus = path_to_skus

	def get_info(self) -> list:
		pass

	def get_skus_from_file(self) -> dict:
		workbook = load_workbook(self.path_to_skus)
		sheet = workbook.active

		skus = {}
		for pos, row in enumerate(sheet.iter_rows()):
			if pos < 2: continue

			sku = str(row[self.skus_col].value)

			if not sku: continue

			if sku not in skus:
				skus[sku] = pos+1
			
		return skus

	def update_skus_to_file(self) -> None:
		workbook = load_workbook(self.path_to_skus)
		sheet = workbook.active

		skus = self.get_skus_from_file()
		prices = self.get_info()
		
		for item in prices:
			item_info = prices[item]

			if item_info['SKU'] in skus:

				name     	  = item

				quantity 	  = float(item_info['quantity'])
				unit 	 	  = item_info['units']
				cost 	 	  = float(item_info['cost'])
				cost_per_unit = f'{round(cost/quantity, 2)} per {unit}'
				case_size     = item_info['case_size']

				sheet.cell(row=skus[item_info['SKU']], column=self.skus_col+2).value = cost_per_unit
				sheet.cell(row=skus[item_info['SKU']], column=self.skus_col+3).value = cost
				sheet.cell(row=skus[item_info['SKU']], column=self.skus_col+4).value = case_size

		workbook.save(self.path_to_skus)
		return

	def helper_iter_rows(sheet):
		for row in sheet.iter_rows():
			yield [cell.value for cell in row]

	def helper_format_size_units(pack: str, size: str) -> list:
		# Check that pack size is float compatible
		
		# if not isinstance(pack, float):
		# 	try:
		# 		pack = float(pack)
		# 	except:
		# 		raise Exception("Pack size is unable to be coerced to float.")

		# Split value and units from size
		unit_string = ''
		size_string = ''
		
		on_number = False
		numeric_non_numerics = ['-', '.']
		for pos, char in enumerate(size):
			
			if char == ' ':
				continue

			if not on_number and char.isnumeric():
				on_number = True
				size_string += char
				continue

			if not on_number and not char.isnumeric():
				if size_string == '': size_string = '1'
				unit_string += char
				continue

			if on_number and char.isnumeric():
				size_string += char
				continue

			if on_number and not char.isnumeric():
				
				
				if char in numeric_non_numerics:
					size_string += char
					continue

				if char == '#':
					unit_string = char
					break

				on_number = False
				unit_string += char
				continue

		if '-' in size_string:
			lower, upper = size_string.split('-')
			size_string  = (float(lower) + float(upper))/2
		else:
			size_string = float(size_string)
		
		return [size_string * float(pack), unit_string]

	def normalize_units(unit: str) -> str:
		unit = unit.upper()
		units = {
			'#': 'LB',
			'LB': 'LB',
			'CNT': 'EA',
			'EA': 'EA',
			'CT': 'EA',
			'DOZ': 'DZ',
			'PINT': 'PT'
		}

		return units[unit] if unit in units else unit


class Performance(Vendor):

	skus_col = 7

	special_cases = {
		'75104': {'unit': 'EA', 'pack': 13.5},
		'16146': {'unit': 'EA', 'pack': 42}
	}

	def get_info(self):
		workbook = load_workbook(self.path_to_pricing_sheet)
		sheet    = workbook.active

		row_info  = list(Vendor.helper_iter_rows(sheet))[9:-3] # We remove the "metadata" from the top and the ancillary information from the bottom of the sheet. 
		item_info = {}
		for row in row_info:
			item_sku  = row[4]
			item_name = row[1]


			'''
			This is the ugliest hack ever
			'''
			pack_size_info = row[5].split('/')
			pack_info      = pack_size_info[1].split(' ')
			
			if len(pack_info) == 2 and pack_info[1] == 'Bu':
				quantity = .9
				units = 'Bu'
			else:
				if item_sku in self.special_cases:
					quantity, units = Vendor.helper_format_size_units(self.special_cases[item_sku]['pack'], self.special_cases[item_sku]['unit'])
				else:
					quantity, units = Vendor.helper_format_size_units(pack_size_info[0], pack_size_info[1])

			cost = float(row[7].replace('$', ''))

			if item_name not in item_info:
				item_info[item_name] = {
					'SKU': item_sku,
					'quantity': quantity,
					'units': Vendor.normalize_units(units),
					'cost': cost,
					'case_size': row[5]

				}

		return item_info


class Renzi(Vendor):

	skus_col = 11

	special_cases = {
		'88076': {'unit': 'EA', 'pack': 1},
		'88055': {'unit': 'EA', 'pack': 36}
	}

	def get_info(self):
		item_info = {}
		with open(self.path_to_pricing_sheet) as file:
			for pos, row in enumerate(file):

				if pos == 0: continue

				row_info = row.split('\t')

				item_sku  		= row_info[0].split('"')[1]
				item_name 		= row_info[7]
				cost   	  		= float(row_info[8])
				
				if item_sku in self.special_cases:
					quantity, units = Vendor.helper_format_size_units(self.special_cases[item_sku]['pack'], self.special_cases[item_sku]['unit'])
				else:
					quantity, units = Vendor.helper_format_size_units(row_info[4], row_info[5])

			
				if item_name not in item_info:
					item_info[item_name] = {
						'SKU': item_sku,
						'quantity': quantity,
						'units': Vendor.normalize_units(units),
						'cost': cost,
						'case_size': f'{row_info[4]} / {row_info[5]}'
					}

		return item_info


class Sysco(Vendor):

	skus_col = 3

	def get_info(self):
		item_info = {}
		with open(self.path_to_pricing_sheet) as sysco_info:
			reader = csv.reader(sysco_info, delimiter=',')
			for pos, row in enumerate(reader):

				if pos < 2: continue

				item_sku 		= row[1]
				item_name 		= row[12]
				quantity, units = Vendor.helper_format_size_units(row[7], row[8])
				cost 			= row[14]

				if item_name not in item_info:
					item_info[item_name] = {
						'SKU': item_sku,
						'quantity': quantity,
						'units': Vendor.normalize_units(units),
						'cost': cost,
						'case_size': f'{row[7]} / {row[8]}'
					}

		return item_info


'''
Maps the vendor to its corresponding function.

Assumes the name to be in the following format:
	"[Vendor Name] Produce - mmddyyyy"

Args:
	name: string representation of the name of the file
	path: string representation of the absolute path to the file

Returns:
	The output of the respective vendor function
'''
def vendor_factory(name: str, path_to_pricing_sheet: str, path_to_skus: str):
	vendor_name = name.split(' ')[0]
	vendors = {
		'Sysco': Sysco,
		'Performance': Performance,
		'Renzi': Renzi,
	}
	if vendor_name not in vendors: return
	return vendors[vendor_name](path_to_pricing_sheet, path_to_skus)

def get_files(path: str) -> list:
	return [file for file in listdir(path) if isfile(join(path, file))]

def clear_old_data(path_to_pricing_sheet: str) -> None:
	pass

def price_compare(path_to_pricing_sheet: str) -> None:
	workbook = load_workbook(path_to_pricing_sheet)
	sheet = workbook.active

	vendors = []
	for pos, row in enumerate(sheet.iter_rows()):

		if pos == 0:
			excluded_values = ['Item', 'Buy From']
			for col in row:
				if col.value not in excluded_values and col.value:
					vendors.append(col.value)
			
			continue
		if pos in [1, 2]: continue			
	
		prices = []
		#unit_in_question = '' # Eventually make corrections based on differing units
		for i in range(3, 4*(len(vendors)), 4):
			price_string = sheet.cell(row=pos, column=i+1).value
			price = price_string.split(' per ')[0] if price_string else 100000
			prices.append(float(price))
	
		if prices: 
			min_price = min(prices)
			indices = [i for i, x in enumerate(prices) if x == min_price]

			# Refactor into string builder factory
			# ---------------------------------------- #
			for index in indices:
				if sheet.cell(row=pos, column=2).value:
					sheet.cell(row=pos, column=2).value = f'{sheet.cell(row=pos, column=2).value} or {vendors[index]} '
				else:
					sheet.cell(row=pos, column=2).value = f'{vendors[index]} '
			# ---------------------------------------- #

	workbook.save(path_to_pricing_sheet)
	return

if __name__ == '__main__':

	base_path = Path(__file__).parent.resolve()

	pricing_sheet_path		 = join(base_path, 'ProducePricing.xlsx')
	vendor_spreadsheets_path = join(base_path, 'VendorSheets')
	vendor_spreadsheets 	 = get_files(vendor_spreadsheets_path)

	for spreadsheet in vendor_spreadsheets:
		vendor_name   = spreadsheet.split(' ')[0]
		print(vendor_name)

		vendor_object = vendor_factory(vendor_name, f'{vendor_spreadsheets_path}/{spreadsheet}', pricing_sheet_path)
		
		if not vendor_object: continue
		
		vendor_object.update_skus_to_file()


	price_compare(pricing_sheet_path)
	