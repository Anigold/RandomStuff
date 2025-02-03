from ..vendor_bots.VendorBot import PricingBotMixin
from openpyxl import load_workbook, Workbook
import shutil
import pprint

class PriceSheetTemplate:
    
    def __init__(self) -> None:
        self.skus = []

class PriceComparator:

    def __init__(self) -> None:
        self.item_skus_file_path = './ItemSkus.xlsx'
    
    def get_skus_from_vendor_sheet(self, vendor_sheet_path) -> dict:
        vendor_workbook = load_workbook(vendor_sheet_path)
        sheet = vendor_workbook.active

        item_info = {}
        for row in sheet.iter_rows():
            name, sku, cost_per, case_cost, case_size = row

            if sku.value not in item_info:
                item_info[sku.value] = {
                    'name': name.value,
                    'cost_per': cost_per.value,
                    'case_cost': case_cost.value,
                    'case_size': case_size.value
                }

        return item_info
    
    def get_all_skus(self) -> dict:

        workbook = load_workbook(self.item_skus_file_path)
        sheet = workbook.active

        sku_info = {}
        vendors  = []
        top_row = list(sheet.rows)[0]
        for pos, val in enumerate(top_row):
            if not val.value: continue
            vendors.append(val.value)
            sku_info[val.value] = {}

        if not sku_info: return None

        
        for item_pos, row in enumerate(sheet.iter_rows(min_row=2)):
            item_name = sheet.cell(row=item_pos+1, column=1).value
         
            for vendor_col, vendor in enumerate(vendors):
                item_sku = sheet.cell(row=item_pos+1, column=vendor_col+2).value
                if item_sku and item_name:
                    if item_name not in sku_info[vendor]:
                        sku_info[vendor][item_name] = item_sku

        return sku_info
 
    def generate_pricing_sheet(self, pricing_sheet_template_path: str, vendor_sheets_paths: list[str], output_file_path: str) -> None:

        # Make copy of template and open
        shutil.copyfile(pricing_sheet_template_path, output_file_path)

        template_workbook = load_workbook(output_file_path)
        template_sheet    = template_workbook.active

        vendor_skus = {}
        vendors = []
        for vendor_path in vendor_sheets_paths:
            vendor_name = vendor_path.split('\\')[-1].split('_')[0]
   
            item_info = self.get_skus_from_vendor_sheet(vendor_path)
            if vendor_name not in vendor_skus:
                vendor_skus[vendor_name] = item_info
                vendors.append(vendor_name)
        

    
        first_row        = list(template_sheet.rows)[0]
        ignore_list      = ['Item', 'Preferred', 'Buy From', None]
        template_vendors = []
        template_offset  = lambda col, offset : (offset*col)+offset
        for column in first_row:
            if column.value not in ignore_list:
                template_vendors.append(column.value.strip())

        for pos, row in enumerate(template_sheet.iter_rows(min_row=3)):
            for vendor_pos, vendor in enumerate(template_vendors):
                sku = str(row[template_offset(vendor_pos, 4)].value)
                if (vendor in vendor_skus) and (sku in vendor_skus[vendor]): 
                    template_sheet.cell(row=pos+3, column=template_offset(vendor_pos, 4)+2).value = vendor_skus[vendor][sku]['cost_per']
                    template_sheet.cell(row=pos+3, column=template_offset(vendor_pos, 4)+3).value = vendor_skus[vendor][sku]['case_cost']
                    template_sheet.cell(row=pos+3, column=template_offset(vendor_pos, 4)+4).value = vendor_skus[vendor][sku]['case_size']
        
        
        template_workbook.save(output_file_path)

        return self.compare_prices(output_file_path)

    def compare_prices(self, path_to_pricing_sheet: str) -> None:
        workbook = load_workbook(path_to_pricing_sheet)
        sheet = workbook.active

        vendors = []
        for pos, row in enumerate(sheet.iter_rows()):

            if pos == 0:
                excluded_values = ['Item', 'Buy From', 'Preferred', None]
                for col in row:
                    if col.value not in excluded_values and col.value:
                        vendors.append(col.value)
                print(vendors)
                continue
            
            if pos in [1, 2]: continue			
            
            prices = []
            #unit_in_question = '' # Eventually make corrections based on differing units
            for i in range(5, 4*(len(vendors)+1), 4):
                price_string = sheet.cell(row=pos, column=i+1).value
                price 		 = price_string.split(' per ')[0] if price_string else 100000
                prices.append(float(price))
            print(pos, prices)
            if prices: 
                min_price = min(prices)
                if min_price == 100000: continue
                indices = [i for i, x in enumerate(prices) if x == min_price]

                # Refactor into string builder factory
                # ---------------------------------------- #
                for index in indices:
                    if sheet.cell(row=pos, column=4).value:
                        sheet.cell(row=pos, column=4).value = f'{sheet.cell(row=pos, column=4).value} or {vendors[index]} '
                    else:
                        sheet.cell(row=pos, column=4).value = f'{vendors[index]} '
                # ---------------------------------------- #

        workbook.save(path_to_pricing_sheet)
        return

    def find_price_fluctuations(self, path):
        pass