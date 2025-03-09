from .VendorBot import VendorBot, SeleniumBotMixin, PricingBotMixin
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium import webdriver
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from csv import writer
from datetime import date
from pprint import pprint

class PerformanceFoodBot(VendorBot, SeleniumBotMixin, PricingBotMixin):

    def __init__(self, driver: webdriver = None, username: str = None, password: str = None) -> None:
        VendorBot.__init__(self)
        PricingBotMixin.__init__(self)
        SeleniumBotMixin.__init__(self, driver, username, password)
        self.name = "Performance Food"
        self.minimum_order_case = 20

        self.store_ids = {}

        self.special_cases = {
            '75104': {'unit': 'EA', 'pack': 13.5},
            '16146': {'unit': 'EA', 'pack': 42}
        }

        

    def login(self) -> None:
        self.driver.get('https://www.customerfirstsolutions.com')

        time.sleep(4)

        username_input = self.driver.find_element(By.ID, 'signInName')
        password_input = self.driver.find_element(By.ID, 'password')

        username_input.send_keys(self.username)
        password_input.send_keys(self.password)

        submit_button = self.driver.find_element(By.ID, 'next')
        submit_button.click()
        
        time.sleep(5)
        self.is_logged_in = True
        return

    def logout(self) -> None:
        pass
    
    def switch_store(self, store_id: str) -> None:
        pass

    def format_for_file_upload(self, item_data: dict, path_to_save: str, store: str) -> None:
        
        with open(f'{path_to_save}.txt', 'w', newline='') as csv_file:

            csv_writer = writer(csv_file, delimiter='\t')

            for item in item_data:
                
                quantity = item['quantity']
                sku      = item['sku']

                csv_writer.writerow([sku, int(float(quantity)), 'CS'])

        return
    
    def retrieve_pricing_sheet(self, guide_name: str) -> str:

        if not self.is_logged_in:
            self.login()
       
        # Ensure we're on the Ithaca Bakery location
        current_location = self.driver.find_element(By.CLASS_NAME, 'MuiBox-root')
        current_location_name = current_location.find_element(By.XPATH, '//span[2]')
    
        if current_location_name.text != 'Ithaca Bakery':
            time.sleep(3)
            current_location_name.click()

            time.sleep(3)

            location_dropdown = self.driver.find_element(By.XPATH, './/div[@data-testid="location-menu"]')
            location_input = location_dropdown.find_element(By.XPATH, './/input[@data-testid="location-search-menu-search"]')
            location_input.send_keys('bakery')
            location_input.send_keys(Keys.ENTER)

            time.sleep(2)

            location = location_dropdown.find_element(By.XPATH, './/span[@data-testid="location-menu-customer-name"]')
            location.click()

            time.sleep(5)

        lists_button = self.driver.find_element(By.XPATH, '//span[@data-testid="nav-item-Lists"]')
        lists_button.click()
        time.sleep(5)

        lists = self.driver.find_elements(By.XPATH, '//h5[@data-testid="product-list-management-card-name"]')
        selected_list = [i for i in lists if guide_name.upper() == i.text][0]
        selected_list.click()
        time.sleep(2)

        export_options_button = self.driver.find_element(By.XPATH, '//button[@data-testid="product-list-management-detail-menu-btn"]')
        export_options_button.click()

        time.sleep(1)

        export_button = self.driver.find_element(By.XPATH, '//div[@data-testid="product-list-menu-export-list"]')
        export_button.click()

        time.sleep(3)

        export_report_button = self.driver.find_element(By.XPATH, '//button[@data-testid="export-report-btn"]')
        export_report_button.click()

        time.sleep(8)
        today = date.today()
        today_mm_dd_yyyy = today.strftime('%m-%d-%Y')

        return f'{guide_name}_RFS Eastern PA-07110_{today_mm_dd_yyyy}.xlsx'
    
    def get_pricing_info_from_sheet(self, path_to_pricing_sheet: str) -> dict:
        workbook = load_workbook(path_to_pricing_sheet)
        sheet    = workbook.active

        row_info  = list(PricingBotMixin.helper_iter_rows(sheet))[9:-3] # We remove the "metadata" from the top and the ancillary information from the bottom of the sheet. 
        item_info = {}
        for row in row_info:
            print(row, flush=True)
            item_sku  = row[5]
            item_name = row[1]


            '''
            This is the ugliest hack ever
            '''
            if not row[6]: continue
            pack_size_info = row[6].split('/')
            print(pack_size_info, flush=True)
            pack_info      = pack_size_info[1].split(' ')
            
            if len(pack_info) == 2 and pack_info[1] == 'Bu':
                quantity = .9
                units = 'Bu'
            else:
                if item_sku in self.special_cases:
                    quantity, units = PricingBotMixin.helper_format_size_units(self.special_cases[item_sku]['pack'], self.special_cases[item_sku]['unit'])
                else:
                    quantity, units = PricingBotMixin.helper_format_size_units(pack_size_info[0], pack_size_info[1])

            cost = float(row[8].replace('$', ''))

            if item_name not in item_info:
                item_info[item_name] = {
                    'SKU': item_sku,
                    'quantity': quantity,
                    'units': PricingBotMixin.normalize_units(units),
                    'cost': cost,
                    'case_size': row[6]

                }

        return item_info