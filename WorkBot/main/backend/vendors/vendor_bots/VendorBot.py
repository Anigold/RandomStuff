from abc import ABC, abstractmethod
from openpyxl import load_workbook, Workbook
import re
from pprint import pprint

class VendorBot(ABC):

    def __init__(self) -> None:
        self.minimum_order_amount = 0 # By cents
        self.minimum_order_case   = 0 # By each

    @abstractmethod
    def format_for_file_upload(self, item_data: dict, path_to_save: str, store: str = None):
        pass

    
class SeleniumBotMixin(ABC):

    def __init__(self, driver, username, password) -> None:
        self.is_logged_in = False
        self.driver       = driver
        self.username     = username
        self.password     = password

    def login(self) -> None:
        pass

    def logout(self) -> None:
        pass

    def end_session(self) -> None:
        return self.driver.close()
    

class PricingBotMixin(ABC):

    UNIT_NORMALIZATION = {
        "#": "LB",
        "LB": "LB",
        "LBS": "LB",
        "OZ": "OZ",
        "GAL": "GAL",
        "GALLON": "GAL",
        "EA": "EA",
        "EACH": "EA",
        "CT": "EA",
        "CNT": "EA",
        "DOZ": "DZ",
        "PT": "PT",
        "QT": "QT",
        "HD": "EA",
        "HEAD": "EA",
        "CS": "CS",
        "CASE": "CS",
    }

    PACK_SIZE_PATTERN = re.compile(
        r"""
        ^(?:                   # Start match
            (?P<packs>\d+)[/x]\s*  # Pack multiplier (e.g., "4/", "2x") (optional)
        )?
        (?P<low>(?:\d+)?(?:\.\d+)?|\d+)    # Main quantity (now supports ".5")
        (?:-(?P<high>(?:\d+)?(?:\.\d+)?|\d+))?  # Optional range (also supports ".5")
        \s*
        (?P<unit>[a-zA-Z#]+)   # Unit of measure (e.g., "LB", "EA", "QT")
        """,
        re.VERBOSE,
    )

    def __init__(self) -> None:
        self.special_cases = {}
    
    def get_pricing_info_from_sheet(self, path_to_pricing_sheet: str) -> dict:
        pass
    
    def get_skus_from_file(self, path_to_skus: str) -> dict:
        pass

    def normalize_units(unit: str) -> str:

        if not unit: return unit

        unit = unit.upper()
        units = {
			'#': 'LB',
			'LB': 'LB',
			'CNT': 'EA',
			'EA': 'EA',
			'CT': 'EA',
			'DOZ': 'DZ',
			'PINT': 'PT',
            'HD': 'EA',
            'HEAD': 'EA'
		}

        return units[unit] if unit in units else unit
    
    def helper_iter_rows(sheet):
        for row in sheet.iter_rows():
            yield [cell.value for cell in row]

    def get_skus_from_file(self, path_to_skus) -> dict:
        workbook = load_workbook(path_to_skus)
        sheet = workbook.active

        skus = {}
        for pos, row in enumerate(sheet.iter_rows()):
            if pos < 2: continue

            sku = str(row[self.skus_col].value)

            if not sku: continue

            if sku not in skus:
                skus[sku] = pos+1
            
        return skus
    
    def format_vendor_pricing_sheet(self, pricing_sheet_path: str, save_path: str) -> None:
        price_info = self.get_pricing_info_from_sheet(pricing_sheet_path)

        # print('creating workbook',  flush=True)
        workbook = Workbook()
        sheet = workbook.active

        # print('importing pricing data', flush=True)
        for row_pos, item_name in enumerate(price_info):

          
            item_info     = price_info[item_name]
            print(item_name)
            pprint(item_info)
            sku           = item_info['SKU']
            cost          = item_info['cost']
            case_size     = item_info['case_size']
            quantity 	  = float(item_info['quantity'])
            unit 	 	  = item_info['units']
            cost_per_unit = f'{round(cost/quantity, 2)} per {unit}'
			
            row = [item_name, sku, cost_per_unit, cost, case_size]
            for col_position, col_value in enumerate(row):
                sheet.cell(row=row_pos+1, column=col_position+1).value = col_value

        # print('Saving workbook', flush=True)
        # print(save_path, flush=True)
        return workbook.save(filename=f'{save_path}')
    
    @classmethod
    def helper_format_size_units(cls, pack_size_str):


        # Expected pack strings:

        #   pack/size unit (12/6 oz)
        #   low-high unit (5-6 EA)
        #   size unit (5 LB)
        #

        if not pack_size_str:
            return None, None
        
        # Ensure space exists between number and unit (e.g., "10Gal" → "10 Gal")
        pack_size_str = re.sub(r"(?<=\d)(?=[A-Za-z#])", " ", pack_size_str)

        match = cls.PACK_SIZE_PATTERN.search(pack_size_str)
        if not match:
            return None, None  # Return None if no valid match is found

        packs = int(match.group("packs")) if match.group("packs") else 1
        low   = float(match.group("low")) if match.group("low") else 1
        high  = float(match.group("high")) if match.group("high") else low
        unit  = match.group("unit").upper()

        # Normalize the unit
        unit = cls.UNIT_NORMALIZATION.get(unit, unit)
        
        # Compute final quantity (average if range is given)
        avg_size = (low + high) / 2
        total_quantity = avg_size * packs

        # print(f'{pack_size_str} -> ({total_quantity}, {unit})')
        return total_quantity, unit

    # def helper_format_size_units(pack: str, size: str) -> list:
	# 	# Check that pack size is float compatible
		
	# 	# if not isinstance(pack, float):
	# 	# 	try:
	# 	# 		pack = float(pack)
	# 	# 	except:
	# 	# 		raise Exception("Pack size is unable to be coerced to float.")

	# 	# Split value and units from size
    #     unit_string = ''
    #     size_string = ''
		
    #     on_number = False
    #     numeric_non_numerics = ['-', '.']
    #     for pos, char in enumerate(size):
			
    #         if char == ' ':
    #             continue

    #         if not on_number and char.isnumeric():
    #             on_number = True
    #             size_string += char
    #             continue

    #         if not on_number and not char.isnumeric():
    #             if size_string == '': size_string = '1'
    #             unit_string += char
    #             continue

    #         if on_number and char.isnumeric():
    #             size_string += char
    #             continue

    #         if on_number and not char.isnumeric():
                
				
    #             if char in numeric_non_numerics:
    #                 size_string += char
    #                 continue

    #             if char == '#':
    #                 unit_string = char
    #                 break

    #             on_number = False
    #             unit_string += char
    #             continue

    #     if '-' in size_string:
    #         lower, upper = size_string.split('-')
    #         size_string  = (float(lower) + float(upper))/2
    #     else:
    #         size_string = float(size_string)

    #     return [size_string * float(pack), unit_string]

    @abstractmethod
    def retrieve_pricing_sheet(self) -> None:
        pass

    def _special_case_info(self, unit: str, pack: float) -> dict:
        return {'unit': unit, 'pack': pack}
    