from .VendorBot import VendorBot, SeleniumBotMixin, PricingBotMixin
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium import webdriver
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from csv import writer
from datetime import date
from pprint import pprint

class BehlogProduceBot(VendorBot, PricingBotMixin):

    def __init__(self, **kwargs) -> None:
        VendorBot.__init__(self)
        PricingBotMixin.__init__(self)
        self.name = 'Behlog Produce'
        self.minimum_order_case = 5

        self.store_ids = {}

        self.special_cases = {
            '1425': self._special_case_info('EA', 1), # Honeydew
            '1081': self._special_case_info('EA', 109), # Empire Apple
            '253': self._special_case_info('EA', 23), # Cilantro
            '176': self._special_case_info('EA', 15), # Collard Greens
            '1305': self._special_case_info('EA', 34.5), # Kiwi
            '1150': self._special_case_info('LB', 8), # Strawberry
            '340': self._special_case_info('EA', 24), # Romaine Heart (Whole Thing)
            '347': self._special_case_info('LB', 36), # Romaine Heart (Just the heart)
            '404': self._special_case_info('LB', 3) # Portobello Mushroom Cap

        }

    def format_for_file_upload(self, item_data: dict, path_to_save: str):
        pass
    
    def get_pricing_info_from_sheet(self, path_to_pricing_sheet: str) -> dict:

        workbook = load_workbook(path_to_pricing_sheet)
        sheet    = workbook.active

        row_info  = list(PricingBotMixin.helper_iter_rows(sheet))[1:]
        item_info = {}

        numeric_non_numerics = ['.', '-', '/']
        for row in row_info:

            if (row[0] == None) or (row[1] == None) or (row[2] == None) or (row[3] == None) or (not row[3]): continue

            item_sku  = str(row[1])
            item_name = str(row[0])
            if not item_name: continue
            pack_size = ''
            unit = ''
            if item_sku not in self.special_cases:

                pack_size_info = row[2]
                
                # Extract units
                for char in pack_size_info:
                    if char.isnumeric() or char in numeric_non_numerics:
                        pack_size += char

                    else: unit += char

                # Average if necessary
                if '-' in pack_size:
                    lower, upper = pack_size.split('-')
                    pack_size = ( int(lower) + int(upper) ) / 2

    
                # Combine if necessary
                elif '/' in pack_size:
                    packs = pack_size.split('/')[0] or 1
                    size = pack_size.split('/')[1] or 1

                    pack_size = ( float(packs) * float(size) )

            else:
                print(item_name)
                unit = self.special_cases[item_sku]['unit']
                pack_size = self.special_cases[item_sku]['pack']

            if row[3] == '' or row[3] == ' ' or row[3]=='sp': continue
            cost = float(row[3])
            if pack_size == '': pack_size = 1
            if item_name not in item_info:
                item_info[item_name] = {
                    'SKU': item_sku,
                    'quantity': pack_size,
                    'units': PricingBotMixin.normalize_units(unit),
                    'cost': cost,
                    'case_size': row[2]
                }
            

        return item_info
    
    def retrieve_pricing_sheet(self, pricing_guide) -> None:
        return 'Behlog Produce_IBProduce.xlsx'