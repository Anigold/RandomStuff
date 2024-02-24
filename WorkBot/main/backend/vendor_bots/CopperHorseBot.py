from .VendorBot import VendorBot
import os
from pprint import pprint
from openpyxl import Workbook, load_workbook

class CopperHorseBot(VendorBot):

    def __init__(self) -> None:
        super().__init__()
        self.name = 'Copper Horse Coffee'

    def format_for_file_upload(self, item_data: dict, path_to_save: str) -> None:
        # CSV-style Excel file with "Item Code, Quantity"
        workbook = Workbook()
        sheet = workbook.active
        pprint(item_data)
        for pos, sku in enumerate(item_data):
            quantity = item_data[sku]

            sheet.cell(row=pos+1, column=1).value = int(sku)
            sheet.cell(row=pos+1, column=2).value = int(quantity)
        
        workbook.save(filename=f'{path_to_save}.xlsx')

    '''
    Takes the pre-formatted orders and merges them to one sheet.

    Args:
        path_to_orders: string representation of the path to the folder containing all the orders
    '''
    def combine_orders(self, path_to_orders: str) -> None:
        orders = []
        for file in os.listdir(path_to_orders):
            if file.endswith('.xlsx'):
                orders.append(file)

        workbook = Workbook()
        sheet = workbook.active
        for pos, order in enumerate(orders):
            current_workbook = load_workbook(f'{path_to_orders}\\{order}')
            current_sheet = current_workbook.active
            order_info = current_sheet.values
            pprint(list(order_info))
