from .VendorBot import VendorBot
import os
from pprint import pprint
from openpyxl import Workbook, load_workbook

class DutchValleyBot(VendorBot):

    def __init__(self, driver, username, password) -> None:
        super().__init__(driver, username, password)
        
        self.name                 = 'DUTCH VALLEY FOOD DIST'
        self.minimum_order_amount = 1_000_00 # $1000 in cents


    def format_for_file_upload(self, item_data: dict, path_to_save: str) -> None:
        # CSV-style Excel file with "Item Code, Quantity"
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1).value = 'sku'
        sheet.cell(row=1, column=2).value = 'qty'
        for pos, sku in enumerate(item_data):
            quantity = item_data[sku]['quantity']
        
            sheet.cell(row=pos+2, column=1).value = sku
            sheet.cell(row=pos+2, column=2).value = int(quantity)
        
        workbook.save(filename=f'{path_to_save}.xlsx')

   