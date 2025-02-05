from .VendorBot import VendorBot
import os
from pprint import pprint
from openpyxl import Workbook, load_workbook
import ntpath

class IthacaBakeryBot(VendorBot):

    def __init__(self) -> None:
        super().__init__()
        
        self.name = 'Ithaca Bakery'


    def format_for_file_upload(self, item_data: dict, path_to_save: str) -> None:
        # CSV-style Excel file with "Item Code, Quantity"
        workbook = Workbook()
        sheet = workbook.active

        for pos, name in enumerate(item_data):
            quantity = item_data[name]['quantity']
            sheet.cell(row=pos+1, column=1).value = True
            sheet.cell(row=pos+1, column=2).value = name
            sheet.cell(row=pos+1, column=3).value = float(quantity)
        
        workbook.save(filename=f'{path_to_save}.xlsx')

    '''
    File name format:

    {MAIN PATH} / {vendor} _ {store} {date}.xlsx

    '''
    def get_store_from_file_name(self, file_name: str) -> str:
        #return ntpath.basename(file_name)
        return file_name.split(' _ ')[1].split(' ')[0]
        
    '''
    Takes the pre-formatted orders and merges them to one sheet.

    Args:
        path_to_orders: string representation of the path to the folder containing all the orders
    '''
    def combine_orders(self, path_to_orders: list, path_to_save: str) -> None:
        
        combined_orders = {}
        '''
        {
            item: {
                "COLLEGETOWN": 3, 
                "DOWNTOWN": 1,
                "EASTHILL": 4,
                "TRIPHAMMER": 1
            },
            .
            .
            .
        }
        '''

        for order_file in path_to_orders:
            
            store_name = self.get_store_from_file_name(order_file)
            print(store_name)
            order_book = load_workbook(order_file)
            order_sheet = order_book.active

            for row in order_sheet.iter_rows(min_row=2):
                name = row[1].value.split('\n')[0]
                quantity = float(row[2].value)

                if name not in combined_orders:
                    combined_orders[name] = {store_name: quantity}
                else:
                    if store_name not in combined_orders[name]:
                        combined_orders[name][store_name] = quantity
                    else:
                        combined_orders[name][store_name] += quantity # This shouldn't be hit, but is here just in case.

        combined_orders_book  = Workbook()
        combined_orders_sheet = combined_orders_book.active
        combined_orders_sheet.cell(row=1, column=1).value = 'Item'
        combined_orders_sheet.cell(row=1, column=2).value = 'COLLEGETOWN'
        combined_orders_sheet.cell(row=1, column=3).value = 'TRIPHAMMER'
        combined_orders_sheet.cell(row=1, column=4).value = 'EASTHILL'
        combined_orders_sheet.cell(row=1, column=5).value = 'DOWNTOWN'

        for pos, item in enumerate(combined_orders):
  
            combined_orders_sheet.cell(row=pos+2, column=1).value = item
            combined_orders_sheet.cell(row=pos+2, column=2).value = combined_orders[item]['COLLEGETOWN'] if 'COLLEGETOWN' in combined_orders[item] else ''
            combined_orders_sheet.cell(row=pos+2, column=3).value = combined_orders[item]['TRIPHAMMER'] if 'TRIPHAMMER' in combined_orders[item] else ''
            combined_orders_sheet.cell(row=pos+2, column=4).value = combined_orders[item]['EASTHILL'] if 'EASTHILL' in combined_orders[item] else ''
            combined_orders_sheet.cell(row=pos+2, column=5).value = combined_orders[item]['DOWNTOWN'] if 'DOWNTOWN' in combined_orders[item] else ''

        combined_orders_book.save(f'{path_to_save}\\combined_order.xlsx')

        return
    