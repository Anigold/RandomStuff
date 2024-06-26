from .VendorBot import VendorBot, SeleniumBotMixin, PricingBotMixin
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium import webdriver
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from csv import writer
from datetime import date
from pprint import pprint

class RussoProduceBot(VendorBot, PricingBotMixin):

    def __init__(self, **kwargs) -> None:
        VendorBot.__init__(self)
        PricingBotMixin.__init__(self)
        self.name = 'Russo Produce'
        self.minimum_order_case = 5

        self.store_ids = {}

        self.special_cases = {
            'Alfalfa Sprouts': {'unit': 'LB', 'pack': 2.25},
            'Bananas': {'unit': 'LB', 'pack': 40},
            'Basil (Cleaned)': {'unit': 'LB', 'pack': .15625},
            'Brussel Sprouts': {'unit': 'LB', 'pack': 25},
            'Cabbage Shredded': {'unit': 'LB', 'pack': 25},
            'Cauliflower': {'unit': 'EA', 'pack': 12},
            'Cilantro': {'unit': 'EA', 'pack': 1},
            'Collard Greens': {'unit': 'EA', 'pack': 9},
            'Cucumbers English': {'unit': 'EA', 'pack': 12},
            'Garlic Peeled': {'unit': 'LB', 'pack': 5},
            'Ginger Root': {'unit': 'LB', 'pack': 1},
            'Grapes Seedless': {'unit': 'LB', 'pack': 18},
            'Honeydew Melon': {'unit': 'EA', 'pack': 5.5},
            'Kale': {'unit': 'EA', 'pack': 24},
            'Kiwi': {'unit': 'EA', 'pack': 1},
            'Lemon': {'unit': 'EA', 'pack': 140},
            'Limes': {'unit': 'LB', 'pack': 10},
            'Mushrooms Thick Sliced': {'unit': 'LB', 'pack': 10},
            'Mushrooms Whole Button': {'unit': 'LB', 'pack': 10},
            'Onion Red': {'unit': 'LB', 'pack': 25},
            'Onion Spanish': {'unit': 'LB', 'pack': 50},
            'Pepper Green Bell (Large) (RENZI = 20lb)': {'unit': 'LB', 'pack': 25},
            'Pepper Green Bell Xtra Large': {'unit': 'LB', 'pack': 25},
            'Pepper Red Bell (11 lb)': {'unit': 'LB', 'pack': 11},
            'Peppers Jalapeno': {'unit': 'LB', 'pack': 1},
            'Pineapple': self._special_case_info('EA', 7.5),
            'Spinach Baby': self._special_case_info('LB', 10),
            'Strawberries Fresh': self._special_case_info('LB', 8),
            'Tomatos 5x6': self._special_case_info('LB', 25),
        }

    def format_for_file_upload(self, item_data: dict, path_to_save: str):
        pass
    
    def get_pricing_info_from_sheet(self, path_to_pricing_sheet: str) -> dict:
        workbook = load_workbook(path_to_pricing_sheet)
        sheet    = workbook.active

        row_info  = list(PricingBotMixin.helper_iter_rows(sheet))[1:]
        item_info = {}

        numeric_non_numerics = ['.', '-', '/']
        for row in row_info:
            item_sku  = str(row[0])
            item_name = str(row[0])

            pack_size = ''
            unit = ''
            if item_sku not in self.special_cases:
                continue
                print(item_name)
                pack_size_info = row[2]
                
                # Extract units
                for char in pack_size_info:
                    if char.isnumeric() or char in numeric_non_numerics:
                        pack_size += char

                    else: unit += char

                # Average if necessary
                if '-' in pack_size:
                    lower, upper = pack_size.split('-')
                    pack_size = ( int(lower) + int(upper) ) / 2

    
                # Combine if necessary
                elif '/' in pack_size:
                    packs = pack_size.split('/')[0] or 1
                    size = pack_size.split('/')[1] or 1

                    pack_size = ( float(packs) * float(size) )

            else:
                unit = self.special_cases[item_sku]['unit']
                pack_size = self.special_cases[item_sku]['pack']

            cost = float(row[1])

            if item_name not in item_info:
                item_info[item_name] = {
                    'SKU': item_sku,
                    'quantity': pack_size,
                    'units': PricingBotMixin.normalize_units(unit),
                    'cost': cost,
                    'case_size': row[2]
                }
            

        return item_info
    
    def retrieve_pricing_sheet(self, pricing_guide) -> None:
        return 'Russo Produce_IBProduce.xlsx'