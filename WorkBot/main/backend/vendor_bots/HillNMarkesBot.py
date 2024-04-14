from .VendorBot import VendorBot, SeleniumBotMixin
from selenium.webdriver.common.by import By
import time
from pprint import pprint
from openpyxl import load_workbook, Workbook
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

class HillNMarkesBot(VendorBot, SeleniumBotMixin):

    def __init__(self):
        super().__init__()
        
        self.name                 = "Hill & Markes"
        #self.order_manager       = order_manager
        self.minimum_order_amount = 300_00 # $300 in cents

        self.store_ids = {
            'BAKERY': 0,
            'COLLEGETOWN': 1,
            'TRIPHAMMER': 4,
            'EASTHILL': 6,
            'DOWNTOWN': 9
        }

    def login(self) -> None:
	
        self.driver.get('https://www.hillnmarkes.com/Welcome.action')
        time.sleep(5)

        print('Finding login dropdown...')
        login_dropdown_button = self.driver.find_elements(By.CLASS_NAME, 'dropdown-toggle')
        login_dropdown_button[0].click()
        print('...login dropdown found.')

        print('')
        time.sleep(2)

        print('Finding login form...')
        login_form    = self.driver.find_element(By.ID, 'popLoginForm')
        username_form = self.driver.find_element(By.ID, 'popUserName')
        password_form = self.driver.find_element(By.ID, 'popPassword')
        print('...login form found.')

        print('')

        print('Sending login information...')
        username_form.send_keys(self.username)
        password_form.send_keys(self.password)
        login_form.find_element(By.XPATH, './/button[@type="submit"]').click()
        print('...credentials sent.')

        time.sleep(10)
        print('')

        print('Selecting store...')
        store_selection_table = self.driver.find_element(By.ID, 'example')
        store_rows            = store_selection_table.find_elements(By.XPATH, './/tbody/tr')

        # Choosing IB for testing
        store_rows[0].find_element(By.XPATH, './/td').click()
        time.sleep(10)
        print('...store selected.')

        print('')

        print('Waiting for login success...')

        return

    def go_to_quick_cart_file_upload(self) -> None:
        self.driver.get('https://www.hillnmarkes.com/QuickCartStandard')
        return
        
    def upload_quick_cart_file(self, file_to_upload: str) -> None:

        if self.driver.current_url != 'https://www.hillnmarkes.com/QuickCartStandard':
            self.go_to_quick_cart_file_upload()
            time.sleep(5)
            return self.upload_quick_cart_file(file_to_upload)

        quick_order_tab = self.driver.find_element(By.ID, 'quickOrderTab')
        file_upload_tab = quick_order_tab.find_elements(By.XPATH, './/ul/li')[2]
        file_upload_tab.click()

        time.sleep(2)

        file_upload_form  = self.driver.find_element(By.ID, 'uploadForm')
        file_upload_input = self.driver.find_element(By.ID, 'datafile')
        file_upload_input.send_keys(file_to_upload)
        time.sleep(5)
        file_submit = file_upload_form.find_element(By.XPATH, './input[@value="Upload"]')

        file_submit.click()

        # We have to wait for items to show beneatch the upload.
        # Sometimes takes up to a minute.
        WebDriverWait(self.driver, 300).until(EC.element_to_be_clickable((By.ID, "itemsInCart")))
 
        return

    def switch_store(self, store: str) -> None:

        if store not in self.store_ids:
            return
        
        my_account_button = self.driver.find_element(By.XPATH, './/a[@data-target="myAccountMenu"]')
        my_account_button.click()
        time.sleep(3)

        my_account_menu       = self.driver.find_element(By.CLASS_NAME, 'myAccountMenu')
        account_menu_choices  = my_account_menu.find_elements(By.TAG_NAME, 'li')
        change_account_button = account_menu_choices[2]
        change_account_button.click()
        time.sleep(5)

        store_change_table = self.driver.find_element(By.ID, 'example')
        store_change_rows  = store_change_table.find_elements(By.TAG_NAME, 'tr')
        store_to_change_to = store_change_rows[self.store_ids[store]]

        row_data                = store_to_change_to.find_elements(By.TAG_NAME, 'td')
        change_to_button_column = row_data[0]
        change_to_button        = change_to_button_column.find_element(By.TAG_NAME, 'label')
        change_to_button.click()
        time.sleep(7)

        return

    def format_for_file_upload(self, item_data: dict, path_to_save: str) -> None:
        # First row needs to be "Key Word, Quantity"
        # Then just CSV-style rows saved as an .XLSX file
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1).value = 'Key Word'
        sheet.cell(row=1, column=2).value = 'Quantity'

        for pos, sku in enumerate(item_data):
            quantity = item_data[sku]['quantity']
            sheet.cell(row=pos+2, column=1).value = sku
            sheet.cell(row=pos+2, column=2).value = quantity
        
        workbook.save(filename=f'{path_to_save}.xlsx')

        return

    def retrieve_pricing_sheet(self) -> None:
        pass