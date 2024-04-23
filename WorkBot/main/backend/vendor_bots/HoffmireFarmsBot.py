from .VendorBot import VendorBot, SeleniumBotMixin, PricingBotMixin
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium import webdriver
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from csv import writer
from datetime import date
from pprint import pprint
from ..craftable_bot import CraftableBot

class HoffmireFarmsBot(VendorBot, SeleniumBotMixin):

    def __init__(self, driver: webdriver = None, username: str = None, password: str = None) -> None:
        VendorBot.__init__(self)
        PricingBotMixin.__init__(self)
        SeleniumBotMixin.__init__(self, driver, username, password)
        self.name = 'Hoffmire Farms'
        self.minimum_order_amount = 400_00

        self.store_ids = {}
        self.special_cases = {
            'Kiwi - Fresh 1case': {'unit': 'EA', 'pack': 120},
            'Herb - Parsley (Fresh Curly) 1lb': {'unit': 'EA', 'pack': 60},
            'Cilantro - Fresh 1lb': {'unit': 'EA', 'pack': 60},
        }

    def login(self):
        with CraftableBot(self.driver, self.username, self.password) as craft_bot:
            craft_bot.login()

    def retrieve_pricing_sheet(self, guide_name: str) -> None:
        if not self.is_logged_in:
            self.login()

        self.driver.get('https://app.craftable.com/director/2/583/reports/purchase-log/0/0/0/period')
        time.sleep(3)

        
    def get_units_from_pack_size(self, pack_size: str) -> str:

        quantity = ''
        units = ''

        quantity_symbols = ['.']
        if 'x' not in pack_size:
            for char in pack_size:
                if char.isnumeric() or (char in quantity_symbols): quantity += char
                else: units += char
            return [float(quantity), units]


        quantity, units = pack_size.replace(' ', '').split('x')

        temp = ''
        final_units = ''
        for char in units:
            if char.isnumeric() or (char in quantity_symbols): temp += char
            else: final_units += char


        return [float(temp) * float(quantity), final_units]
        
    def get_info(self):
        workbook = load_workbook(self.path_to_pricing_sheet)
        sheet = workbook.active

        items = {}
        for row in sheet.iter_rows(min_row=7):

            name 		  = row[3].value
            if not name or name == '':
                continue

            if name not in self.special_cases:

                unit_value 	  = row[4].value
                tempunits 	  = self.get_units_from_pack_size(unit_value)
                units 		  = tempunits[1]
                quantity 	  = float(row[11].value)

            else:
                units = self.special_cases[name]['unit']
                quantity = self.special_cases[name]['pack']

            purchase_date = row[9].value
            cost = float(row[12].value) / quantity

            if name not in items:
                items[name] = {
                    'SKU': name,
                    'units': units,
                    'purchase_date': purchase_date,
                    'quantity': tempunits[0],
                    'cost': cost,
                    'case_size': unit_value
                }
            else:
                new_month, new_day, new_year 			 = purchase_date.split('/') 
                current_month, current_day, current_year = items[name]['purchase_date'].split('/')
                
                if datetime(int(new_year), int(new_month), int(new_day)) > datetime(int(current_year), int(current_month), int(current_day)):
                    items[name] = {
                        'SKU': name,
                        'units': units,
                        'purchase_date': purchase_date,
                        'quantity': tempunits[0],
                        'cost': cost,
                        'case_size': unit_value
                    }

        return items
