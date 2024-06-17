from .VendorBot import VendorBot, SeleniumBotMixin, PricingBotMixin
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium import webdriver
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from csv import writer
from datetime import date
from pprint import pprint

class CortlandProduceBot(VendorBot, SeleniumBotMixin, PricingBotMixin):

    def __init__(self, driver: webdriver = None, username: str = None, password: str = None) -> None:
        VendorBot.__init__(self)
        PricingBotMixin.__init__(self)
        SeleniumBotMixin.__init__(self, driver, username, password)
        self.name = 'Cortland Produce Inc.'
        self.minimum_order_amount = 400_00

        self.store_ids = {}

        self.special_cases = {
            '8203': {'unit': 'LB', 'pack': 20}, # This is a place holder until we get the real measurement
            '4060': {'unit': 'LB', 'pack': 30}, # This is a place holder until we get the real measurement
            '4193': {'unit': 'LB', 'pack': 8}, # This is a place holder until we get the real measurement
            '8275': {'unit': 'LB', 'pack': 10},
            '4219': {'unit': 'EA', 'pack': 1},
            '4166': {'unit': 'EA', 'pack': 80},
            '8465': {'unit': 'CS', 'pack': 1},
            '027398': {'unit': 'GAL', 'pack': 1},
        }

    def login(self):
        self.driver.get('https://net3.necs.com/cortlandproduce/login')

        time.sleep(5)

        username_input = self.driver.find_element(By.ID, 'username')
        password_input = self.driver.find_element(By.XPATH, '//input[@type="password"]')

        username_input.send_keys(self.username)
        password_input.send_keys(self.password)

        submit_button = self.driver.find_element(By.XPATH, '//button[text()="Sign In"]')
        submit_button.click()
        
        time.sleep(5)
        self.is_logged_in = True
        return

    def logout(self):
        pass

    def format_for_file_upload(self, item_data: dict, path_to_save: str):
        pass
    
    def retrieve_pricing_sheet(self, guide_name: str) -> None:
        if not self.is_logged_in:
            self.login()

        self.driver.get('https://net3.necs.com/cortlandproduce/customer/report/setup/3')
        time.sleep(4)

        guide_name_dropdown = self.driver.find_element(By.XPATH, '//*[@id="ext-comp-1031"]')
        guide_name_dropdown.click()
        time.sleep(2)

        guide_names_container = self.driver.find_element(By.XPATH, '//*[@id="ext-gen194"]')
        guide_names = [i for i in guide_names_container.find_elements(By.TAG_NAME, 'div') if guide_name == i.text][0]
        guide_names.click()
        time.sleep(2)

        unit_price_selection = self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div/div/div[1]/div/div[2]/div[1]/div/div/div[2]/div/div/div[3]/table/tbody/tr[2]/td[1]/div/div/div/div/div/div[1]/div[2]/div/div[3]/table/tbody/tr/td/div')
        unit_price_selection.click()
        time.sleep(1)
   
        unit_price_selection = self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div/div/div[1]/div/div[2]/div[1]/div/div/div[2]/div/div/div[3]/table/tbody/tr[2]/td[2]/div/div/div/div/table[3]/tbody/tr[2]/td[2]/em/button')
        unit_price_selection.click()
        time.sleep(2)

        format_dropdown = self.driver.find_element(By.XPATH, '//*[@id="ext-comp-1007"]')
        format_dropdown.click()
        time.sleep(3)

        excel_dropdown_option = self.driver.find_element(By.XPATH, '//div[text()="Excel 2007 (.xlsx)"]')
        excel_dropdown_option.click()
        time.sleep(2)

        download_report_button = self.driver.find_element(By.XPATH, '//*[@id="ext-gen63"]')
        download_report_button.click()
        time.sleep(3)

        self.driver.switch_to.window(window_name=self.driver.window_handles[1])
        self.driver.close()

        self.driver.switch_to.window(window_name=self.driver.window_handles[0])
        time.sleep(2)
        
        return f'OrderGuide_ProductListing-{date.today().strftime('%m-%d-%Y')}.xlsx'
        

    def get_pricing_info_from_sheet(self, path_to_pricing_sheet: str) -> dict:
        workbook = load_workbook(path_to_pricing_sheet)
        sheet    = workbook.active

        row_info  = list(PricingBotMixin.helper_iter_rows(sheet))[1:]
        item_info = {}

        numeric_non_numerics = ['.', '-', '/']
        for row in row_info:
            item_sku  = str(row[1])
            item_name = str(row[2])

            pack_size = ''
            unit = ''
            if item_sku not in self.special_cases:

                pack_size_info = row[6]
                
                # Extract units
                for char in pack_size_info:
                    if char.isnumeric() or char in numeric_non_numerics:
                        pack_size += char

                    else: unit += char

                # Average if necessary
                if '-' in pack_size:
                    lower, upper = pack_size.split('-')
                    pack_size = ( int(lower) + int(upper) ) / 2

    
                # Combine if necessary
                elif '/' in pack_size:
                    packs = pack_size.split('/')[0] or 1
                    size = pack_size.split('/')[1] or 1

                    pack_size = ( float(packs) * float(size) )

            else:
                unit = self.special_cases[item_sku]['unit']
                pack_size = self.special_cases[item_sku]['pack']

            cost = float(row[8])

            if item_name not in item_info:
                item_info[item_name] = {
                    'SKU': item_sku,
                    'quantity': pack_size,
                    'units': PricingBotMixin.normalize_units(unit),
                    'cost': cost,
                    'case_size': row[6]
                }
            

        return item_info