from __future__ import annotations
from dataclasses import dataclass
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pathlib import Path


class TransferItem:

    def __init__(self, name: str, quantity: float) -> None:
        self.name     = name
        self.quantity = quantity

class Transfer:

    def __init__(self, store_from: str, store_to: str, items: set[TransferItem], date: datetime) -> None:
        self.store_from = store_from
        self.store_to   = store_to
        self.items      = items
        self.date       = date

    def to_excel_workbook(self) -> Workbook:
        
        workbook = Workbook()
        sheet = workbook.active

        col_headers = ['Item', 'Quantity']

        # Insert headers
        for pos, header in enumerate(col_headers):
            sheet.cell(row=1, column=pos+1).value = header

        # Insert item data
        for pos, item in enumerate(self.items):
            for info_pos, item_info in enumerate(item):
                sheet.cell(row=pos+2, column=info_pos+1).value = item_info

        return workbook
    
    @staticmethod
    def from_excel_workbook(excel_file_path: Path, store_from: str, store_to: str, date: datetime) -> Transfer:
        
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active

         # Read headers (Assuming headers in the first row)
        headers = [cell.value for cell in sheet[1]]
        col_index = {col: idx for idx, col in enumerate(headers)}

        # Ensure necessary columns exist
        required_cols = {"Item", "Quantity"}
        if not required_cols.issubset(col_index.keys()):
            raise ValueError(f"Excel file must contain the columns: {required_cols}")
        
        items = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip headers
            item_name = row[col_index["Item"]]
            quantity = row[col_index["Quantity"]]

            if item_name and quantity is not None:
                items.add(TransferItem(item_name, float(quantity)))

        return Transfer(store_from, store_to, items, date)


    
