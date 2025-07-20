from __future__ import annotations # Lazy type-hint evaluation.
from ..stores.Store import Store
from pathlib import Path
import re
from openpyxl import load_workbook, Workbook

class OrderItem:

    def __init__(self, sku: str, name: str, quantity: float, cost_per: float, total_cost: float = None) -> None:
        self.sku        = sku
        self.name       = name
        self.quantity   = quantity
        self.cost_per   = cost_per
        self.total_cost = total_cost or (quantity * cost_per)

    def to_dict(self) -> dict:
        return {
            'sku':        self.sku,
            'name':       self.name,
            'quantity':   self.quantity,
            'cost_per':   self.cost_per,
            'total_cost': self.total_cost
        }
    
    def __repr__(self):
        return f'< OrderItem sku={self.sku}, name={self.name}, quantity={self.quantity}, cost_per={self.cost_per}, total_cost={self.total_cost} >'


class Order:

    def __init__(self, store: str, vendor: str, date: str, items: list[OrderItem] = None) -> None:
        self.store  = store
        self.vendor = vendor
        self.date   = date
        self.items  = items or []

    def load_items_from_csv(self, load_path: Path) -> None:
        pass

    def generate_item_csv(self, save_path: Path) -> None:
        pass

    def to_dict(self) -> dict:
        return {
            'store':  self.store,
            'vendor': self.vendor,
            'date':   self.date,
            'items':  self.items
        }

    def load_items_from_excel(self, excel_path: Path) -> None:

        workbook = load_workbook(excel_path)
        sheet = workbook.active

        # col_headers = ['SKU', 'Item', 'Quantity', 'Cost Per', 'Total Cost']
        for row in sheet.iter_rows(min_row=2):
            
            item_sku, item_name, item_quantity, item_cost_per, item_total_cost = row[0:5]
            # print([item_sku.value, item_name.value, item_quantity.value, item_cost_per.value, item_total_cost.value])
            order_item = OrderItem(
                sku=item_sku.value,
                name=item_name.value,
                quantity=item_quantity.value,
                cost_per=item_cost_per.value,
                total_cost=item_total_cost.value
            )
            # print(order_item)
            self.items.append(order_item)
            
        return

    # def is_valid_date_format(self, given_date: str) -> bool:
    #     date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    #     return bool(date_pattern.match(given_date))



    def to_excel_workbook(self) -> Workbook:

        workbook = Workbook()
        sheet = workbook.active

        col_headers = ['SKU', 'Item', 'Quantity', 'Cost Per', 'Total Cost']

        # Insert headers
        for pos, header in enumerate(col_headers):
            sheet.cell(row=1, column=pos+1).value = header

        for pos, item in enumerate(self.items):
            sheet.cell(row=pos+2, column=1).value = item.sku
            sheet.cell(row=pos+2, column=2).value = item.name
            sheet.cell(row=pos+2, column=3).value = item.quantity
            sheet.cell(row=pos+2, column=4).value = item.cost_per
            sheet.cell(row=pos+2, column=5).value = item.total_cost


        return workbook
    
    def __repr__(self) -> str:
        return f'< Order store={self.store}, vendor={self.vendor}, date={self.date}, items={len(self.items)} >'
    
