import re
from pathlib import Path
from .Order import Order

from openpyxl import Workbook

from backend.logger.Logger import Logger
from config.paths import DOWNLOADS_DIR, ORDER_FILES_DIR

@Logger.attach_logger
class OrderManager:

    file_pattern = re.compile(r"^(?P<vendor>[^-]+?)_(?P<store>.+?)_(?P<date>\d{8})$")

    def get_order_files_directory(self) -> Path:
        return ORDER_FILES_DIR
    
    def get_downloads_directory(self) -> Path:
        return DOWNLOADS_DIR
    
    def get_vendor_orders_directory(self, vendor: str) -> Path:

        self.logger.info(f'Looking up directory for: {vendor}.')
        
        vendor_orders_directory = self.get_order_files_directory() / vendor
        if vendor_orders_directory.exists() and vendor_orders_directory.is_dir():
            self.logger.info(f'Directory found: {vendor_orders_directory}')
            return vendor_orders_directory
        self.logger.info(f'Directory not found for: {vendor}.')

        return None

    def get_store_orders(self, stores: list, vendors: list = []) -> list:

        orders_dir = self.get_order_files_directory()

        orders = []

        for vendor_order_dir in orders_dir.iterdir():
            for vendor_order_file in vendor_order_dir.iterdir():
                
                if vendor_order_file.suffix != '.xlsx' or not self.is_valid_filename(vendor_order_file): 
                    continue
               
                file_metadata = OrderManager.parse_file_name(vendor_order_file)
                if 'store' not in file_metadata or file_metadata['store'] not in stores: continue
                
                if vendors and file_metadata['vendor'] not in vendors: break

                order = OrderManager.create_order_from_excel(vendor_order_file)

                orders.append(order)

        return orders  

    def get_vendor_orders(self, vendor: str, file_extension: str = '.xlsx') -> list[Path]:
    
        vendor_orders_directory = self.get_vendor_orders_directory(vendor)

        if not vendor_orders_directory:
            return []

        return [file for file in vendor_orders_directory.iterdir() if file.is_file() and self.is_valid_filename(file) and file.suffix == file_extension]

    def generate_filename(self, order: Order = None, store: str = None, vendor: str = None, date: str = None, file_extension: str = '.xlsx') -> str:
        return f'{order.vendor}_{order.store}_{order.date}{file_extension}' if order else f'{vendor}_{store}_{date}{file_extension}'

    def get_file_path(self, order: Order, file_extension: str = '.xlsx') -> Path:
        return self.get_order_files_directory() / self.generate_filename(order, file_extension)

    def save_order(self, order: Order, file_extension: str = '.xlsx') -> None:

        extensions = {
            '.xlsx': self._save_as_excel
        }

        if file_extension in extensions: extensions[file_extension](order)

        else: return None

    def _save_as_excel(self, order: Order) -> None:
        
        workbook = order.to_excel_workbook()
        workbook.save(self.get_file_path(order, '.xlsx'))

        return 
    
    def sort_orders(self) -> None:
        """Sorts order files into vendor-specific subdirectories within the orders directory."""
        orders_dir = self.get_order_files_directory()
        
        for file in orders_dir.iterdir():
            
            if file.is_file() and OrderManager.is_valid_filename(file):
                
                order_info = self.parse_file_name(file)
                vendor_name = order_info.get('vendor')
                
                if not vendor_name:
                    raise ValueError(f'Filename "{file.stem}" does not contain a valid vendor name.')

                vendor_dir = orders_dir / vendor_name  # Target directory for vendor
                vendor_dir.mkdir(exist_ok=True)  # Create vendor directory if it doesn't exist
                
                new_path = vendor_dir / file.name
                file.rename(new_path)  # Move file into the vendor directory

    def combine_orders(self, vendors: list) -> None:
        for vendor in vendors:
            vendor_order_dir = self.get_vendor_orders_directory(vendor)
            order_files      = self.get_vendor_orders(vendor)

            combined_orders = {}

            for order_file in order_files:
                order = self.create_order_from_excel(order_file)

                for item in order.items:
                    if item.name not in combined_orders:
                        combined_orders[item.name] = {order.store: item.quantity}

            combined_excel = self._create_combined_orders_excel(combined_orders)
            combined_excel.save(f'{vendor_order_dir / 'combined_orders.xlsx'}')

        return
                
    def _create_combined_orders_excel(self, combined_orders: dict):

        workbook = Workbook()
        sheet = workbook.active

        store_names = {store for item in combined_orders.values() for store in item.keys()}
        headers = ['Item']
        for i in store_names:
            headers.append(i)
        
        for pos, header in enumerate(headers):
            sheet.cell(row=1, column=pos+1).value = header

        for pos, item in enumerate(combined_orders):
            for store in combined_orders[item]:
                sheet.cell(row=pos+2, column=(headers.index(store)+1)).value = combined_orders[item][store]

        return workbook

                

    @classmethod
    def create_order_from_excel(cls, file_path: Path) -> Order:

        file_metadata = cls.parse_file_name(file_path)
        store_name    = file_metadata['store']
        vendor_name   = file_metadata['vendor']
        order_date    = file_metadata['date']

        order = Order(store_name, vendor_name, order_date)
        order.load_items_from_excel(file_path)

        return order
    
    @classmethod
    def is_valid_filename(cls, filename: Path) -> bool:
        return bool(re.fullmatch(cls.file_pattern, filename.stem))
    
    @classmethod
    def parse_file_name(cls, filename: Path) -> dict:
        if cls.is_valid_filename(filename): 
            return cls.file_pattern.match(filename.stem).groupdict()
        raise ValueError(f'Filename "{filename}" does not match expected pattern.')
        