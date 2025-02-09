import re
from pathlib import Path
from .Order import Order
from openpyxl import Workbook

SOURCE_PATH      = Path(__file__).parent.parent
ORDERS_DIRECTORY = SOURCE_PATH / 'orders' / 'OrderFiles'
DOWNLOADS_DIRECTORY = SOURCE_PATH / 'downloads'

class OrderManager:

    def __init__(self):
        self.orders_directory = self.get_order_files_directory()
        self.file_pattern     = re.compile(r"^(.+?) _ (.+?) (\d{2}\d{2}\d{4})$")

    def get_order_files_directory(self) -> Path:
        return ORDERS_DIRECTORY
    
    def get_downloads_directory(self) -> Path:
        print(f'OrderManager: {DOWNLOADS_DIRECTORY}', flush=True)
        return DOWNLOADS_DIRECTORY
    
    def generate_filename(self, order: Order = None, store: str = None, vendor: str = None, date: str = None, file_extension: str = '.xlsx') -> str:

        if order:
            return f'{order.vendor} _ {order.store} {order.date}{file_extension}'
        
        return f'{vendor} _ {store} {date}{file_extension}'
    
    def get_file_path(self, order: Order, file_extension: str = '.xlsx') -> Path:
        return self.orders_directory / self.generate_filename(order, file_extension)

    def save_order(self, order: Order, file_extension: str = '.xlsx') -> None:

        extensions = {
            '.xlsx': self._save_as_excel
        }

        if file_extension in extensions: extensions[file_extension]()

        else: return None

    def _save_as_excel(self, order: Order) -> None:
        
        workbook = self._order_to_excel(order)
        workbook.save(self.get_file_path(order, '.xlsx'))

        return 
    
    def _order_to_excel(self, order: Order) -> Workbook:

        workbook = Workbook()
        sheet = workbook.active

        col_headers = ['SKU', 'Item', 'Quantity', 'Cost Per', 'Total Cost']

        # Insert headers
        for pos, header in enumerate(col_headers):
            sheet.cell(row=1, column=pos+1).value = header

        # Insert item data
        for pos, item in enumerate(order.items):
            for info_pos, item_info in enumerate(item):
                sheet.cell(row=pos+2, column=info_pos+1).value = item_info
        
        return workbook
    
    # def extract_general_order_info(self, order_file_name: Path) -> dict:
    #     filename = order_file_name.stem
    #     print(filename)
    #     match = self.file_pattern.match(filename)
    #     print(match)
    #     if not match: return None

    #     store_name, vendor_name, date = match.groups()

    #     return {'store': store_name,
    #             'vendor': vendor_name,
    #             'date': date}