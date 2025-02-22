from __future__ import annotations # Lazy type-hint evaluation.
from ..stores.Store import Store
from pathlib import Path
import re
from openpyxl import load_workbook, Workbook


class Order:

    def __init__(self, store: str, vendor: str, date: str, items: list = None) -> None:
        self.store  = store
        self.vendor = vendor
        self.date   = date
        self.items  = items or []

    def load_items_from_csv(self, load_path: Path) -> None:
        pass

    def generate_item_csv(self, save_path: Path) -> None:
        pass

    def to_dict(self) -> dict:
        return {
            'store':  self.store,
            'vendor': self.vendor,
            'date':   self.date,
            'items':  self.items
        }

    def load_items_from_excel(self, excel_path: Path) -> None:

        workbook = load_workbook(excel_path)
        sheet = workbook.active

        # col_headers = ['SKU', 'Item', 'Quantity', 'Cost Per', 'Total Cost']

        for row in sheet.iter_rows(min_row=2):
            sku, name, quantity, cost_per, total_cost = row[0:5]
            self.items.append({
                'SKU': sku,
                'Item': name,
                'Quantity': quantity,
                'Cost Per': cost_per,
                'Total Cost': total_cost
                })
            
        return

    # def is_valid_date_format(self, given_date: str) -> bool:
    #     date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    #     return bool(date_pattern.match(given_date))



    def to_excel_workbook(self) -> Workbook:

        workbook = Workbook()
        sheet = workbook.active

        col_headers = ['SKU', 'Item', 'Quantity', 'Cost Per', 'Total Cost']

        # Insert headers
        for pos, header in enumerate(col_headers):
            sheet.cell(row=1, column=pos+1).value = header

        # Insert item data
        for pos, item in enumerate(self.items):
            for info_pos, item_info in enumerate(item):
                sheet.cell(row=pos+2, column=info_pos+1).value = item_info

        return workbook
    
    def __repr__(self) -> str:
        return f'< Order store={self.store}, vendor={self.vendor}, date={self.date}, items={len(self.items)} >'
    
