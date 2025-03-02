from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains 
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.webdriver import WebDriver

import time
from pynput.keyboard import Key, Controller
import os


from openpyxl import Workbook, load_workbook

from backend.transferring import Transfer
from backend.transferring.TransferManager import TransferManager
from backend.orders.OrderManager import OrderManager
from backend.orders.Order import Order

from datetime import datetime

from pathlib import Path

from backend.logger.Logger import Logger
from backend.helpers.DatetimeFormattingHelper import convert_date_format


def temporary_login(func):
    def wrapper(self, *args, **kwargs):
        self.login()
        try:
            return func(self, *args, **kwargs)
        finally:
            self.close_session()
    return wrapper


def login_necessary(func):
    def wrapper(self, *args, **kwargs):
        if not self.is_logged_in:
            self.login()
        return func(self, *args, **kwargs)
    return wrapper

'''
Craftable Bot utlizes Selenium to interact with the Craftable website. 
'''
class CraftableBot:

    _logger = None

    site_map = {
            'login_page': 'https://app.craftable.com/signin',
            'orders_page': 'https://app.craftable.com/buyer/2/{store_id}/orders/list',
            'transfer_page': 'https://app.craftable.com/buyer/2/{store_id}/transfers/list',
        }
    
    stores = {
            'BAKERY':      '14376',
            'DOWNTOWN':    '14373',
            'EASTHILL':    '14374',
            'TRIPHAMMER':  '14375',
            'COLLEGETOWN': '14372',
        }
    
    @classmethod
    def get_logger(cls):
        if cls._logger is None:
            cls._logger = Logger.get_logger('CraftableBot', log_file='logs/craftable_bot.log')
        return cls._logger
    
    def __init__(self, driver: WebDriver, username: str, password: str, 
                 order_manager: OrderManager = None, transfer_manager: TransferManager = None):
        
        self.logger = self.get_logger()
        self.driver           = driver
        self.username         = username
        self.password         = password
        self.order_manager    = order_manager or OrderManager()
        self.transfer_manager = transfer_manager or TransferManager()

        self.is_logged_in = False

    # def __enter__(self):
    #     self.logger.info('Starting CraftableBot session.')
    #     self.login()
    #     return self
    
    # def __exit__(self, type, value, traceback):
    #     self.logger.info('Ending CraftableBot session.')
    #     try:
    #         self.close_session()
    #     except Exception as e:
    #         self.logger.warning(f'Issue with closing session: {e}')

    #     time.sleep(2) # Prevents race condition with OS program manager
    #     return True
    
    '''
    Login to the Craftable website using the standard landing page.

    This method assumes the bot knows the username and password.

    Args:
        None

    Returns:
        None
    '''
    @Logger.log_exceptions
    def login(self) -> None:
        self.logger.info('Logging into Craftable.')
        self.driver.get(self.site_map['login_page'])

        WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, './/form')))
        
        login_form = self.driver.find_element(By.XPATH, './/form')
        
        fieldsets         = login_form.find_elements(By.XPATH, './/fieldset')
        email_fieldset    = fieldsets[0]
        password_fieldset = fieldsets[1]

        email_input    = email_fieldset.find_element(By.XPATH, './/input')
        password_input = password_fieldset.find_element(By.XPATH, './/input')

        submit_button_div = login_form.find_elements(By.XPATH, './/button')[0]
        
        email_input.send_keys(self.username)
        password_input.send_keys(self.password)

        submit_button_div.click()
        time.sleep(5)

        self.is_logged_in = True

        self.logger.info('Login successful.')
        return
    
    '''
    Close the driver session.

    Args:
        None

    Returns:
        None
    '''
    def close_session(self) -> None:
        self.logger.info('Closing WebDriver session.')

        try: 
            self.driver.close()
        except Exception as e: 
            self.logger.error(f'Error closing WebDriver session: {e}')

        time.sleep(2)
        return
    
    '''
    Switch to another store's page.

    Args:
        store_code: store name

    Returns:
        None
    '''
    @Logger.log_exceptions
    def switch_store(self, store: str) -> None:
        
        if store not in self.stores: 
            self.logger.warning(f'Invalid store name: {store}.')
            return

        if store == 'DIRECTOR':
            self.driver.get('https://app.craftable.com/director/2/583')
            return
        
        self.logger.info(f'Switching to store: {store}.')
        store_url = f'https://app.craftable.com/buyer/2/{self.stores[store]}'
        self.driver.get(store_url)

        #WebDriverWait(self.driver, 45).until(EC.element_to_be_clickable((By.CLASS_NAME, 'row home-cards-orders-invoices')))
        time.sleep(5)
        return
    
    ''' Download orders from Craftable.

    ARGS

    stores: a list of stores for which to retrieve orders.
    vendors: a list of vendors to retrieve orders of. If none supplied, then all possible orders are retrieved.
    download_pdf: a boolean of whether to download an accompanying PDF file.
    update: a boolean of whether to overwrite currently saved order data.

    RETURN

    None

    Will scrape the Craftable order page for each store and selected vendors, download and generate the appropriate
    files, and save in the ORDER_FILES_PATH.

    '''
    @login_necessary
    @Logger.log_exceptions
    def download_orders(self, stores: list, vendors=list, download_pdf=True, update=True) -> None:

        self.logger.info('Starting order download protocol.')
        for store in stores:
            
            self.logger.info(f'Accessing order page for {store}.')
            store_order_url = self.get_url('orders_page', store=store)
            self.driver.get(store_order_url)
            time.sleep(6)
            
            table_rows = self._get_order_table_rows()
            if not table_rows:
                self.logger.info(f'No more orders found for {store}. Moving to next store.')
                break  # Exit while loop and move to the next store

            for pos in range(len(table_rows)):
                stale_reference_table_rows = self._get_order_table_rows() # Refresh to avoid stale references
                self._process_order_row(store, stale_reference_table_rows[pos], vendors, download_pdf, update)

        self.logger.info("Order download complete.")

        return
   
    def _get_order_table_rows(self) -> list:
        """Returns all order rows in the table, handling stale references."""
        try:
            table_body = self.driver.find_element(By.XPATH, './/tbody')
            return table_body.find_elements(By.XPATH, './tr')
        except Exception as e:
            self.logger.warning(f"Failed to retrieve order table: {e}")
            return []
    
    def _process_order_row(self, store: str, row, vendors: list, download_pdf: bool, update: bool) -> None:
        """Processes a single order row: extracts data, downloads order, and saves it."""

        row_data = row.find_elements(By.XPATH, './td')
        row_date_text = row_data[2].text
        row_vendor_name = row_data[3].text
        row_date_formatted = convert_date_format(row_date_text, '%m/%d/%Y', '%Y%m%d')

        if vendors and row_vendor_name not in vendors:
            self.logger.debug(f"Skipping vendor {row_vendor_name}.")
            return 

        self.logger.info(f"Retrieving order for {row_vendor_name} ({row_date_text}).")

        row_data[2].click()
        

        WebDriverWait(self.driver, 45).until(EC.element_to_be_clickable((By.TAG_NAME, 'table')))
        items = self._scrape_order()

        if update and self._update_existing_order(store, row_vendor_name, row_date_formatted, items):
            self.logger.info(f"Order for {row_vendor_name} is up-to-date. Skipping.")
            self.driver.back()
            time.sleep(2)
            return 

        self.logger.info(f"Saving order for {row_vendor_name}.")
        order_to_save = Order(store=store, vendor=row_vendor_name, date=row_date_formatted, items=items)
        self.order_manager.save_order(order_to_save)

        # self._save_order_as_excel(items, store=store, vendor=row_vendor_name, date=row_date_formatted)
        time.sleep(1)

        if download_pdf:
            self._download_order_pdf()
            time.sleep(3)
            self._rename_new_order_file(store=store, vendor=row_vendor_name, date=row_date_formatted)

        self.driver.back()
        time.sleep(3)
        return 

    def _find_order_row(self, table_rows, vendor_name, date_text):
        """Finds the correct order row by vendor name and order date."""
        for row in table_rows:
            row_data = row.find_elements(By.XPATH, './td')
            if len(row_data) < 4:
                continue  # Skip malformed rows
            if row_data[2].text == date_text and row_data[3].text == vendor_name:
                return row
        return None
    
    @login_necessary
    @Logger.log_exceptions
    def delete_orders(self, stores: list[str], vendors: list[str] = None) -> None:

        for store in stores:
            self.logger.info(f'Starting order deletion for store: {store}.')
            
            if store not in stores:
                self.logger.warning(f'Invalid store name: {store}, Skipping.')
                continue

            store_orders_url = self.get_url('orders_page', store=store)
            self.logger.debug(f'Navigating to: {store_orders_url}')
            self.driver.get(store_orders_url)
            time.sleep(6)

            table_body = self.driver.find_element(By.XPATH, './/tbody')
            table_rows = table_body.find_elements(By.XPATH, './tr')

            if not table_rows:
                self.logger.info(f'No orders found for store: {store}. Skipping.')
                continue

            self.logger.info(f'Found {len(table_rows)} orders for deletion.')

            while table_rows:
                row             = table_rows[-1]  # Always delete the last order first
                row_data        = row.find_elements(By.XPATH, './td')
                row_date_text   = row_data[2].text
                row_vendor_name = row_data[3].text

                # If vendors list is provided, only delete matching vendor orders
                if vendors and row_vendor_name not in vendors:
                    self.logger.debug(f"Skipping order from vendor: {row_vendor_name}.")
                    table_rows.pop()
                    continue

                self.logger.info(f"Deleting order: {row_vendor_name} ({row_date_text})")
                self._delete_order(row)
                
                # self.driver.back()
                # time.sleep(5)
                # Refresh table after deletion
                try:
                    table_body = self.driver.find_element(By.XPATH, './/tbody')
                    table_rows = table_body.find_elements(By.XPATH, './tr')
                except:
                    self.logger.info('No orders table found. All orders deleted.')
                    break

            self.logger.info(f"Finished deleting orders for store: {store}.")

    def _delete_order(self, row) -> None:
        """Handles the deletion of a single order."""
        try:
            row.find_elements(By.XPATH, './td')[2].click()
            WebDriverWait(self.driver, 35).until(EC.element_to_be_clickable((By.CLASS_NAME, 'btn-danger')))
        except Exception as e:
            self.logger.warning(f"Could not open order details: {e}")
            return

        self._click_delete_button()
        self._confirm_delete()

    def _click_delete_button(self) -> None:
        """Clicks the delete button within the order page."""
        for attempt in range(3):
            try:
                delete_button = self.driver.find_element(By.CLASS_NAME, 'btn-danger')
                delete_button.click()
                self.logger.debug("Delete button clicked.")
                return
            except Exception as e:
                self.logger.warning(f"Attempt {attempt+1} failed to click delete button: {e}")
                time.sleep(2)

        self.logger.error("Failed to click delete button after 3 attempts.")

    def _confirm_delete(self) -> None:
        """Handles confirming the deletion of an order."""
        try:
            WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.ID, 'confirmOrderDeleteModal')))
            confirm_modal = self.driver.find_element(By.ID, 'confirmOrderDeleteModal')
            modal_footer  = confirm_modal.find_element(By.CLASS_NAME, 'modal-footer')
            delete_button = modal_footer.find_elements(By.TAG_NAME, 'button')[1]

            delete_button.click()
            self.logger.info("Order deletion confirmed.")
            time.sleep(5)
        except Exception as e:
            self.logger.error(f"Failed to confirm deletion: {e}")

    @classmethod
    @Logger.log_exceptions
    def get_url(cls, key: str, store: str = None) -> str:

        cls.logger.debug('Retrieving URL from site map.')

        if key not in cls.site_map:
            cls.logger.error(f'INvalid site map key: {key}')
            return ''
        
        if '{store_id}' in cls.site_map[key]:
            if store not in cls.stores:
                cls.logger.error(f'Invalid store name: {store}')
                return ''
            return cls.site_map[key].format(store_id=cls.stores[store])
        
        return cls.site_map[key]
    
    @temporary_login
    @Logger.log_exceptions
    def input_transfers(self, transfers: list) -> None:

        self.logger.info(f'Starting protocol for {len(transfers)} transfers.')
        for transfer in transfers:
            self.input_transfer(transfer)
        self.logger.info('Transfers complete.')
        return
    
    @Logger.log_exceptions
    def input_transfer(self, transfer: Transfer) -> None:
        
            
        self.logger.info(f'Starting transfer protocol for {len(transfer.items)} items from {transfer.store_from} to {transfer.store_to} on {transfer.date}')
        if not self.is_logged_in: self.login()

        time.sleep(3)

        transfer_url = self.get_url('transfer_page', store=transfer.store_from)
        self.driver.get(transfer_url)

        time.sleep(5)

        self._start_new_transfer()
        self._enter_transfer_details(transfer)
        self._submit_transfer()
        self._input_transfer_items(transfer)

        self.logger.info(f'Transfer for {len(transfer.items)} items from {transfer.store_from} to {transfer.store_to} on {transfer.date} completed successfully.')
        
        # submit_transfer_button = self.driver.find_element(By.XPATH, './/a[text()="Request"]')
        return

    def _start_new_transfer(self) -> None:
        self.logger.info('Creating new transfer in Craftable.')
        new_transfer_button = self.driver.find_element(By.XPATH, '//a[text()="New Transfer"]')
        new_transfer_button.click()

        time.sleep(4) # WebDriver wait and exception handling...or we just wait and pray.

        return
    
    def _enter_transfer_details(self, transfer: Transfer) -> None:

        self.logger.info(f'Entering transfer details: {transfer.store_from} to {transfer.store_to} on {transfer.date}')

        transfer_modal = self.driver.find_element(By.ID, 'transferNewModal')

        transfer_form = transfer_modal.find_element(By.XPATH, './/form')
        transfer_form_inputs = transfer_form.find_elements(By.XPATH, './div')

        date_input = transfer_form_inputs[0].find_element(By.XPATH, './/input')
        toggle_out = transfer_form_inputs[1].find_element(By.XPATH, './/input') # We can do this because we only want the first input.
        
        self.logger.info('Opening transfer date calendar.')
        date_input.click()
        
        WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.CLASS_NAME, 'flatpickr-calendar')))
        # date_input_current_month = self.driver.find_element(By.CLASS_NAME, 'cur-month')
        
        # calendar_current_month = self.driver.find_element(By.CLASS_NAME, 'cur-month')
        # calendar_current_year_input = self.driver.find_element(By.CLASS_NAME, 'cur_year')

        # calendar_current_year_input.send_keys('2025')

        # NEED TO UPDATE SO IT CAN CHOOSE THE CORRECT MONTH!!!!!
        self.logger.info(f'Selecting calendar date: {transfer.date}')
        self._change_calendar_date(transfer.date)

        # calendar = self.driver.find_element(By.CLASS_NAME, 'dayContainer')
        # print(transfer.date.day)
        # today = calendar.find_element(By.XPATH, f'.//span[@class="flatpickr-day "][text()="{transfer.date.day}"]')
        # today.click()
        
        self.logger.info('Marking transfer as "Out".')
        time.sleep(2)
        toggle_out.click()
        time.sleep(2)

        '''Have to search these again to avoid stale references...'''
        transfer_modal = self.driver.find_element(By.ID, 'transferNewModal')

        transfer_form = transfer_modal.find_element(By.XPATH, './/form')
        transfer_form_inputs = transfer_form.find_elements(By.XPATH, './div')

        self.logger.info(f'Selecting outbound store: {transfer.store_to}')
        store_to_select = transfer_form_inputs[2].find_element(By.XPATH, './/div[@class="search ember-view input-select-searchable"]')
        store_to_select.click()

        time.sleep(2)
    
        options = self.driver.find_elements(By.XPATH, './/li[@class="select2-results__option"]')
        for option in options:
            choice_id_full = option.get_attribute('data-select2-id')
            choice_id = choice_id_full.split('-')[-1]
            if choice_id == self.stores[transfer.store_to]:
                choice = self.driver.find_element(By.XPATH, f'//li[@data-select2-id="{choice_id_full}"]')
                choice.click()
                break

        time.sleep(4)
        return
    
    def _submit_transfer(self) -> None:
        self.logger.info('Submitting transfer form.')
        WebDriverWait(self.driver, 120).until(EC.element_to_be_clickable((By.ID, 'transferNewModal')))
        transfer_modal        = self.driver.find_element(By.ID, 'transferNewModal')
        transfer_modal_footer = transfer_modal.find_element(By.XPATH, './/div[@class="modal-footer "]')

        submit_button = transfer_modal_footer.find_elements(By.TAG_NAME, 'button')[1]
        submit_button.click()

        time.sleep(4)
        return
    
    def _input_transfer_items(self, transfer: Transfer) -> None:
        self.logger.info(f'Starting input for {len(transfer.items)} transfer items.')

        for item in transfer.items:
            print(item.name, flush=True)
            if item.quantity <= 0: return # Skip items that weren't transferred.
            try:
                self.logger.debug(f'Adding transfer item: {item.name} ({item.quantity})')
                self._add_transfer_item(item)
            except Exception as e:
                self.logger.warning(f'Failed to add item {item.name}: {e}', exc_info=True)
        
        self.logger.info('All items processed.')
        return
    
    def _add_transfer_item(self, item) -> None:


        self.logger.debug(f'Clicking "Add Transfer Line" button for {item.name}.')
        # Click the add-item button
        page_body = self.driver.find_element(By.XPATH, './/div[@class="card-body"]')
        WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, '//a[text()="Add Transfer Line"]')))
        add_item_button = page_body.find_element(By.XPATH, '//a[text()="Add Transfer Line"]')
        add_item_button.click()

        time.sleep(2)

        # Enter the item name   
        item_input = self.driver.find_element(By.ID, 'typeahead')
        item_input.send_keys(item.name)
        time.sleep(5)

        # Select the item from the dropdown
        item_choice_container = self.driver.find_element(By.XPATH, './/div[@class="input-group input-typeahead-container input-group-merge"]')
        item_choice = item_choice_container.find_elements(By.XPATH, './following-sibling::div/div[@class="input-type-ahead "]/div[@class="input-type-ahead-row"]')[0]
        item_choice.click()
        
        transfer_modal           = self.driver.find_element(By.ID, 'transferItemModal')
        transfer_form            = transfer_modal.find_element(By.XPATH, './/form')
        quantity_input_container = transfer_form.find_elements(By.XPATH, './div')[2]
        quantity_input           = quantity_input_container.find_element(By.XPATH, './/input')

        quantity_input.send_keys(Keys.BACKSPACE)
        quantity_input.send_keys(item.quantity)
        time.sleep(3)

        for attempt in range(3):
            try:
                add_transfer_item_button = self.driver.find_element(By.XPATH, './/button[text()="Add Transfer Line"]')
                add_transfer_item_button.click()
                WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, './/div[@class="_c-notification__content"][text()="Successfully added a Transfer Line"]')))
                self.logger.debug(f'Item {item.name} addedd successfully.')
                break
            except Exception as e:
                self.logger.warning(f'Attempt {attempt+1} failed for {item.name}: {e}')
            
        return
            
    '''HELPER FUNCTIONS'''

    def _rename_new_order_file(self, store: str, vendor: str, date: str) -> None:

        original_file = self.order_manager.get_downloads_directory() / 'Order.pdf'

        if not original_file.exists():
            self.logger.warding(f'Expected file not found: {original_file}')
            return
        
        new_filename = self.order_manager.generate_filename(store=store, vendor=vendor, date=date, file_extension='.pdf')
        new_filepath = self.order_manager.get_order_files_directory() / new_filename
        
        original_file.rename(new_filepath)
        return
    
    def _run_save_protocol() -> None:
        keyboard = Controller()

        keyboard.press(Key.ctrl)
        keyboard.press('s')
        keyboard.release(Key.ctrl)
        keyboard.release('s')

        time.sleep(2)
              
        keyboard.press(Key.enter)
        keyboard.release(Key.enter)

        return
    
    '''
    Assumes the current driver is 'looking at' an order page.

    Scrapes the current page for order data and returns the data 
    as a 2-D list.
    '''
    def _scrape_order(self) -> list:
        order_table = self.driver.find_element(By.TAG_NAME, 'tbody')
        item_rows   = order_table.find_elements(By.TAG_NAME, 'tr')

        items = []
        for item in item_rows:
            item_info = item.find_elements(By.TAG_NAME, 'td')
        
            item_sku   = item_info[2].text
            item_name  = item_info[3].find_element(By.XPATH, './/a').text
            quantity   = item_info[5].text
            cost_per   = item_info[6].text.replace('$', '').replace(',', '')
            total_cost = item_info[7].text.replace('$', '').replace(',', '')

            items.append([
                item_sku, 
                item_name, 
                quantity, 
                cost_per, 
                total_cost
                ])
            
        return items
    
    '''
    Assumes the current driver is 'looking at' an order page.

    Clicks the download button on the order page.
    '''
    def _download_order_pdf(self) -> None:
        download_button = self.driver.find_element(By.CLASS_NAME, 'fa-download')
        ActionChains(self.driver).key_down(Keys.CONTROL).click(download_button).perform()
        return
    
    def _save_order_as_excel(self, item_data, store: str, vendor: str, date: str) -> None:

        self.order_manager.save_order()
        new_filename = self.order_manager.generate_filename(store=store, vendor=vendor, date=date, file_extension='.xlsx')
        new_filepath = self.order_manager.get_order_files_directory() / new_filename
        
        self.logger.info(f'Saving order as Excel file: {new_filename}')
        order = Order(store, vendor, date, items=item_data)
     
        order_as_workbook = order.to_excel_workbook()

        order_as_workbook.save(new_filepath)
        order_as_workbook.close()
        
        return
    
    def _delete_order_protocol(self) -> None:
        pass
        
    def _extract_order_table_rows(self, row) -> tuple:
        pass

    def _update_existing_order(self, store, vendor_name, date_formatted, items) -> bool:

        """Handles the update protocol for checking existing orders and removing outdated files."""

        self.logger.info(f'[Order Update] Initiating update protocol for {store} for vendor: {vendor_name}, on {date_formatted}.')

        preexisting_workbook_filename = self.order_manager.generate_filename(
            store=store, 
            vendor=vendor_name, 
            date=date_formatted, 
            file_extension='.xlsx'
        )
        
        self.logger.info(f'[Order Update] Checking for pre-existing order file: {preexisting_workbook_filename}')
        preexisting_workbook_path = self.order_manager.get_order_files_directory() / vendor_name / preexisting_workbook_filename
        preexisting_file_exists   = preexisting_workbook_path.exists()
        
        if not self.order_manager.get_vendor_orders_directory(vendor_name): 
            self.logger.info(f'[Order Update] No order directory found for vendor: {vendor_name}. Proceeding with new download.')
            return False
        if not preexisting_file_exists: 
            self.logger.info(f'[Order Update] No existing order file found for {store} for {vendor_name} on {date_formatted}. Proceeding with new download.')
            return False
        
        self.logger.info(f'[Order Update] Extracting data from existing order file: {preexisting_workbook_filename}')
        workbook    = load_workbook(preexisting_workbook_path, read_only=True)
        sheet       = workbook.active
        saved_items = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]
        workbook.close()

        self.logger.info(f'[Order Update] Comparing new data with the existing file for {store} for {vendor_name} on {date_formatted}.')
        if set(map(tuple, items)) == set(map(tuple, saved_items)):
            self.logger.info(f'[Order Update] No changes detected for {store} for {vendor_name} on {date_formatted}. Skipping update protocol.')
            return True  # Skip further processing

        self.logger.info(f'[Order Update] Removing outdated order file: {preexisting_workbook_path}')
        self._remove_old_file(preexisting_workbook_path)
        self._remove_old_file(preexisting_workbook_path.with_suffix('.pdf'))
        return False  # Continue with saving new data
    
    def _remove_old_file(self, file_path: Path) -> None:
        '''Used to remove outdated order files.'''
        try:

            
            self.logger.info(f'[File Cleanup] Attempting to remove file: {file_path}')

            if file_path.is_file():
                os.remove(file_path)
                self.logger.info(f'[File Cleanup] Successfully removed file: {file_path}')
            else:
                self.logger.warning(f'[File Cleanup] File does not exist: {file_path}')

        except OSError as e:
            self.logger.error(f'[File Cleanup] Error removing file: {e}')

            if e.errno == 13:
                self.logger.warning(f'[File Cleanup] File is in use by another program. Aborting deletion.')
            
    def _change_calendar_date(self, transfer_datetime: datetime) -> None:
       
        try:
            WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.CLASS_NAME, 'flatpickr-calendar')))
        except:
            print('Calendar not found.', flush=True)

        print('Checking year, month, and days.', flush=True)
        calendar_current_month      = self.driver.find_element(By.CLASS_NAME, 'cur-month')
        calendar_current_year_input = self.driver.find_element(By.CLASS_NAME, 'cur-year')

        calendar_current_year_input.send_keys(transfer_datetime.year)

        current_month_value = self._get_month_value(calendar_current_month.text)
        # desired_month_value = self._get_month_value(transfer_datetime.month)

        # print(current_month_value, flush=True)
        # print(desired_month_value, flush=True)
        if current_month_value != transfer_datetime.month:
            print('Changing month', flush=True)
            if current_month_value > transfer_datetime.month:
                # Click back
                for _ in range(current_month_value - transfer_datetime.month + 1):
                    previous_month_button = self.driver.find_element(By.CLASS_NAME, 'flatpickr-prev-month')
                    previous_month_button.click()
                    time.wait(1)
            if current_month_value < transfer_datetime.month:
                for _ in range(transfer_datetime.month - current_month_value + 1):
                    next_month_button = self.driver.find_element(By.CLASS_NAME, 'flatpickr-next-month')
                    next_month_button.click()
                    time.wait(1)

        calendar = self.driver.find_element(By.CLASS_NAME, 'dayContainer')
        today = calendar.find_element(By.XPATH, f'.//span[@class="flatpickr-day "][text()="{transfer_datetime.day}"]')
        today.click()
        time.sleep(2)

        return
    
    def _get_month_value(self, month: str) -> int:
        print(month, flush=True)
        months = {
            'January': 1,
            'February': 2,
            'March': 3,
            'April': 4,
            'May': 5, 
            'June': 6, 
            'July': 7, 
            'August': 8, 
            'September': 9, 
            'October': 10, 
            'November': 11, 
            'December': 12
        }

        if month not in months: return None
        return months[month]