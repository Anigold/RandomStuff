from io import BytesIO
from pathlib import Path
from typing import Any, Dict
from openpyxl import Workbook, load_workbook

from backend.core.interfaces.formatter import BaseFormatter


class ExcelFormatter(BaseFormatter[Dict[str, Any]]):
    """
    Excel formatter supporting either:

    1. A single table:
        {
            "headers": [...],
            "rows": [...],
            "metadata": {...}
        }

    2. A workbook of named tables:
        {
            "sheet1": {"headers": [...], "rows": [...], "metadata": {...}},
            "sheet2": {"headers": [...], "rows": [...], "metadata": {...}},
        }
    """

    def format_name(self) -> str:
        return "xlsx"

    def dumps(self, obj: Dict[str, Any], context: dict | None = None) -> bytes:
        wb = Workbook()

        if self._is_single_table(obj):
            ws = wb.active
            ws.title = (context or {}).get("sheet_name", "Sheet1")
            self._write_table_to_sheet(ws, obj)

        elif self._is_multi_table(obj):
            default_ws = wb.active
            wb.remove(default_ws)

            for sheet_name, table in obj.items():
                ws = wb.create_sheet(title=self._safe_sheet_name(sheet_name))
                self._write_table_to_sheet(ws, table)

        else:
            raise ValueError(
                "ExcelFormatter.dumps expected either a single table "
                "({'headers': ..., 'rows': ...}) or a workbook mapping "
                "({'sheet': {'headers': ..., 'rows': ...}, ...})."
            )

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def loads(self, data: bytes, format: str | None = None, context: dict | None = None) -> Dict[str, Any]:
        buf = BytesIO(data)
        wb = load_workbook(buf)

        sheets = {
            ws.title: self._merge_context_into_table(
                self._read_table_from_sheet(ws),
                context=context,
            )
            for ws in wb.worksheets
        }

        if len(sheets) == 1:
            return next(iter(sheets.values()))

        return sheets

    def load_path(self, path: Path, context: dict | None = None) -> Dict[str, Any]:
        wb = load_workbook(path)

        auto_context = {
            "source_path": str(path),
            "filename": path.name,
            "stem": path.stem,
            "suffix": path.suffix.lstrip(".").lower(),
        }

        merged_context = {**auto_context, **(context or {})}

        sheets = {
            ws.title: self._merge_context_into_table(
                self._read_table_from_sheet(ws),
                context=merged_context,
            )
            for ws in wb.worksheets
        }

        if len(sheets) == 1:
            return next(iter(sheets.values()))

        return sheets

    # -------------------------
    # Internal helpers
    # -------------------------

    def _write_table_to_sheet(self, ws, table: Dict[str, Any]) -> None:
        headers = table.get("headers", [])
        rows = table.get("rows", [])
        metadata = table.get("metadata", {})

        if metadata:
            for key, value in metadata.items():
                ws.append([f"@{key}", value])
            ws.append([])

        if headers:
            ws.append(headers)

        for row in rows:
            ws.append(row)

    def _read_table_from_sheet(self, ws) -> Dict[str, Any]:
        raw_rows = [list(row) for row in ws.iter_rows(values_only=True)]

        if not raw_rows:
            return {"headers": [], "rows": [], "metadata": {}}

        metadata: dict[str, Any] = {}
        data_start_idx = 0

        for i, row in enumerate(raw_rows):
            if not row or all(cell is None for cell in row):
                data_start_idx = i + 1
                continue

            first = row[0]
            if isinstance(first, str) and first.startswith("@"):
                metadata[first[1:]] = row[1] if len(row) > 1 else None
                data_start_idx = i + 1
                continue

            break

        table_rows = raw_rows[data_start_idx:]

        if not table_rows:
            return {"headers": [], "rows": [], "metadata": metadata}

        headers = [
            cell if cell is not None else ""
            for cell in table_rows[0]
        ]
        body = table_rows[1:]

        return {
            "headers": headers,
            "rows": body,
            "metadata": metadata,
        }

    def _merge_context_into_table(
        self,
        table: Dict[str, Any],
        context: dict | None = None,
    ) -> Dict[str, Any]:
        existing_metadata = dict(table.get("metadata", {}))
        merged_metadata = {**existing_metadata, **(context or {})}

        return {
            "headers": table.get("headers", []),
            "rows": table.get("rows", []),
            "metadata": merged_metadata,
        }

    def _is_single_table(self, obj: Dict[str, Any]) -> bool:
        return isinstance(obj, dict) and "headers" in obj and "rows" in obj

    def _is_multi_table(self, obj: Dict[str, Any]) -> bool:
        if not isinstance(obj, dict) or not obj:
            return False

        return all(
            isinstance(v, dict) and "headers" in v and "rows" in v
            for v in obj.values()
        )

    def _safe_sheet_name(self, name: str) -> str:
        invalid = ['\\', '/', '*', '?', ':', '[', ']']
        cleaned = str(name)
        for char in invalid:
            cleaned = cleaned.replace(char, "_")
        return cleaned[:31] or "Sheet"