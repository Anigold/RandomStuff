from pathlib import Path
from backend.infra.paths import DOWNLOADS_PATH, ITEM_FILES_DIR
from openpyxl import Workbook, load_workbook
import json
from backend.domain.models.items.item import Item
from backend.core.normalization.ids import IdGenerator
from dataclasses import asdict


CRATABLE_ITEM_LIST_FILE_NAME: str = 'Ithaca Bakery Director - Foodager - Items.xlsx'

CRAFTABLE_ITEM_LIST_PATH: Path = DOWNLOADS_PATH / CRATABLE_ITEM_LIST_FILE_NAME


CURRENT_IDS: set = set()

def _check_id(item_id: str):
    return item_id in CURRENT_IDS

def run_indexing():
    index = {}
    file_list = ITEM_FILES_DIR.glob('*.json')
    for file in file_list:

        with open(file, 'r') as f:
            item_json = json.load(f)
            item_id = item_json['id']
            item_name = item_json['name']

            index[item_name] = item_id

    with open(ITEM_FILES_DIR/'index.json', 'w', encoding='utf-8') as f:
        json.dump(index, f, indent=4)

def run_import():
    last_completed_item = None
    craftable_item_list_workbook = load_workbook(CRAFTABLE_ITEM_LIST_PATH)
    item_list_sheet = craftable_item_list_workbook.active

    for row in item_list_sheet.iter_rows(min_row=2, values_only=True):

        is_active, item_id, is_inventoried, item_name, category, subcategory, upc, item_yield, vendor, notes = row
        # print([is_active, item_id, is_inventoried, item_name, category, subcategory, upc, item_yield, vendor, notes], flush=True)
        # return
        item_id = IdGenerator.generate_unique(entity_type='item', exists=_check_id)
        item_name_sani = ' '.join(item_name.split(' ')[0:-1])
        count_unit = ''.join(item_name.split(' ')[-1])

        item_obj = Item(
            id=item_id, 
            name=item_name_sani,
            category=category, 
            subcategory=subcategory, 
            count_unit=count_unit,
            store_info=None,
            vendor_info=None,
            is_active=is_active,
            is_inventoried=is_inventoried,
            notes=notes,
            aliases=[]
        )

        item_dict = asdict(item_obj)

        filepath = ITEM_FILES_DIR / f'{item_id}.json'
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(item_dict, f, indent=4)

            last_completed_item = item_obj.name
            CURRENT_IDS.add(item_obj.id)
        
            # print([is_active, item_id, is_inventoried, item_name, category, subcategory, upc, item_yield, vendor, notes], flush=True)
        except:
            print(f'{item_name} failed to import.')
            print(f'Last completed item: {last_completed_item}')