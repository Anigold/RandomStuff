from __future__ import annotations
from pathlib import Path
from typing import Optional
from openpyxl import Workbook, load_workbook

from backend.models.order import Order
from backend.models.order_item import OrderItem
from backend.app.files.ports_generic import Serializer

class OrderExcelSerializer(Serializer[Order]):
    '''Excel <-> Order. Minimal; add formatting as needed'''
    
    def preferred_format(self) -> str:
        return "xlsx"

    def dumps(self, obj: Order, format: Optional[str] = None) -> bytes:
        wb = Workbook(); ws = wb.active
        ws.title = "order"
        ws.append(["SKU", "Name", "Qty", "Cost", "Total"])
        for it in obj.items:
            ws.append([it.sku, it.name, it.quantity, it.cost_per, it.total_cost])

        meta = wb.create_sheet("meta")
        meta.append(["store", obj.store])
        meta.append(["vendor", obj.vendor])
        meta.append(["date", obj.date])

        tmp = Path("_tmp_order.xlsx")
        wb.save(tmp)
        data = tmp.read_bytes()
        tmp.unlink(missing_ok=True)
        return data

    def loads(self, data: bytes, format: Optional[str] = None) -> Order:
        tmp = Path("_tmp_read_order.xlsx"); tmp.write_bytes(data)
        order = self.load_path(tmp)
        tmp.unlink(missing_ok=True)
        return order

    def load_path(self, path: Path) -> Order:
        wb = load_workbook(path)
        ws = wb.active
        items: list[OrderItem] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:  # header
                continue
            sku, name, qty, cp, total = row
            items.append(OrderItem(str(sku or ""), str(name or ""), str(qty or 0), str(cp or 0), str(total or 0)))

        store = vendor = date = ""
        if "meta" in wb.sheetnames:
            meta = wb["meta"]
            meta_map = {str(r[0].value or ""): str(r[1].value or "") for r in meta.iter_rows(values_only=True)}
            store  = meta_map.get("store", "")
            vendor = meta_map.get("vendor", "")
            date   = meta_map.get("date", "")

        return Order(store=store, vendor=vendor, date=date, items=items)
